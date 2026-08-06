from flask import Blueprint, render_template, render_template_string, request, redirect, url_for, session, jsonify, flash
import json
from functools import wraps
from datetime import date, datetime, timedelta
from database import get_db, get_setting, set_setting

bp = Blueprint("owner", __name__, url_prefix="/owner")


ORDERS_PAGE = """{% extends 'base.html' %}
{% block title %}Orders & Progress — Owner{% endblock %}
{% block content %}
<div class="page-header">
  <div><h1>📋 Orders & Progress</h1><div class="header-sub">{{ total }} orders total</div></div>
  <div style="display:flex;gap:8px;align-items:center;">
    <a href="/owner/dashboard" class="btn btn-ghost btn-sm">← Dashboard</a>
    <button class="menu-toggle" onclick="openSidebar()">☰</button>
  </div>
</div>
{% with messages = get_flashed_messages(with_categories=true) %}{% if messages %}<div style="padding:8px 24px 0;">{% for cat,msg in messages %}<div style="background:{{'#d1fae5' if cat=='success' else '#fee2e2'}};color:{{'#065f46' if cat=='success' else '#dc2626'}};padding:10px 16px;border-radius:10px;font-weight:700;font-size:13px;margin-bottom:4px;">{{ msg }}</div>{% endfor %}</div>{% endif %}{% endwith %}

<div class="page-body" style="padding:12px 24px;">

  <!-- ═══ TAB SWITCHER ═══ -->
  <div style="display:flex;gap:0;border-bottom:2px solid var(--border);margin-bottom:16px;">
    <button id="tab-orders-btn" onclick="switchTab('orders')"
      style="padding:10px 24px;font-size:14px;font-weight:800;border:none;background:none;border-bottom:3px solid var(--accent);color:var(--accent);cursor:pointer;margin-bottom:-2px;">
      📋 Order Management
    </button>
    <button id="tab-wp-btn" onclick="switchTab('workprogress')"
      style="padding:10px 24px;font-size:14px;font-weight:800;border:none;background:none;border-bottom:3px solid transparent;color:var(--text-muted);cursor:pointer;margin-bottom:-2px;">
      📊 Work Progress
    </button>
  </div>

  <!-- Search + Filter (always visible) -->
  <div id="orders-filter-bar" style="display:flex;gap:10px;flex-wrap:wrap;margin-bottom:14px;align-items:center;">
    <input type="text" id="srch" placeholder="Search #code, name, mobile..." oninput="filterOrders()"
      style="flex:1;min-width:200px;padding:10px 16px;font-size:14px;border:2px solid var(--border);border-radius:12px;outline:none;">
    <span id="dues-badge" style="display:none;background:#fef3c7;color:#b45309;border:1.5px solid #fde68a;border-radius:10px;padding:6px 14px;font-size:12px;font-weight:800;">💰 Pending dues only</span>
    <div style="display:flex;gap:6px;flex-wrap:wrap;">
      {% for key,label in [('all','All'),('pending','Pending'),('ready','Ready'),('delivered','Delivered'),('cancelled','Cancelled')] %}
      <button class="ftab" data-key="{{ key }}" onclick="setTab('{{ key }}',this)"
        style="padding:7px 14px;border-radius:10px;border:2px solid {% if key=='all' %}var(--accent){% else %}var(--border){% endif %};background:{% if key=='all' %}var(--accent){% else %}#fff{% endif %};color:{% if key=='all' %}#fff{% else %}var(--text-muted){% endif %};font-size:12px;font-weight:800;cursor:pointer;">{{ label }}</button>
      {% endfor %}
    </div>
  </div>

  <!-- ═══ ORDERS TAB ═══ -->
  <div id="tab-orders">
    <div style="overflow-x:auto;">
      <table style="width:100%;border-collapse:collapse;font-size:13px;">
        <thead><tr style="background:#f8fafc;border-bottom:2px solid var(--border);">
          <th style="padding:10px 14px;text-align:left;font-size:10px;font-weight:700;text-transform:uppercase;color:var(--text-muted);">Order</th>
          <th style="padding:10px 14px;text-align:left;font-size:10px;font-weight:700;text-transform:uppercase;color:var(--text-muted);">Customer</th>
          <th style="padding:10px 14px;text-align:left;font-size:10px;font-weight:700;text-transform:uppercase;color:var(--text-muted);">Garments</th>
          <th style="padding:10px 14px;text-align:left;font-size:10px;font-weight:700;text-transform:uppercase;color:var(--text-muted);">Dates</th>
          <th style="padding:10px 14px;text-align:right;font-size:10px;font-weight:700;text-transform:uppercase;color:var(--text-muted);">Amount</th>
          <th style="padding:10px 14px;text-align:center;font-size:10px;font-weight:700;text-transform:uppercase;color:var(--text-muted);">Status</th>
          <th style="padding:10px 14px;text-align:center;font-size:10px;font-weight:700;text-transform:uppercase;color:var(--text-muted);">Actions</th>
        </tr></thead>
        <tbody id="tbody">
          {% for o in orders %}
          <tr class="orow" data-status="{{ o.status }}" data-remaining="{{ o.remaining }}"
            data-s="{{ o.order_code }} {{ o.cname|lower }} {{ o.mobile }} {{ o.garments|lower }}"
            style="border-bottom:1px solid var(--border);cursor:pointer;"
            onclick="toggleDetail('{{ o.order_code }}')"
            onmouseover="this.style.background='#f8fafc'" onmouseout="this.style.background=''">
            <td style="padding:12px 14px;">
              <div style="display:flex;align-items:center;gap:6px;">
                <span id="arrow-{{ o.order_code }}" style="font-size:10px;color:var(--text-muted);transition:transform 0.2s;display:inline-block;">▶</span>
                <div style="font-size:15px;font-weight:900;color:var(--accent);">#{{ o.display_code }}</div>
                {% if o.entry_code %}<span style="background:#f1f5f9;color:#64748b;font-size:10px;font-weight:700;padding:1px 6px;border-radius:6px;">(#{{ o.entry_code }})</span>{% endif %}
                {% if o.is_urgent %}<span style="background:#fee2e2;color:#dc2626;font-size:9px;font-weight:800;padding:1px 6px;border-radius:4px;">🔥</span>{% endif %}
              </div>
              {% if o.note %}<div style="font-size:11px;color:var(--text-muted);font-style:italic;padding-left:18px;">📝 {{ o.note[:30] }}</div>{% endif %}
            </td>
            <td style="padding:12px 14px;"><div style="font-weight:700;">{{ o.cname }}</div>{% if o.mobile %}<div style="font-size:11px;color:var(--text-muted);">{{ o.mobile }}</div>{% endif %}</td>
            <td style="padding:12px 14px;color:var(--text-secondary);max-width:140px;">{{ o.garments }}</td>
            <td style="padding:12px 14px;"><div style="font-size:11px;color:var(--text-muted);">Order: {{ o.order_date }}</div><div style="font-size:11px;color:var(--text-muted);">Delivery: <strong>{{ o.delivery_date }}</strong></div></td>
            <td style="padding:12px 14px;text-align:right;"><div style="font-weight:800;">₹{{ o.payable|int }}</div>{% if o.remaining > 0 %}<div style="font-size:11px;color:var(--danger);font-weight:700;">Due ₹{{ o.remaining|int }}</div>{% else %}<div style="font-size:11px;color:var(--success);font-weight:700;">✓ Paid</div>{% endif %}</td>
            <td style="padding:12px 14px;text-align:center;"><span style="font-size:11px;padding:3px 10px;border-radius:8px;font-weight:800;background:{% if o.status=='delivered' %}#d1fae5;color:#065f46{% elif o.status=='ready' %}#ede9fe;color:#6d28d9{% elif o.status=='cancelled' %}#fee2e2;color:#dc2626{% else %}#dbeafe;color:#1e40af{% endif %};">{{ o.status|upper }}</span></td>
            <td style="padding:12px 14px;text-align:center;" onclick="event.stopPropagation()">
              <div style="display:flex;gap:4px;justify-content:center;flex-wrap:wrap;">
                <a href="/owner/orders/edit/{{ o.order_code }}" style="background:#f0fdf4;color:#16a34a;border-radius:7px;padding:4px 9px;font-size:11px;font-weight:700;text-decoration:none;display:inline-block;">✏️</a>
                <button onclick="window.open('/print-slip/{{ o.order_code }}','_blank')" style="background:var(--accent-light);color:var(--accent);border:none;border-radius:7px;padding:4px 9px;font-size:11px;font-weight:700;cursor:pointer;">🖨️</button>
                {% if o.status != 'delivered' and o.status != 'cancelled' %}
                <form action="/owner/orders/cancel/{{ o.order_code }}" method="POST" style="margin:0;" onsubmit="return confirm('Cancel #{{ o.order_code }}?')"><button type="submit" style="background:var(--danger-light);color:var(--danger);border:none;border-radius:7px;padding:4px 9px;font-size:11px;font-weight:700;cursor:pointer;">✕</button></form>
                {% endif %}
                <form action="/owner/orders/delete/{{ o.order_code }}" method="POST" style="margin:0;" onsubmit="return confirm('DELETE #{{ o.order_code }}?')"><button type="submit" style="background:#1f2937;color:#fff;border:none;border-radius:7px;padding:4px 9px;font-size:11px;font-weight:700;cursor:pointer;">🗑️</button></form>
              </div>
            </td>
          </tr>
          <!-- Expandable Detail Row -->
          <tr id="detail-{{ o.order_code }}" style="display:none;background:#f8faff;border-bottom:2px solid var(--accent);">
            <td colspan="7" style="padding:0;">
              <div id="detail-content-{{ o.order_code }}" style="padding:16px 20px;">
                <div style="color:var(--text-muted);font-size:13px;text-align:center;padding:16px;">Loading...</div>
              </div>
            </td>
          </tr>
          {% else %}
          <tr><td colspan="7" style="padding:40px;text-align:center;color:var(--text-muted);">No orders yet</td></tr>
          {% endfor %}
        </tbody>
      </table>
    </div>
  </div>

  <!-- ═══ WORK PROGRESS TAB ═══ -->
  <div id="tab-workprogress" style="display:none;">
    <div style="display:flex;gap:8px;flex-wrap:wrap;margin-bottom:14px;">
      {% for key,label in [('all','All Active'),('naap','📐 नाप Pending'),('cut','✂️ कटाई Pending'),('stitch','🪡 सिलाई Pending'),('done','✅ All Done')] %}
      <button class="wpf" data-key="{{ key }}" onclick="setWpFilter('{{ key }}',this)"
        style="padding:7px 14px;border-radius:10px;border:2px solid {% if key=='all' %}var(--accent){% else %}var(--border){% endif %};background:{% if key=='all' %}var(--accent){% else %}#fff{% endif %};color:{% if key=='all' %}#fff{% else %}var(--text-muted){% endif %};font-size:12px;font-weight:800;cursor:pointer;">{{ label }}</button>
      {% endfor %}
    </div>
    <div style="overflow-x:auto;-webkit-overflow-scrolling:touch;">
    <table style="width:100%;border-collapse:collapse;font-size:13px;">
      <thead><tr style="background:#f8fafc;border-bottom:2px solid var(--border);">
        <th style="padding:10px 14px;text-align:left;font-size:10px;font-weight:700;text-transform:uppercase;color:var(--text-muted);">Order</th>
        <th style="padding:10px 14px;text-align:left;font-size:10px;font-weight:700;text-transform:uppercase;color:var(--text-muted);">Customer</th>
        <th style="padding:10px 14px;text-align:left;font-size:10px;font-weight:700;text-transform:uppercase;color:var(--text-muted);">Garments</th>
        <th style="padding:10px 14px;text-align:left;font-size:10px;font-weight:700;text-transform:uppercase;color:var(--text-muted);">Delivery</th>
        <th style="padding:10px 14px;text-align:left;font-size:10px;font-weight:700;text-transform:uppercase;color:var(--text-muted);">Progress</th>
        <th style="padding:10px 14px;text-align:center;font-size:10px;font-weight:700;text-transform:uppercase;color:var(--text-muted);">Pending</th>
      </tr></thead>
      <tbody id="wp-tbody">
        {% for o in wp_orders %}
        {% set all_done = o.naap_pct >= 100 and o.cut_pct >= 100 and o.stitch_pct >= 100 %}
        <tr class="wrow"
          data-naap="{{ 'done' if o.naap_pct>=100 else 'pending' }}"
          data-cut="{{ 'done' if o.cut_pct>=100 else 'pending' }}"
          data-stitch="{{ 'done' if o.stitch_pct>=100 else 'pending' }}"
          data-alldone="{{ 'yes' if all_done else 'no' }}"
          style="border-bottom:1px solid var(--border);"
          onmouseover="this.style.background='#fafbff'" onmouseout="this.style.background=''">
          <td style="padding:12px 14px;"><div style="font-size:15px;font-weight:900;color:var(--accent);">#{{ o.display_code }}</div>{% if o.entry_code %}<span style="background:#f1f5f9;color:#64748b;font-size:9px;font-weight:700;padding:1px 6px;border-radius:5px;">(#{{ o.entry_code }})</span>{% endif %}{% if o.is_urgent %}<span style="background:#fee2e2;color:#dc2626;font-size:9px;font-weight:800;padding:1px 6px;border-radius:4px;">🔥</span>{% endif %}</td>
          <td style="padding:12px 14px;"><div style="font-weight:700;">{{ o.cname }}</div><div style="font-size:11px;color:var(--text-muted);">{{ o.mobile }}</div></td>
          <td style="padding:12px 14px;color:var(--text-secondary);max-width:130px;font-size:12px;">{{ o.garments }}</td>
          <td style="padding:12px 14px;"><div style="font-size:13px;font-weight:700;">{{ o.delivery_date }}</div><div style="font-size:11px;color:var(--text-muted);">{{ o.status|upper }}</div></td>
          <td style="padding:12px 14px;">
            <div style="display:flex;justify-content:space-between;margin-bottom:4px;">
              <span style="font-size:9px;font-weight:800;color:{% if o.naap_pct>=100 %}#4f46e5{% else %}#9ca3af{% endif %};">नाप {% if o.naap_pct>=100 %}✓{% else %}{{ o.naap_pct }}%{% endif %}</span>
              <span style="font-size:9px;font-weight:800;color:{% if o.cut_pct>=100 %}#ea580c{% else %}#9ca3af{% endif %};">कटाई {% if o.cut_pct>=100 %}✓{% else %}{{ o.cut_pct }}%{% endif %}</span>
              <span style="font-size:9px;font-weight:800;color:{% if o.stitch_pct>=100 %}#16a34a{% else %}#9ca3af{% endif %};">सिलाई {% if o.stitch_pct>=100 %}✓{% else %}{{ o.stitch_pct }}%{% endif %}</span>
            </div>
            <div style="display:flex;gap:2px;height:10px;">
              <div style="flex:1;background:#e5e7eb;border-radius:4px;overflow:hidden;"><div style="height:100%;background:#4f46e5;width:{{ o.naap_pct }}%;"></div></div>
              <div style="flex:1;background:#e5e7eb;border-radius:4px;overflow:hidden;"><div style="height:100%;background:#ea580c;width:{{ o.cut_pct }}%;"></div></div>
              <div style="flex:1;background:#e5e7eb;border-radius:4px;overflow:hidden;"><div style="height:100%;background:#16a34a;width:{{ o.stitch_pct }}%;"></div></div>
            </div>
          </td>
          <td style="padding:12px 14px;text-align:center;">
            {% if all_done %}<span style="background:#d1fae5;color:#065f46;border-radius:8px;padding:4px 10px;font-size:11px;font-weight:800;">✅ All Done</span>
            {% else %}<div style="display:flex;flex-direction:column;gap:3px;align-items:center;">
              {% if o.naap_pct < 100 %}<span style="background:#eef2ff;color:#4f46e5;border-radius:6px;padding:2px 8px;font-size:10px;font-weight:800;">📐 नाप</span>{% endif %}
              {% if o.cut_pct < 100 %}<span style="background:#fff7ed;color:#ea580c;border-radius:6px;padding:2px 8px;font-size:10px;font-weight:800;">✂️ कटाई</span>{% endif %}
              {% if o.stitch_pct < 100 %}<span style="background:#f0fdf4;color:#16a34a;border-radius:6px;padding:2px 8px;font-size:10px;font-weight:800;">🪡 सिलाई</span>{% endif %}
            </div>{% endif %}
          </td>
        </tr>
        {% else %}
        <tr><td colspan="6" style="padding:40px;text-align:center;color:var(--text-muted);">All orders delivered! 🎉</td></tr>
        {% endfor %}
      </tbody>
    </table>
    </div>
  </div>
</div>

<!-- Image Overlay -->
<div id="img-overlay-orders" onclick="this.style.display='none'" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,0.92);z-index:9999;align-items:center;justify-content:center;">
  <img id="ov-img-orders" style="max-width:90%;max-height:90%;border-radius:12px;object-fit:contain;">
  <button onclick="document.getElementById('img-overlay-orders').style.display='none';event.stopPropagation();" style="position:absolute;top:16px;right:16px;background:rgba(255,255,255,0.15);border:none;color:#fff;width:44px;height:44px;border-radius:50%;font-size:22px;cursor:pointer;">✕</button>
</div>

{% endblock %}
{% block extra_js %}<script>
// ─── STATE ───────────────────────────────────────
var activeTab = "all";
var activeFilter = "{{filter_mode}}";
var expandedCode = null;

// ─── ON LOAD ──────────────────────────────────────
window.addEventListener("DOMContentLoaded", function() {
  if (activeFilter === "dues") {
    var badge = document.getElementById("dues-badge");
    if (badge) badge.style.display = "inline-block";
    filterOrders();
  }
});

// ─── TAB SWITCHING (Orders / Work Progress) ───────
function switchTab(tab) {
  var ordersDiv = document.getElementById("tab-orders");
  var wpDiv     = document.getElementById("tab-workprogress");
  var filterBar = document.getElementById("orders-filter-bar");
  var ordBtn    = document.getElementById("tab-orders-btn");
  var wpBtn     = document.getElementById("tab-wp-btn");
  var isOrders  = (tab === "orders");

  if (ordersDiv) ordersDiv.style.display = isOrders ? "block" : "none";
  if (wpDiv)     wpDiv.style.display     = isOrders ? "none"  : "block";
  if (filterBar) filterBar.style.display = isOrders ? "flex"  : "none";

  ordBtn.style.borderBottomColor = isOrders ? "var(--accent)" : "transparent";
  ordBtn.style.color = isOrders ? "var(--accent)" : "var(--text-muted)";
  ordBtn.style.fontWeight = "800";

  wpBtn.style.borderBottomColor = isOrders ? "transparent" : "#ea580c";
  wpBtn.style.color = isOrders ? "var(--text-muted)" : "#ea580c";
  wpBtn.style.fontWeight = "800";
}

// ─── STATUS FILTER TABS ───────────────────────────
function setTab(k, btn) {
  activeTab = k;
  document.querySelectorAll(".ftab").forEach(function(b) {
    b.style.background = "#fff";
    b.style.color = "var(--text-muted)";
    b.style.borderColor = "var(--border)";
  });
  btn.style.background = "var(--accent)";
  btn.style.color = "#fff";
  btn.style.borderColor = "var(--accent)";
  filterOrders();
}

// ─── URL STATE SYNC ────────────────────────────────
// Keeps search text + active tab in the URL (?q=...&tab=...) so that
// after editing an order and saving, we can return to this exact
// search/filter state instead of a blank list (needed for bulk edits
// like going through order codes 3720-3730 one by one).
function syncOrdersUrlState() {
  var srch = document.getElementById("srch");
  var q = srch ? srch.value : "";
  var params = new URLSearchParams(window.location.search);
  if (q) params.set("q", q); else params.delete("q");
  if (activeTab && activeTab !== "all") params.set("tab", activeTab); else params.delete("tab");
  var newUrl = window.location.pathname + (params.toString() ? "?" + params.toString() : "");
  history.replaceState(null, "", newUrl);
}

function restoreOrdersUrlState() {
  var params = new URLSearchParams(window.location.search);
  var q   = params.get("q");
  var tab = params.get("tab");
  var srch = document.getElementById("srch");
  if (q && srch) srch.value = q;
  if (tab) {
    activeTab = tab;
    var btn = document.querySelector('.ftab[data-key="' + tab + '"]');
    if (btn) {
      document.querySelectorAll(".ftab").forEach(function(b) {
        b.style.background = "#fff"; b.style.color = "var(--text-muted)"; b.style.borderColor = "var(--border)";
      });
      btn.style.background = "var(--accent)"; btn.style.color = "#fff"; btn.style.borderColor = "var(--accent)";
    }
  }
  if (q || tab) filterOrders();
}
window.addEventListener("DOMContentLoaded", restoreOrdersUrlState);

// ─── SEARCH + FILTER ──────────────────────────────
function filterOrders() {
  var srch = document.getElementById("srch");
  var q = srch ? srch.value.toLowerCase().trim() : "";
  document.querySelectorAll("#tbody .orow").forEach(function(r) {
    var hasDue   = parseFloat(r.dataset.remaining || 0) > 0;
    var matchQ   = !q || (r.dataset.s || "").includes(q);
    var matchF   = activeTab === "all" || r.dataset.status === activeTab;
    var matchDues = activeFilter !== "dues" || (hasDue && r.dataset.status !== "delivered" && r.dataset.status !== "cancelled");
    var show = matchQ && matchF && matchDues;
    r.style.display = show ? "" : "none";
    // hide detail row if parent hidden
    var arrow = r.querySelector("[id^='arrow-']");
    if (arrow) {
      var code = arrow.id.replace("arrow-", "");
      var dr = document.getElementById("detail-" + code);
      if (dr && !show) { dr.style.display = "none"; }
    }
  });
  syncOrdersUrlState();
}

// ─── EXPANDABLE ROWS ──────────────────────────────
function toggleDetail(code) {
  var dr    = document.getElementById("detail-" + code);
  var arrow = document.getElementById("arrow-" + code);
  if (!dr || !arrow) return;

  if (expandedCode && expandedCode !== code) {
    var prev      = document.getElementById("detail-" + expandedCode);
    var prevArrow = document.getElementById("arrow-" + expandedCode);
    if (prev)      prev.style.display = "none";
    if (prevArrow) prevArrow.style.transform = "rotate(0deg)";
    expandedCode = null;
  }

  if (dr.style.display === "none" || dr.style.display === "") {
    dr.style.display = "table-row";
    arrow.style.transform = "rotate(90deg)";
    expandedCode = code;
    loadDetail(code);
  } else {
    dr.style.display = "none";
    arrow.style.transform = "rotate(0deg)";
    expandedCode = null;
  }
}

function loadDetail(code) {
  var container = document.getElementById("detail-content-" + code);
  if (!container || container.dataset.loaded === "1") return;
  container.innerHTML = '<div style="color:var(--text-muted);font-size:13px;text-align:center;padding:16px;">Loading...</div>';

  fetch("/owner/api/order-detail/" + code)
    .then(function(r) { return r.json(); })
    .then(function(d) {
      if (d.error) { container.innerHTML = '<div style="color:red;padding:12px;">Error loading</div>'; return; }
      var o = d.order;
      var html = '<div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:16px;">';

      // Col 1: Customer + Payment
      html += '<div>';
      html += '<div style="font-size:10px;font-weight:700;text-transform:uppercase;color:var(--text-muted);margin-bottom:8px;">👤 Customer & Payment</div>';
      html += '<div style="font-size:15px;font-weight:800;">' + o.cname + '</div>';
      html += '<div style="font-size:12px;color:var(--text-muted);">📱 ' + o.mobile + '</div>';
      html += '<div style="margin-top:10px;background:#f8fafc;border-radius:8px;padding:10px;">';
      html += '<div style="display:flex;justify-content:space-between;font-size:12px;padding:3px 0;"><span style="color:var(--text-muted);">Total Bill</span><span style="font-weight:700;">₹' + Math.round(o.payable) + '</span></div>';
      html += '<div style="display:flex;justify-content:space-between;font-size:12px;padding:3px 0;"><span style="color:var(--text-muted);">Advance</span><span style="font-weight:700;color:#16a34a;">₹' + Math.round(o.advance) + '</span></div>';
      html += '<div style="display:flex;justify-content:space-between;font-size:13px;padding:4px 0;border-top:1px solid var(--border);margin-top:3px;"><span style="font-weight:800;color:' + (o.remaining > 0 ? "#dc2626" : "#16a34a") + ';">' + (o.remaining > 0 ? "Due" : "✓ Paid") + '</span><span style="font-weight:900;color:' + (o.remaining > 0 ? "#dc2626" : "#16a34a") + ';">₹' + Math.round(o.remaining) + '</span></div>';
      html += '</div></div>';

      // Col 2: Garments
      html += '<div>';
      html += '<div style="font-size:10px;font-weight:700;text-transform:uppercase;color:var(--text-muted);margin-bottom:8px;">👕 Garments & Measurements</div>';
      d.garments.forEach(function(g) {
        html += '<div style="background:#fff;border:1px solid var(--border);border-radius:8px;padding:10px;margin-bottom:8px;">';
        html += '<div style="display:flex;justify-content:space-between;margin-bottom:6px;"><span style="font-weight:800;">' + g.type + ' x' + g.qty + '</span><span style="color:var(--accent);font-weight:700;">₹' + Math.round(g.amount) + '</span></div>';
        if (g.measurements) {
          var mkeys = Object.keys(g.measurements);
          if (mkeys.length > 0) {
            html += '<div style="display:flex;flex-wrap:wrap;gap:4px;">';
            mkeys.forEach(function(k) { if (g.measurements[k]) html += '<span style="background:#f1f5f9;border-radius:5px;padding:2px 8px;font-size:11px;"><span style="color:var(--text-muted);">' + k + ':</span> <b>' + g.measurements[k] + '</b></span>'; });
            html += '</div>';
          }
        }
        html += '</div>';
      });
      html += '</div>';

      // Col 3: Progress + Photos
      html += '<div>';
      html += '<div style="font-size:10px;font-weight:700;text-transform:uppercase;color:var(--text-muted);margin-bottom:8px;">📊 Progress & Photos</div>';
      html += '<div style="display:flex;justify-content:space-between;margin-bottom:4px;">';
      html += '<span style="font-size:10px;font-weight:800;color:' + (o.naap_pct>=100?"#4f46e5":"#9ca3af") + ';">नाप ' + (o.naap_pct>=100?"✓":o.naap_pct+"%") + '</span>';
      html += '<span style="font-size:10px;font-weight:800;color:' + (o.cut_pct>=100?"#ea580c":"#9ca3af") + ';">कटाई ' + (o.cut_pct>=100?"✓":o.cut_pct+"%") + '</span>';
      html += '<span style="font-size:10px;font-weight:800;color:' + (o.stitch_pct>=100?"#16a34a":"#9ca3af") + ';">सिलाई ' + (o.stitch_pct>=100?"✓":o.stitch_pct+"%") + '</span>';
      html += '</div>';
      html += '<div style="display:flex;gap:2px;height:10px;margin-bottom:12px;">';
      html += '<div style="flex:1;background:#e5e7eb;border-radius:4px;overflow:hidden;"><div style="height:100%;background:#4f46e5;width:' + o.naap_pct + '%;"></div></div>';
      html += '<div style="flex:1;background:#e5e7eb;border-radius:4px;overflow:hidden;"><div style="height:100%;background:#ea580c;width:' + o.cut_pct + '%;"></div></div>';
      html += '<div style="flex:1;background:#e5e7eb;border-radius:4px;overflow:hidden;"><div style="height:100%;background:#16a34a;width:' + o.stitch_pct + '%;"></div></div>';
      html += '</div>';
      if (d.images && d.images.length > 0) {
        html += '<div style="display:flex;flex-wrap:wrap;gap:6px;">';
        d.images.forEach(function(src) {
          html += '<img src="' + src + '" onclick="openImgOverlay(this.src)" style="width:70px;height:70px;object-fit:cover;border-radius:8px;border:2px solid var(--border);cursor:zoom-in;">';
        });
        html += '</div>';
      } else {
        html += '<div style="font-size:12px;color:var(--text-muted);">No photos</div>';
      }
      html += '</div></div>';

      container.innerHTML = html;
      container.dataset.loaded = "1";
    })
    .catch(function() { container.innerHTML = '<div style="color:red;padding:12px;">Could not load details.</div>'; });
}

function openImgOverlay(src) {
  var ov = document.getElementById("img-overlay-orders");
  if (ov) { document.getElementById("ov-img-orders").src = src; ov.style.display = "flex"; }
}

// ─── WORK PROGRESS FILTER ─────────────────────────
function setWpFilter(f, btn) {
  document.querySelectorAll(".wpf").forEach(function(b) {
    b.style.background = "#fff"; b.style.color = "var(--text-muted)"; b.style.borderColor = "var(--border)";
  });
  btn.style.background = "var(--accent)"; btn.style.color = "#fff"; btn.style.borderColor = "var(--accent)";
  document.querySelectorAll(".wrow").forEach(function(r) {
    var show = f === "all"
      || (f === "naap"   && r.dataset.naap   === "pending")
      || (f === "cut"    && r.dataset.cut    === "pending")
      || (f === "stitch" && r.dataset.stitch === "pending")
      || (f === "done"   && r.dataset.alldone === "yes");
    r.style.display = show ? "" : "none";
  });
}

// ─── SESSION TIMEOUT ──────────────────────────────
const SECS = 5 * 60; let last = Date.now();
["click","keydown","mousemove","touchstart"].forEach(function(ev) {
  document.addEventListener(ev, function() { last = Date.now(); }, {passive:true});
});
setInterval(function() {
  if (Math.floor((Date.now() - last) / 1000) >= SECS) window.location.href = "/owner/login?expired=1";
}, 5000);
window.addEventListener("pageshow", function(e) {
  if (e.persisted) { fetch("/owner/logout", {method:"POST",keepalive:true}).finally(function() { window.location.href = "/owner/login"; }); }
});
</script>{% endblock %}
"""


CUSTOMERS_PAGE = """{% extends 'base.html' %}
{% block title %}Customers — Owner{% endblock %}
{% block content %}
<div class="page-header">
  <div><h1>👥 Customers</h1><div class="header-sub">{{ total }} total</div></div>
  <div style="display:flex;gap:8px;align-items:center;"><a href="/owner/export/orders" class="btn btn-ghost btn-sm" style="background:#d1fae5;color:#065f46;">📥 Export</a><a href="/owner/dashboard" class="btn btn-ghost btn-sm">← Dashboard</a><button class="menu-toggle" onclick="openSidebar()">☰</button></div>
</div>
<div class="page-body" style="padding:16px 24px;">
  <input type="text" id="srch" placeholder="Search name, mobile, order code..." oninput="filterRows()" style="padding:11px 16px;font-size:14px;border:2px solid var(--border);border-radius:12px;width:100%;max-width:400px;margin-bottom:16px;">
  <div style="overflow-x:auto;"><table style="width:100%;border-collapse:collapse;font-size:13px;">
    <thead><tr style="background:#f8fafc;border-bottom:2px solid var(--border);"><th style="padding:10px 16px;text-align:left;font-size:10px;font-weight:700;text-transform:uppercase;color:var(--text-muted);">Customer</th><th style="padding:10px 16px;text-align:left;font-size:10px;font-weight:700;text-transform:uppercase;color:var(--text-muted);">Mobile</th><th style="padding:10px 16px;text-align:left;font-size:10px;font-weight:700;text-transform:uppercase;color:var(--text-muted);">Address</th><th style="padding:10px 16px;text-align:center;font-size:10px;font-weight:700;text-transform:uppercase;color:var(--text-muted);">Orders</th><th style="padding:10px 16px;text-align:left;font-size:10px;font-weight:700;text-transform:uppercase;color:var(--text-muted);">Codes</th><th style="padding:10px 16px;text-align:right;font-size:10px;font-weight:700;text-transform:uppercase;color:var(--text-muted);">Billed</th><th style="padding:10px 16px;text-align:right;font-size:10px;font-weight:700;text-transform:uppercase;color:var(--text-muted);">Due</th><th style="padding:10px 16px;text-align:right;font-size:10px;font-weight:700;text-transform:uppercase;color:var(--text-muted);">Last Order</th></tr></thead>
    <tbody id="tbody">{% for c in customers %}<tr class="crow" data-s="{{ c.name|lower }} {{ c.mobile }} {{ c.address|lower }} {{ c.order_codes }}" onclick="window.location='/owner/customers/{{ c.id }}'" style="border-bottom:1px solid var(--border);cursor:pointer;" onmouseover="this.style.background='#f8fafc'" onmouseout="this.style.background=''"><td style="padding:12px 16px;"><div style="font-size:14px;font-weight:800;">{{ c.name }}</div></td><td style="padding:12px 16px;color:var(--text-muted);">{{ c.mobile or '—' }}</td><td style="padding:12px 16px;color:var(--text-muted);max-width:140px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;">{{ c.address or '—' }}</td><td style="padding:12px 16px;text-align:center;font-weight:800;color:var(--accent);">{{ c.order_count }}</td><td style="padding:12px 16px;color:var(--text-muted);font-size:11px;">{{ c.order_codes or '—' }}</td><td style="padding:12px 16px;text-align:right;font-weight:700;">₹{{ c.total_billed|int }}</td><td style="padding:12px 16px;text-align:right;font-weight:700;color:{% if c.total_due > 0 %}var(--danger){% else %}var(--success){% endif %};">{% if c.total_due > 0 %}₹{{ c.total_due|int }}{% else %}✓ Paid{% endif %}</td><td style="padding:12px 16px;text-align:right;color:var(--text-muted);font-size:12px;">{{ c.last_order_date or '—' }}</td></tr>{% else %}<tr><td colspan="8" style="padding:40px;text-align:center;color:var(--text-muted);">No customers yet</td></tr>{% endfor %}</tbody>
  </table></div>
</div>
{% endblock %}
{% block extra_js %}<script>
function filterRows(){
  var q=document.getElementById("srch").value.toLowerCase().trim();
  document.querySelectorAll(".crow").forEach(function(r){r.style.display=(!q||r.dataset.s.includes(q))?"":"none";});
  var params = new URLSearchParams(window.location.search);
  if (q) params.set("q", q); else params.delete("q");
  var newUrl = window.location.pathname + (params.toString() ? "?" + params.toString() : "");
  history.replaceState(null, "", newUrl);
}
window.addEventListener("DOMContentLoaded", function(){
  var params = new URLSearchParams(window.location.search);
  var q = params.get("q");
  if (q) { document.getElementById("srch").value = q; filterRows(); }
});
const SECS=5*60;let last=Date.now();["click","keydown","mousemove","touchstart"].forEach(ev=>document.addEventListener(ev,()=>{last=Date.now();},{passive:true}));setInterval(()=>{if(Math.floor((Date.now()-last)/1000)>=SECS)window.location.href="/owner/login?expired=1";},5000);window.addEventListener("pageshow",function(e){if(e.persisted){fetch("/owner/logout",{method:"POST",keepalive:true}).finally(()=>{window.location.href="/owner/login";})}});
</script>{% endblock %}
"""

WORK_PROGRESS_PAGE = """{% extends 'base.html' %}
{% block title %}Work Progress — Owner{% endblock %}
{% block content %}
<div class="page-header">
  <div><h1>📊 Work Progress</h1><div class="header-sub">{{ total }} active orders</div></div>
  <div style="display:flex;gap:8px;"><a href="/owner/dashboard" class="btn btn-ghost btn-sm">← Dashboard</a><button class="menu-toggle" onclick="openSidebar()">☰</button></div>
</div>
<div class="page-body" style="padding:16px 24px;">
  <div style="display:flex;gap:8px;flex-wrap:wrap;margin-bottom:16px;">
    <button class="wtab active" onclick="setWTab('all',this)" style="padding:7px 14px;border-radius:10px;border:2px solid var(--accent);background:var(--accent);color:#fff;font-size:12px;font-weight:800;cursor:pointer;">All</button>
    <button class="wtab" onclick="setWTab('naap',this)" style="padding:7px 14px;border-radius:10px;border:2px solid var(--border);background:#fff;color:var(--text-muted);font-size:12px;font-weight:800;cursor:pointer;">📐 नाप Pending</button>
    <button class="wtab" onclick="setWTab('cut',this)" style="padding:7px 14px;border-radius:10px;border:2px solid var(--border);background:#fff;color:var(--text-muted);font-size:12px;font-weight:800;cursor:pointer;">✂️ कटाई Pending</button>
    <button class="wtab" onclick="setWTab('stitch',this)" style="padding:7px 14px;border-radius:10px;border:2px solid var(--border);background:#fff;color:var(--text-muted);font-size:12px;font-weight:800;cursor:pointer;">🪡 सिलाई Pending</button>
    <button class="wtab" onclick="setWTab('done',this)" style="padding:7px 14px;border-radius:10px;border:2px solid var(--border);background:#fff;color:var(--text-muted);font-size:12px;font-weight:800;cursor:pointer;">✅ All Done</button>
  </div>
  <div style="overflow-x:auto;"><table style="width:100%;border-collapse:collapse;font-size:13px;background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 1px 8px rgba(0,0,0,0.07);">
    <thead><tr style="background:#f8fafc;border-bottom:2px solid var(--border);"><th style="padding:10px 16px;text-align:left;font-size:10px;font-weight:700;text-transform:uppercase;color:var(--text-muted);">Order</th><th style="padding:10px 16px;text-align:left;font-size:10px;font-weight:700;text-transform:uppercase;color:var(--text-muted);">Customer</th><th style="padding:10px 16px;text-align:left;font-size:10px;font-weight:700;text-transform:uppercase;color:var(--text-muted);">Garments</th><th style="padding:10px 16px;text-align:left;font-size:10px;font-weight:700;text-transform:uppercase;color:var(--text-muted);">Delivery</th><th style="padding:10px 16px;text-align:left;font-size:10px;font-weight:700;text-transform:uppercase;color:var(--text-muted);min-width:180px;">Progress</th><th style="padding:10px 16px;text-align:center;font-size:10px;font-weight:700;text-transform:uppercase;color:var(--text-muted);">Pending</th></tr></thead>
    <tbody id="wp-tbody">{% for o in orders %}{% set all_done = o.naap_pct >= 100 and o.cut_pct >= 100 and o.stitch_pct >= 100 %}<tr class="wrow" data-naap="{{ 'pending' if o.naap_pct < 100 else 'done' }}" data-cut="{{ 'pending' if o.cut_pct < 100 else 'done' }}" data-stitch="{{ 'pending' if o.stitch_pct < 100 else 'done' }}" data-alldone="{{ 'yes' if all_done else 'no' }}" style="border-bottom:1px solid var(--border);" onmouseover="this.style.background='#fafbff'" onmouseout="this.style.background=''"><td style="padding:12px 16px;"><div style="font-size:15px;font-weight:900;color:var(--accent);">#{{ o.display_code }}</div>{% if o.entry_code %}<span style="background:#f1f5f9;color:#64748b;font-size:10px;font-weight:700;padding:1px 6px;border-radius:5px;">(#{{ o.entry_code }})</span>{% endif %}{% if o.is_urgent %}<span style="background:#fee2e2;color:#dc2626;font-size:9px;font-weight:800;padding:1px 6px;border-radius:4px;">🔥 URGENT</span>{% endif %}</td><td style="padding:12px 16px;"><div style="font-weight:700;">{{ o.cname }}</div>{% if o.mobile %}<div style="font-size:11px;color:var(--text-muted);">{{ o.mobile }}</div>{% endif %}</td><td style="padding:12px 16px;color:var(--text-secondary);max-width:130px;font-size:12px;">{{ o.garments }}</td><td style="padding:12px 16px;"><div style="font-size:13px;font-weight:700;">{{ o.delivery_date }}</div><div style="font-size:11px;color:var(--text-muted);">{{ o.status|upper }}</div></td><td style="padding:12px 16px;"><div style="display:flex;justify-content:space-between;margin-bottom:4px;"><span style="font-size:9px;font-weight:800;color:{% if o.naap_pct>=100 %}#4f46e5{% else %}#9ca3af{% endif %};">नाप{% if o.naap_pct>=100 %} ✓{% else %} {{ o.naap_pct }}%{% endif %}</span><span style="font-size:9px;font-weight:800;color:{% if o.cut_pct>=100 %}#ea580c{% else %}#9ca3af{% endif %};">कटाई{% if o.cut_pct>=100 %} ✓{% else %} {{ o.cut_pct }}%{% endif %}</span><span style="font-size:9px;font-weight:800;color:{% if o.stitch_pct>=100 %}#16a34a{% else %}#9ca3af{% endif %};">सिलाई{% if o.stitch_pct>=100 %} ✓{% else %} {{ o.stitch_pct }}%{% endif %}</span></div><div style="display:flex;gap:2px;height:10px;"><div style="flex:1;background:#e5e7eb;border-radius:4px;overflow:hidden;"><div style="height:100%;background:#4f46e5;width:{{ o.naap_pct }}%;"></div></div><div style="flex:1;background:#e5e7eb;border-radius:4px;overflow:hidden;"><div style="height:100%;background:#ea580c;width:{{ o.cut_pct }}%;"></div></div><div style="flex:1;background:#e5e7eb;border-radius:4px;overflow:hidden;"><div style="height:100%;background:#16a34a;width:{{ o.stitch_pct }}%;"></div></div></div></td><td style="padding:12px 16px;text-align:center;">{% if all_done %}<span style="background:#d1fae5;color:#065f46;border-radius:8px;padding:4px 10px;font-size:11px;font-weight:800;">✅ All Done</span>{% else %}<div style="display:flex;flex-direction:column;gap:3px;align-items:center;">{% if o.naap_pct < 100 %}<span style="background:#eef2ff;color:#4f46e5;border-radius:6px;padding:2px 8px;font-size:10px;font-weight:800;">📐 नाप</span>{% endif %}{% if o.cut_pct < 100 %}<span style="background:#fff7ed;color:#ea580c;border-radius:6px;padding:2px 8px;font-size:10px;font-weight:800;">✂️ कटाई</span>{% endif %}{% if o.stitch_pct < 100 %}<span style="background:#f0fdf4;color:#16a34a;border-radius:6px;padding:2px 8px;font-size:10px;font-weight:800;">🪡 सिलाई</span>{% endif %}</div>{% endif %}</td></tr>{% else %}<tr><td colspan="6" style="padding:40px;text-align:center;color:var(--text-muted);">All orders delivered!</td></tr>{% endfor %}</tbody>
  </table></div>
</div>
{% endblock %}
{% block extra_js %}<script>function setWTab(key,btn){document.querySelectorAll(".wtab").forEach(function(b){b.style.background="#fff";b.style.color="var(--text-muted)";b.style.borderColor="var(--border)";});btn.style.background="var(--accent)";btn.style.color="#fff";btn.style.borderColor="var(--accent)";document.querySelectorAll(".wrow").forEach(function(r){var show=key==="all"||(key==="naap"&&r.dataset.naap==="pending")||(key==="cut"&&r.dataset.cut==="pending")||(key==="stitch"&&r.dataset.stitch==="pending")||(key==="done"&&r.dataset.alldone==="yes");r.style.display=show?"":"none";});}const SECS=5*60;let last=Date.now();["click","keydown","mousemove","touchstart"].forEach(ev=>document.addEventListener(ev,()=>{last=Date.now();},{passive:true}));setInterval(()=>{if(Math.floor((Date.now()-last)/1000)>=SECS)window.location.href="/owner/login?expired=1";},5000);window.addEventListener("pageshow",function(e){if(e.persisted){fetch("/owner/logout",{method:"POST",keepalive:true}).finally(()=>{window.location.href="/owner/login";})}});</script>{% endblock %}
"""

def fmt_d(d):
    """Format date string from YYYY-MM-DD to DD-MM-YYYY"""
    if not d: return "—"
    try:
        parts = str(d).split("-")
        if len(parts) == 3:
            return f"{parts[2]}-{parts[1]}-{parts[0]}"
    except: pass
    return str(d)


def owner_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("owner_logged_in"):
            # AJAX/API requests expect JSON, not a redirect
            if (request.is_json or
                request.path.startswith("/owner/api/") or
                request.headers.get("X-Requested-With") == "XMLHttpRequest"):
                return jsonify({"ok": False, "error": "Session expired — please refresh and login again"}), 401
            next_url = request.full_path if request.query_string else request.path
            return redirect(url_for("owner.login", next=next_url))
        return f(*args, **kwargs)
    return decorated

@bp.route("/login")
def login():
    if request.args.get("expired"):
        flash("Session expired. Please login again.", "warning")
    next_url = request.args.get("next", "/owner/dashboard")
    # Safety: only allow redirecting within our own site, never to an external URL
    if not next_url.startswith("/"):
        next_url = "/owner/dashboard"
    return render_template("owner/login.html", active_page=None, show_voice=False, urgent_count=0, next_url=next_url)

@bp.route("/verify-pin", methods=["POST"])
def verify_pin():
    data = request.get_json(silent=True) or {}
    entered = str(data.get("pin",""))
    real_pin = get_setting("owner_pin","1234")
    if entered == real_pin:
        session["owner_logged_in"] = True
        session.permanent = False
        return jsonify({"ok": True})
    return jsonify({"ok": False})

@bp.route("/logout", methods=["GET","POST"])
def logout():
    session.pop("owner_logged_in", None)
    return redirect(url_for("owner.login"))

@bp.route("/dashboard")
@owner_required
def dashboard():
    conn = get_db()
    today = date.today().isoformat()
    month_start = date.today().replace(day=1).isoformat()

    # Allow owner to pick a date via ?date=YYYY-MM-DD, default to most recent active date
    selected_date_str = request.args.get("date", "").strip()
    
    # If no date param, find the most recent date that has transactions
    # This fixes the server-date vs laptop-date mismatch
    if not selected_date_str:
        last_tx = conn.execute("SELECT tx_date FROM finance ORDER BY id DESC LIMIT 1").fetchone()
        if last_tx:
            # Use the date of the last transaction as "today" if it's within last 2 days
            from datetime import timedelta
            last_date = last_tx["tx_date"]
            server_today = date.today().isoformat()
            # If last transaction date is today or yesterday, show it
            yesterday = (date.today() - timedelta(days=1)).isoformat()
            if last_date in [server_today, yesterday]:
                selected_date = last_date
            else:
                selected_date = server_today
        else:
            selected_date = today
    else:
        selected_date = selected_date_str
    
    # Recalculate month_start based on selected_date
    try:
        sel_dt = datetime.strptime(selected_date, "%Y-%m-%d").date()
    except:
        sel_dt = date.today()
    month_start = sel_dt.replace(day=1).isoformat()

    # Today income/expense totals (using selected_date)
    rows = conn.execute("SELECT tx_type, SUM(amount) as total FROM finance WHERE tx_date=? GROUP BY tx_type",(selected_date,)).fetchall()
    fin_today = {r["tx_type"]: r["total"] or 0 for r in rows}

    # Month totals
    rows_m = conn.execute("SELECT tx_type, SUM(amount) as total FROM finance WHERE tx_date >= ? GROUP BY tx_type",(month_start,)).fetchall()
    fin_month = {r["tx_type"]: r["total"] or 0 for r in rows_m}

    # Month cash/upi breakdown
    month_cash = conn.execute(
        "SELECT COALESCE(SUM(amount),0) as t FROM finance WHERE tx_date>=? AND tx_type='income' AND LOWER(mode)='cash'",(month_start,)
    ).fetchone()["t"]
    month_upi = conn.execute(
        "SELECT COALESCE(SUM(amount),0) as t FROM finance WHERE tx_date>=? AND tx_type='income' AND LOWER(mode)='upi'",(month_start,)
    ).fetchone()["t"]

    # Selected date cash/upi
    cash_today = conn.execute(
        "SELECT COALESCE(SUM(amount),0) as t FROM finance WHERE tx_date=? AND tx_type='income' AND LOWER(mode)='cash'",(selected_date,)
    ).fetchone()["t"]
    upi_today = conn.execute(
        "SELECT COALESCE(SUM(amount),0) as t FROM finance WHERE tx_date=? AND tx_type='income' AND LOWER(mode)='upi'",(selected_date,)
    ).fetchone()["t"]

    # Full transaction log for selected date
    today_transactions = conn.execute(
        "SELECT f.*, o.order_code, o.repeat_of FROM finance f LEFT JOIN orders o ON o.id=f.order_id WHERE f.tx_date=? ORDER BY f.id DESC",(selected_date,)
    ).fetchall()

    dues = conn.execute("SELECT SUM(remaining) as total, COUNT(*) as cnt FROM orders WHERE status != 'delivered' AND remaining > 0").fetchone()
    work_today = conn.execute("SELECT SUM(qty_done) as total FROM work_logs WHERE log_date=?",(today,)).fetchone()
    # If no work logs yet, count items from today's orders as a proxy
    work_today_val = (work_today["total"] or 0) if work_today else 0
    if work_today_val == 0:
        items_today = conn.execute("""SELECT COALESCE(SUM(oi.quantity),0) as total
            FROM order_items oi JOIN orders o ON o.id=oi.order_id
            WHERE o.order_date=?""",(today,)).fetchone()
        work_today_proxy = items_today["total"] or 0
    low_stock = conn.execute("SELECT * FROM inventory WHERE quantity <= low_alert_at ORDER BY quantity ASC").fetchall()
    urgent_count = conn.execute("SELECT COUNT(*) as c FROM orders WHERE is_urgent=1 AND status != 'delivered' AND delivery_date >= ?",(today,)).fetchone()["c"]

    # Today's order activity
    # New orders: order_date = today (regardless of current status — an order
    # created AND delivered on the same day should still count as "today's order")
    new_orders_today = conn.execute(
        "SELECT COUNT(*) as c FROM orders WHERE order_date=?",
        (today,)
    ).fetchone()["c"]
    # Past orders: created today (entered today) AND already delivered AND old order_date
    past_orders_today = conn.execute(
        "SELECT COUNT(*) as c FROM orders WHERE created_at LIKE ? AND status='delivered' AND order_date != ?",
        (today + "%", today)
    ).fetchone()["c"]

    def _fmtd(d):
        if not d: return "—"
        p = str(d).split("-")
        return f"{p[2]}-{p[1]}-{p[0]}" if len(p)==3 else d

    new_orders_list = [{"code":r["order_code"],
        "display_code": r["repeat_of"] if r["repeat_of"] else r["order_code"],
        "entry_code": r["order_code"] if r["repeat_of"] else "",
        "name":r["name"], "status":r["status"],
        "date":_fmtd(r["order_date"]),"payable":int(r["payable_amount"] or 0),
        "paid":int(r["advance_paid"] or 0),"due":int(r["remaining"] or 0)}
        for r in conn.execute("""
            SELECT o.order_code,o.repeat_of,o.status,c.name,o.order_date,o.payable_amount,o.advance_paid,o.remaining
            FROM orders o JOIN customers c ON c.id=o.customer_id
            WHERE o.order_date=? ORDER BY o.id DESC
        """, (today,)).fetchall()]

    def fmt_12h(ts):
        if not ts or len(ts) < 16: return "—"
        try:
            from datetime import datetime as _dt
            t = _dt.strptime(str(ts)[:16], "%Y-%m-%d %H:%M")
            h, m = t.hour, t.strftime("%M")
            return f"{h%12 or 12}:{m} {'AM' if h<12 else 'PM'}"
        except: return "—"

    past_orders_list = [{"code":r["order_code"],"name":r["name"],
        "date":_fmtd(r["order_date"]),"payable":int(r["payable_amount"] or 0),
        "paid":int(r["advance_paid"] or 0),"due":int(r["remaining"] or 0),
        "time": fmt_12h(r["created_at"])}
        for r in conn.execute("""
            SELECT o.order_code,c.name,o.order_date,o.payable_amount,o.advance_paid,o.remaining,o.created_at
            FROM orders o JOIN customers c ON c.id=o.customer_id
            WHERE o.created_at LIKE ? AND o.status='delivered' AND o.order_date!=? ORDER BY o.id DESC
        """, (today+"%", today)).fetchall()]

    # All rate settings
    custom_rates = conn.execute("SELECT key,value FROM settings WHERE key LIKE '%rate%'").fetchall()
    conn.close()

    today_str = datetime.today().strftime("%A, %d %B %Y")
    d = date.today()
    today_date = f"{d.day:02d}-{d.month:02d}-{d.year}"
    last_backup = get_setting("last_backup_at", "")

    garment_names = [
        "Shirt","Shirt Linen","Pant","Pant Double","Jeans","Suit 2pc","Suit 3pc",
        "Blazer","Kurta","Kurta Pajama","Pajama","Pathani","Sherwani","Safari","Waistcoat",
        "Alteration","Cutting Only"
    ]
    # Customer rates (what customers pay)
    deleted_csv = get_setting("deleted_customer_rates", "")
    deleted_set = set(x.strip() for x in deleted_csv.split(",") if x.strip())
    garment_rates = {}
    for n in garment_names:
        if n in deleted_set:
            continue
        r = get_setting("customer_rate_"+n,"") or get_setting("rate_"+n,"0")
        garment_rates[n] = r
    # Add custom customer rates
    for row in custom_rates:
        if row["key"].startswith("customer_rate_"):
            name = row["key"][14:]
            if name not in garment_rates and name not in deleted_set:
                garment_rates[name] = row["value"]

    # Stitching rates (what employees get paid)
    stitch_rates = {n: get_setting("stitch_rate_"+n,"0") for n in garment_names}
    for row in custom_rates:
        if row["key"].startswith("stitch_rate_"):
            name = row["key"][12:]
            if name not in stitch_rates:
                stitch_rates[name] = row["value"]

    fresh_start_enabled = get_setting("utms_fresh_start", "0") == "1"
    fresh_start_date    = get_setting("utms_fresh_start_date", "2026-06-01")

    return render_template("owner/dashboard.html",
        active_page="owner_dashboard", show_voice=False, urgent_count=urgent_count,
        today_str=today_str, today_date=selected_date,
        selected_date=selected_date,
        low_stock=low_stock,
        garment_rates=garment_rates,
        stitch_rates=stitch_rates,
        today_transactions=today_transactions,
        last_backup=last_backup,
        fresh_start_enabled=fresh_start_enabled,
        fresh_start_date=fresh_start_date,
        stats={
            "today_income":  fin_today.get("income",0),
            "today_expense": fin_today.get("expense",0),
            "today_cash":    cash_today,
            "today_upi":     upi_today,
            "month_income":  fin_month.get("income",0),
            "month_cash":    month_cash,
            "month_upi":     month_upi,
            "month_net":     fin_month.get("income",0) - fin_month.get("expense",0),
            "pending_dues":  dues["total"] or 0 if dues else 0,
            "pending_orders":dues["cnt"] or 0 if dues else 0,
            "work_today":    work_today_proxy if (work_today["total"] or 0)==0 else work_today["total"],
            "new_orders_today":   new_orders_today,
            "past_orders_today":  past_orders_today,
        },
        new_orders_list=new_orders_list,
        past_orders_list=past_orders_list,
    )

@bp.route("/scan-diary")
@owner_required
def scan_diary():
    today = date.today().isoformat()
    urgent_count = conn_urgent = get_db()
    uc = conn_urgent.execute("SELECT COUNT(*) as c FROM orders WHERE is_urgent=1 AND status!='delivered'"
                             ).fetchone()["c"]
    conn_urgent.close()
    api_key_set = bool(get_setting("anthropic_api_key", ""))
    return render_template("owner/scan_diary.html",
        active_page="scan_diary", show_voice=False,
        urgent_count=uc, api_key_set=api_key_set)


@bp.route("/api/scan-diary/extract", methods=["POST"])
@owner_required
def api_scan_diary_extract():
    """Send diary image to Claude Vision and extract order details."""
    import base64 as _b64, urllib.request as _ur, json as _js, os as _os

    api_key = get_setting("anthropic_api_key", "") or _os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        return jsonify({"ok": False, "error": "Anthropic API key not set. Go to Admin → Settings to add it."})

    data = request.get_json(silent=True) or {}
    img_b64   = data.get("image_b64", "")   # base64 encoded image
    img_type  = data.get("image_type", "image/jpeg")  # MIME type

    if not img_b64:
        return jsonify({"ok": False, "error": "No image provided"})

    # Prompt for tailor diary extraction
    prompt = """You are scanning a tailor's diary/order book page (Indian tailoring shop).
Extract the following details from this image. The writing may be in Hindi, English, or Hinglish.

Return ONLY a JSON object with these fields (use null if not found):
{
  "order_code": "the order number/code (just the number, no #)",
  "customer_name": "customer name",
  "mobile": "phone number if visible",
  "order_date": "date in YYYY-MM-DD format if visible",
  "delivery_date": "delivery date in YYYY-MM-DD format if visible",
  "total_amount": numeric amount or null,
  "advance_paid": numeric advance paid or null,
  "garments": [
    {
      "type": "garment type in English (Shirt/Pant/Suit/Kameej/etc)",
      "qty": numeric quantity,
      "rate": numeric rate per piece or null,
      "measurements": {
        "field_name": "value"
      },
      "notes": "any style notes"
    }
  ],
  "notes": "any other notes"
}

Common measurement fields: Lambai/Length, Seena/Chest, Kamar/Waist, Shoulder, Aastin/Sleeve, Collar, Seat, Jangh, Goda, Mori, Cough/Kaf.
Return ONLY the JSON, no explanation."""

    payload = {
        "model": "claude-sonnet-4-6",
        "max_tokens": 1024,
        "messages": [{
            "role": "user",
            "content": [
                {"type": "image", "source": {"type": "base64", "media_type": img_type, "data": img_b64}},
                {"type": "text", "text": prompt}
            ]
        }]
    }

    try:
        req = _ur.Request(
            "https://api.anthropic.com/v1/messages",
            data=_js.dumps(payload).encode(),
            headers={
                "Content-Type": "application/json",
                "x-api-key": api_key,
                "anthropic-version": "2023-06-01"
            },
            method="POST"
        )
        with _ur.urlopen(req, timeout=30) as r:
            resp = _js.loads(r.read())

        text = resp["content"][0]["text"].strip()
        # Strip markdown fences if present
        import re as _re
        text = _re.sub(r'^```json\s*|^```\s*|```$', '', text, flags=_re.MULTILINE).strip()
        extracted = _js.loads(text)
        return jsonify({"ok": True, "data": extracted})

    except Exception as e:
        return jsonify({"ok": False, "error": f"AI extraction failed: {str(e)[:200]}"})


@bp.route("/api/settings/anthropic-key", methods=["POST"])
@owner_required
def save_anthropic_key():
    data = request.get_json(silent=True) or {}
    key  = (data.get("key") or "").strip()
    set_setting("anthropic_api_key", key)
    return jsonify({"ok": True})


@bp.route("/bulk-import")
@owner_required
def bulk_import():
    today = date.today().isoformat()
    conn = get_db()
    uc = conn.execute("SELECT COUNT(*) as c FROM orders WHERE is_urgent=1 AND status!='delivered'").fetchone()["c"]
    conn.close()
    return render_template("owner/bulk_import.html",
        active_page="bulk_import", show_voice=False, urgent_count=uc)


@bp.route("/bulk-import/template")
@owner_required
def bulk_import_template():
    """Download ALL current UTMS orders as Excel — add new rows and re-upload to import."""
    import openpyxl as xl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import io

    wb = xl.Workbook()
    ws = wb.active
    ws.title = "Orders"

    # Headers
    headers = [
        "Order Code", "Customer Name", "Mobile", "Address",
        "Order Date\n(DD-MM-YYYY)", "Delivery Date\n(DD-MM-YYYY)",
        "Garment Type", "Quantity", "Rate (₹)",
        "Lambai", "Seena", "Kamar", "Shoulder", "Aastin",
        "Collar", "Cough", "Seat", "Mori", "Jangh", "Goda", "Langot",
        "Total Amount (₹)", "Advance Paid (₹)", "Notes",
        "Status\n(new/delivered)"
    ]

    # Style header row
    hdr_fill = PatternFill("solid", fgColor="3B3EA1")
    hdr_font = Font(color="FFFFFF", bold=True, size=10)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 32

    # Column widths
    widths = [12, 18, 14, 18, 14, 14, 14, 8, 10,
              8, 8, 8, 10, 8, 8, 8, 8, 8, 8, 8, 8,
              16, 16, 20, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[xl.utils.get_column_letter(i)].width = w

    # Export ACTUAL data from DB
    import json as _json
    conn_exp = get_db()
    MEAS_FIELDS_EXP = ["Lambai","Seena","Kamar","Shoulder","Aastin",
                        "Collar","Cough","Seat","Mori","Jangh","Goda","Langot"]

    rows_db = conn_exp.execute("""
        SELECT o.order_code, c.name, c.mobile, c.address,
               o.order_date, o.delivery_date, o.status,
               o.payable_amount, o.advance_paid, o.note
        FROM orders o JOIN customers c ON c.id=o.customer_id
        ORDER BY o.id DESC
    """).fetchall()

    row_idx = 2
    alt_fill  = PatternFill("solid", fgColor="F0F9FF")
    norm_fill = PatternFill("solid", fgColor="FFFFFF")

    def fmtd(d):
        if not d: return ""
        p = str(d).split("-")
        return f"{p[2]}-{p[1]}-{p[0]}" if len(p)==3 else d

    for oi, o in enumerate(rows_db):
        items = conn_exp.execute(
            "SELECT garment_type, quantity, rate, measurements, notes FROM order_items WHERE order_id=(SELECT id FROM orders WHERE order_code=?) ORDER BY id",
            (o["order_code"],)
        ).fetchall()
        if not items:
            # Order with no items — still export one row
            items = [{"garment_type":"","quantity":1,"rate":0,"measurements":"{}","notes":""}]

        for item in items:
            try: meas = _json.loads(item["measurements"] or "{}")
            except: meas = {}
            fill = alt_fill if oi % 2 == 0 else norm_fill
            row_data = [
                o["order_code"], o["name"], o["mobile"] or "", o["address"] or "",
                fmtd(o["order_date"]), fmtd(o["delivery_date"]),
                item["garment_type"], item["quantity"], item["rate"],
            ]
            for mf in MEAS_FIELDS_EXP:
                row_data.append(meas.get(mf, ""))
            row_data += [
                int(o["payable_amount"] or 0),
                int(o["advance_paid"] or 0),
                item["notes"] or "",
                "delivered" if o["status"]=="delivered" else "new"
            ]
            for c_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=c_idx, value=val)
                cell.alignment = Alignment(horizontal="center")
                cell.border = border
                cell.fill = fill
            row_idx += 1

    conn_exp.close()

    # Instructions sheet
    ws2 = wb.create_sheet("Instructions")
    instructions = [
        ["📋 UTMS BULK IMPORT — INSTRUCTIONS", ""],
        ["", ""],
        ["IMPORTANT RULES:", ""],
        ["1. Same Order Code in multiple rows = SAME ORDER (multiple garments)", ""],
        ["2. Leave Order Code BLANK = auto-assign next available code", ""],
        ["3. Status: 'delivered' = past order  |  'new' = active order", ""],
        ["4. Date format: DD-MM-YYYY  (e.g. 25-06-2026)", ""],
        ["5. Measurements: fill only the ones that apply — rest leave blank", ""],
        ["6. One row = One garment. Add multiple rows for multiple garments per order.", ""],
        ["", ""],
        ["REQUIRED FIELDS:", ""],
        ["✅ Customer Name", "Required"],
        ["✅ Garment Type", "Required (Shirt/Pant/Kurta/Suit/etc.)"],
        ["⬜ Order Code", "Optional — leave blank for auto"],
        ["⬜ All measurements", "Optional — fill what you know"],
        ["", ""],
        ["GARMENT TYPES SUPPORTED:", ""],
        ["Shirt, Shirt Linen, Pant, Pant Double, Kurta, Kurta Pajama,", ""],
        ["Suit 2pc, Suit 3pc, Blazer, Pajama, Pathani, Sherwani,", ""],
        ["Safari, Waistcoat, Alteration, Cutting Only, or any custom type", ""],
    ]
    for r, (a, b) in enumerate(instructions, 1):
        ws2.cell(row=r, column=1, value=a).font = Font(bold=(r==1 or a.endswith(":")))
        ws2.cell(row=r, column=2, value=b)
    ws2.column_dimensions["A"].width = 60
    ws2.column_dimensions["B"].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from flask import send_file
    return send_file(buf, as_attachment=True,
                     download_name=f"UTMS_Orders_{date.today().isoformat()}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@bp.route("/bulk-import/upload", methods=["POST"])
@owner_required
def bulk_import_upload():
    """Process uploaded Excel file and create orders."""
    import openpyxl as xl
    import io

    f = request.files.get("excel_file")
    if not f or not f.filename.endswith((".xlsx",".xls")):
        return jsonify({"ok": False, "error": "Excel file (.xlsx) upload karo"})

    try:
        wb = xl.load_workbook(io.BytesIO(f.read()), data_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({"ok": False, "error": f"File read error: {e}"})

    # Parse rows — skip header row 1
    orders_map = {}  # order_code -> {customer, garments[]}
    errors = []
    row_count = 0

    MEAS_FIELDS = ["Lambai","Seena","Kamar","Shoulder","Aastin",
                   "Collar","Cough","Seat","Mori","Jangh","Goda","Langot"]

    def cell_val(row, idx):
        v = row[idx-1].value if len(row) >= idx else None
        return str(v).strip() if v is not None and str(v).strip() else ""

    for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), 2):
        if not any(c.value for c in row):
            continue
        row_count += 1

        order_code   = cell_val(row, 1)
        cust_name    = cell_val(row, 2)
        mobile       = cell_val(row, 3)
        address      = cell_val(row, 4)
        order_date   = cell_val(row, 5)
        delivery_date= cell_val(row, 6)
        garment_type = cell_val(row, 7)
        qty_raw      = cell_val(row, 8)
        rate_raw     = cell_val(row, 9)
        # Measurements cols 10-21
        meas = {}
        for mi, mf in enumerate(MEAS_FIELDS, 10):
            v = cell_val(row, mi)
            if v:
                meas[mf] = v
        total_raw    = cell_val(row, 22)
        advance_raw  = cell_val(row, 23)
        notes        = cell_val(row, 24)
        status_raw   = cell_val(row, 25).lower()

        if not cust_name:
            errors.append(f"Row {r_idx}: Customer Name missing — skipped")
            continue
        if not garment_type:
            errors.append(f"Row {r_idx}: Garment Type missing — skipped")
            continue

        # Convert date DD-MM-YYYY → YYYY-MM-DD
        def parse_date(d):
            if not d: return ""
            for fmt in ["%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y"]:
                try:
                    from datetime import datetime as dt
                    return dt.strptime(d, fmt).strftime("%Y-%m-%d")
                except: pass
            return ""

        key = order_code if order_code else f"__auto__{r_idx}"
        if key not in orders_map:
            orders_map[key] = {
                "order_code":    order_code,
                "customer_name": cust_name,
                "mobile":        mobile,
                "address":       address,
                "order_date":    parse_date(order_date),
                "delivery_date": parse_date(delivery_date),
                "total_amount":  float(total_raw) if total_raw else 0,
                "advance_paid":  float(advance_raw) if advance_raw else 0,
                "is_delivered":  status_raw in ("delivered","past","d","yes","1",""),
                "garments":      [],
            }

        try: qty = int(float(qty_raw)) if qty_raw else 1
        except: qty = 1
        try: rate = float(rate_raw) if rate_raw else 0
        except: rate = 0

        orders_map[key]["garments"].append({
            "type": garment_type, "qty": qty, "rate": rate,
            "meas": meas, "notes": notes
        })

    if not orders_map:
        return jsonify({"ok": False, "error": "Koi valid row nahi mili. Template check karo."})

    # Now save each order via past_orders_save logic
    created = 0
    skipped = 0
    conn = get_db()

    for key, od in orders_map.items():
        try:
            # Check if order code already exists
            if od["order_code"]:
                exists = conn.execute("SELECT id FROM orders WHERE order_code=?",
                                      (od["order_code"],)).fetchone()
                if exists:
                    errors.append(f"Order #{od['order_code']} already exists — skipped")
                    skipped += 1
                    continue

            # No automatic matching by mobile or name — every imported row
            # gets its own separate, brand-new customer record.
            now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            new_cust = conn.execute(
                "INSERT INTO customers(name,mobile,address,created_at) VALUES(?,?,?,?) RETURNING id",
                (od["customer_name"], od["mobile"] or None, od["address"], now_str)).fetchone()
            cust_id = new_cust["id"] if new_cust else None

            # Get order code — bulk import must have explicit code, never auto-generate
            order_code = od.get("order_code", "").strip()
            if not order_code:
                errors.append(f"Row skipped — no order code provided")
                continue

            payable   = od["total_amount"]
            advance   = od["advance_paid"]
            remaining = max(0, payable - advance)
            status    = "delivered" if od["is_delivered"] else "pending"
            now_str   = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            today_str = date.today().isoformat()

            conn.execute("""
                INSERT INTO orders(order_code,customer_id,order_date,delivery_date,
                    status,payable_amount,advance_paid,remaining,created_at)
                VALUES(?,?,?,?,?,?,?,?,?)
            """, (order_code, cust_id,
                  od["order_date"] or today_str,
                  od["delivery_date"] or today_str,
                  status, payable, advance, remaining, now_str))
            conn.commit()
            order_id = conn.execute("SELECT id FROM orders WHERE order_code=?",
                                    (order_code,)).fetchone()["id"]

            # Save garments + measurements
            import json as _json
            for g in od["garments"]:
                amount = g["qty"] * g["rate"]
                meas_json = _json.dumps(g["meas"])
                conn.execute("""
                    INSERT INTO order_items(order_id,garment_type,quantity,rate,amount,measurements,notes)
                    VALUES(?,?,?,?,?,?,?)
                """, (order_id, g["type"], g["qty"], g["rate"], amount, meas_json, g["notes"]))
            conn.commit()
            created += 1

        except Exception as e:
            errors.append(f"Order {od.get('order_code') or od['customer_name']}: {str(e)[:100]}")
            skipped += 1

    conn.close()
    return jsonify({
        "ok": True,
        "created": created,
        "skipped": skipped,
        "errors": errors[:20],
        "total_rows": row_count
    })


@bp.route("/api/fix-shared-mobile")
@owner_required
def fix_shared_mobile():
    """Show all orders grouped by mobile where multiple orders exist — bulk name fix tool."""
    conn = get_db()

    # Find all mobiles that have 2+ orders
    rows = conn.execute("""
        SELECT c.mobile, c.name, c.id as cust_id,
               o.order_code, o.order_date, o.created_at,
               GROUP_CONCAT(oi.garment_type, ', ') as garments
        FROM customers c
        JOIN orders o ON o.customer_id = c.id
        LEFT JOIN order_items oi ON oi.order_id = o.id
        WHERE c.mobile IS NOT NULL AND c.mobile != ''
          AND c.mobile IN (
              SELECT mobile FROM customers
              WHERE mobile IS NOT NULL AND mobile != ''
              GROUP BY mobile HAVING COUNT(*) >= 1
          )
          AND (SELECT COUNT(*) FROM orders o2 WHERE o2.customer_id = c.id) >= 2
        GROUP BY o.order_code
        ORDER BY c.mobile, o.id DESC
    """).fetchall()

    conn.close()

    # Group by mobile
    from collections import OrderedDict
    groups = OrderedDict()
    for r in rows:
        m = r["mobile"]
        if m not in groups:
            groups[m] = []
        groups[m].append(dict(r))

    if not groups:
        return """<div style='font-family:sans-serif;padding:30px;'>
        <h2>✅ Koi shared mobile orders nahi mile!</h2>
        <a href='/owner/orders'>← Back to Orders</a></div>"""

    html = """<!DOCTYPE html><html><head>
    <meta charset='utf-8'>
    <title>Fix Shared Mobile Names</title>
    <style>
      body{font-family:-apple-system,sans-serif;background:#f8fafc;margin:0;padding:20px;}
      h1{color:#1e293b;font-size:22px;margin-bottom:4px;}
      .subtitle{color:#64748b;font-size:13px;margin-bottom:24px;}
      .group{background:#fff;border-radius:12px;border:1.5px solid #e2e8f0;margin-bottom:20px;overflow:hidden;}
      .group-header{background:#1e293b;color:#fff;padding:12px 16px;font-size:13px;font-weight:700;}
      .mobile-tag{background:#6366f1;color:#fff;padding:2px 10px;border-radius:20px;font-size:12px;}
      table{width:100%;border-collapse:collapse;}
      th{background:#f8fafc;padding:10px 14px;font-size:11px;text-transform:uppercase;color:#64748b;text-align:left;border-bottom:1px solid #e2e8f0;}
      td{padding:10px 14px;border-bottom:1px solid #f1f5f9;font-size:13px;vertical-align:middle;}
      tr:last-child td{border-bottom:none;}
      input.name-inp{padding:6px 10px;border:1.5px solid #e2e8f0;border-radius:7px;font-size:13px;font-weight:600;width:160px;}
      input.name-inp:focus{border-color:#6366f1;outline:none;}
      .fix-btn{padding:6px 14px;background:#16a34a;color:#fff;border:none;border-radius:7px;font-size:12px;font-weight:700;cursor:pointer;}
      .fix-btn:hover{background:#15803d;}
      .done-tag{color:#16a34a;font-size:12px;font-weight:700;display:none;}
      .code-tag{font-weight:800;color:#4f46e5;}
      .garment-tag{font-size:11px;color:#64748b;}
    </style>
    </head><body>
    <h1>🔧 Fix Shared Mobile Names</h1>
    <div class='subtitle'>Yahan wo orders hain jisme same mobile se multiple orders hain. Har order ke sahi naam fix karo.</div>
    """

    for mobile, orders in groups.items():
        html += f"""<div class='group'>
        <div class='group-header'>
          📱 <span class='mobile-tag'>{mobile}</span>
          &nbsp; {len(orders)} orders — same mobile number
        </div>
        <table>
        <tr><th>Order</th><th>Date</th><th>Garments</th><th>Current Name</th><th>Correct Name</th><th></th></tr>
        """
        for o in orders:
            odate = (o.get('order_date') or o.get('created_at') or '')[:10]
            garments = o.get('garments') or '—'
            html += f"""<tr id='row-{o["order_code"]}'>
            <td class='code-tag'>#{o["order_code"]}</td>
            <td>{odate}</td>
            <td class='garment-tag'>{garments[:40]}</td>
            <td>{o["name"]}</td>
            <td><input class='name-inp' id='name-{o["order_code"]}' value='{o["name"]}' placeholder='Sahi naam'></td>
            <td>
              <button class='fix-btn' onclick='fixName("{o["order_code"]}")'>Save</button>
              <span class='done-tag' id='done-{o["order_code"]}'>✅ Done</span>
            </td>
            </tr>"""
        html += "</table></div>"

    html += """
    <script>
    async function fixName(code) {
      var name = document.getElementById('name-' + code).value.trim();
      if (!name) return;
      var btn = event.target;
      btn.disabled = true; btn.textContent = '...';
      var r = await fetch('/owner/api/fix-order-customer', {
        method: 'POST',
        headers: {'Content-Type': 'application/x-www-form-urlencoded'},
        body: 'code=' + encodeURIComponent(code) + '&name=' + encodeURIComponent(name)
      });
      var text = await r.text();
      if (text.includes('Fixed') || text.includes('✅')) {
        btn.style.display = 'none';
        document.getElementById('done-' + code).style.display = 'inline';
        document.getElementById('row-' + code).style.background = '#f0fdf4';
      } else {
        btn.textContent = 'Save'; btn.disabled = false;
        alert('Error fixing #' + code);
      }
    }
    </script>
    </body></html>"""

    return html

@bp.route("/api/fix-order-customer", methods=["GET", "POST"])
@owner_required
def fix_order_customer():
    """Fix customer name for a specific order — useful when same-mobile overwrote a name."""
    if request.method == "GET":
        code = request.args.get("code", "")
        new_name = request.args.get("name", "")
        if not code or not new_name:
            return """<h2>Fix Order Customer Name</h2>
            <form method='GET'>
              Order Code: <input name='code' placeholder='e.g. 3917'><br><br>
              Correct Name: <input name='name' placeholder='e.g. Ronak'><br><br>
              <button type='submit'>Preview</button>
            </form>"""
        conn = get_db()
        order = conn.execute(
            "SELECT o.id, o.order_code, o.customer_id, c.name, c.mobile FROM orders o JOIN customers c ON c.id=o.customer_id WHERE o.order_code=?",
            (code,)).fetchone()
        if not order:
            conn.close()
            return f"<h2>❌ Order #{code} not found</h2><a href='/owner/api/fix-order-customer'>← Back</a>"
        conn.close()
        return f"""<h2>Fix Customer for Order #{code}</h2>
        <p>Current name: <b>{order['name']}</b> (Mobile: {order['mobile']})</p>
        <p>New name: <b>{new_name}</b></p>
        <form method='POST'>
          <input type='hidden' name='code' value='{code}'>
          <input type='hidden' name='name' value='{new_name}'>
          <button type='submit' style='background:#16a34a;color:#fff;padding:10px 20px;border:none;border-radius:8px;font-size:15px;cursor:pointer;'>
            ✅ Confirm Fix
          </button>
          &nbsp;<a href='/owner/api/fix-order-customer'>← Cancel</a>
        </form>"""

    # POST — apply fix
    code = request.form.get("code", "").strip()
    new_name = request.form.get("name", "").strip()
    if not code or not new_name:
        return "<h2>❌ Missing fields</h2>"

    conn = get_db()
    order = conn.execute(
        "SELECT o.id, o.order_code, o.customer_id, c.name, c.mobile, c.address FROM orders o JOIN customers c ON c.id=o.customer_id WHERE o.order_code=?",
        (code,)).fetchone()
    if not order:
        conn.close()
        return f"<h2>❌ Order #{code} not found</h2>"

    # Create new customer with correct name (don't touch existing record)
    conn.execute("INSERT INTO customers(name, mobile, address) VALUES(?,?,?)",
                 (new_name, order["mobile"] or "", order["address"] or ""))
    new_cust = conn.execute(
        "SELECT id FROM customers WHERE name=? ORDER BY id DESC LIMIT 1", (new_name,)).fetchone()
    if new_cust:
        conn.execute("UPDATE orders SET customer_id=? WHERE order_code=?",
                     (new_cust["id"], code))
        conn.commit()
        conn.close()
        return f"""<h2>✅ Fixed!</h2>
        <p>Order #{code} is now linked to <b>{new_name}</b></p>
        <a href='/owner/orders'>← All Orders</a>"""
    conn.close()
    return "<h2>❌ Could not create customer</h2>"


@bp.route("/order-ledger")
@owner_required
def order_ledger():
    """Admin page — search order code to see full payment history."""
    return render_template("owner/order_ledger.html", active_page="order_ledger")


@bp.route("/api/order-ledger")
@owner_required
def api_order_ledger():
    """Return order details + payment history + customer's other orders."""
    code = request.args.get("code", "").strip().lstrip("#")
    if not code:
        return jsonify({"ok": False, "error": "Order code required"})

    conn = get_db()

    # Get order + customer
    order = conn.execute("""
        SELECT o.*, c.name as cname, c.mobile, c.address,
               string_agg(oi.garment_type, ', ') as garments
        FROM orders o
        LEFT JOIN customers c ON c.id = o.customer_id
        LEFT JOIN order_items oi ON oi.order_id = o.id
        WHERE o.order_code = ?
        GROUP BY o.id, c.name, c.mobile, c.address
    """, (code,)).fetchone()

    if not order:
        conn.close()
        return jsonify({"ok": False, "error": f"Order #{code} nahi mila"})

    # Payment history from finance table (linked by order_id)
    payments = conn.execute("""
        SELECT tx_date, amount, mode, note, created_at
        FROM finance
        WHERE order_id = ? AND tx_type = 'income'
        ORDER BY tx_date ASC, id ASC
    """, (order["id"],)).fetchall()

    # Also check advance payment if no finance records
    if not payments and order["advance_paid"] > 0:
        # Fallback: show order_date as payment date
        payments = [{
            "tx_date": order["order_date"] or order["created_at"] or "",
            "amount": order["advance_paid"],
            "mode": order["payment_mode"] or "cash",
            "note": "Advance (order ke waqt)",
            "created_at": ""
        }]
        payments_list = payments
    else:
        payments_list = [dict(p) for p in payments]

    # Other orders by same customer
    other_orders = conn.execute("""
        SELECT order_code, repeat_of, order_date, delivery_date, payable_amount, remaining, status
        FROM orders
        WHERE customer_id = ? AND order_code != ?
        ORDER BY id DESC LIMIT 20
    """, (order["customer_id"], code)).fetchall()

    conn.close()

    other_orders_out = [{
        "order_code": o["order_code"],
        "display_code": o["repeat_of"] if o["repeat_of"] else o["order_code"],
        "entry_code": o["order_code"] if o["repeat_of"] else "",
        "order_date": o["order_date"], "delivery_date": o["delivery_date"],
        "payable_amount": o["payable_amount"], "remaining": o["remaining"], "status": o["status"],
    } for o in other_orders]

    return jsonify({
        "ok": True,
        "order": {
            "order_code":    order["order_code"],
            "display_code":  order["repeat_of"] if order["repeat_of"] else order["order_code"],
            "entry_code":    order["order_code"] if order["repeat_of"] else "",
            "cname":         order["cname"] or "—",
            "mobile":        order["mobile"] or "",
            "order_date":    order["order_date"] or "",
            "delivery_date": order["delivery_date"] or "",
            "delivered_at":  order["delivered_at"] or "",
            "status":        order["status"],
            "is_urgent":     bool(order["is_urgent"]),
            "payable_amount":order["payable_amount"] or 0,
            "advance_paid":  order["advance_paid"] or 0,
            "remaining":     order["remaining"] or 0,
            "payment_mode":  order["payment_mode"] or "cash",
            "garments":      order["garments"] or "—",
            "note":          order["note"] or "",
        },
        "payments": payments_list,
        "other_orders": other_orders_out,
    })


@bp.route("/api/fix-finance-mode/<order_code>", methods=["POST","GET"])
@owner_required
def fix_finance_mode(order_code):
    """Fix payment mode in finance entries for an order."""
    mode = request.args.get("mode", "").strip() or (request.get_json(silent=True) or {}).get("mode","")
    if not mode:
        return jsonify({"ok": False, "error": "mode param required (cash/upi/bank)"})
    conn = get_db()
    order = conn.execute("SELECT id FROM orders WHERE order_code=?", (order_code,)).fetchone()
    if not order:
        conn.close()
        return jsonify({"ok": False, "error": f"Order #{order_code} not found"})
    rows = conn.execute(
        "UPDATE finance SET mode=? WHERE order_id=? AND tx_type='income' RETURNING id",
        (mode, order["id"])
    ).fetchall()
    conn.execute("UPDATE orders SET payment_mode=? WHERE id=?", (mode, order["id"]))
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "updated": len(rows), "message": f"✅ {len(rows)} finance entries updated to {mode}"})


@bp.route("/api/diagnose-finance-bulk")
@owner_required
def api_diagnose_finance_bulk():
    """Same as diagnose-finance but for many order codes at once —
    comma-separated in ?codes=3898,3860,2358 — to quickly spot a pattern
    across several orders instead of checking one at a time."""
    codes_param = request.args.get("codes", "")
    codes = [c.strip().lstrip("#") for c in codes_param.split(",") if c.strip()]
    if not codes:
        return jsonify({"ok": False, "error": "Pass ?codes=3898,3860,..."})
    conn = get_db()
    out = []
    for code in codes:
        order = conn.execute("""
            SELECT id, order_code, repeat_of, status, payable_amount, advance_paid, remaining,
                   payment_mode, delivered_at
            FROM orders WHERE order_code=?
        """, (code,)).fetchone()
        if not order:
            out.append({"code": code, "found": False})
            continue
        finance_rows = conn.execute(
            "SELECT id, order_id, tx_date, tx_type, category, amount, mode, note, created_at FROM finance WHERE order_id=? ORDER BY id DESC",
            (order["id"],)
        ).fetchall()
        # Also check for repeat entries OF this code — if the user actually
        # delivered a repeat visit (displayed as "#code (Entry #xxxx)"), the
        # real data lives under that entry's own order_code, not this one.
        repeats = conn.execute("""
            SELECT id, order_code, status, payable_amount, advance_paid, remaining, delivered_at
            FROM orders WHERE repeat_of=?
            ORDER BY id DESC
        """, (code,)).fetchall()
        repeat_out = []
        for rep in repeats:
            rep_fin = conn.execute(
                "SELECT id, tx_type, category, amount, mode, created_at FROM finance WHERE order_id=? ORDER BY id DESC",
                (rep["id"],)
            ).fetchall()
            repeat_out.append({
                "entry_code": rep["order_code"], "status": rep["status"],
                "payable_amount": rep["payable_amount"], "advance_paid": rep["advance_paid"],
                "remaining": rep["remaining"], "delivered_at": rep["delivered_at"],
                "finance_entry_count": len(rep_fin),
                "finance_entries": [dict(r) for r in rep_fin],
            })
        out.append({
            "code": code, "found": True,
            "order": dict(order),
            "finance_entry_count": len(finance_rows),
            "finance_entries": [dict(r) for r in finance_rows],
            "repeat_entries": repeat_out,
        })
    conn.close()
    return jsonify({"ok": True, "results": out})


@bp.route("/api/diagnose-finance/<code>")
@owner_required
def api_diagnose_finance(code):
    """Read-only: show the order's current stored fields + every finance
    entry linked to it, exactly as stored in the DB. Used to check why an
    amount isn't showing on the Finance page (wrong date? wrong mode? entry
    missing entirely?) without guessing."""
    conn = get_db()
    order = conn.execute("""
        SELECT id, order_code, repeat_of, status, payable_amount, advance_paid, remaining,
               payment_mode, delivered_at
        FROM orders WHERE order_code=?
    """, (code,)).fetchone()
    if not order:
        conn.close()
        return jsonify({"ok": False, "error": f"Order #{code} not found"})
    finance_rows = conn.execute(
        "SELECT id, order_id, tx_date, tx_type, category, amount, mode, note, created_at FROM finance WHERE order_id=? ORDER BY id DESC",
        (order["id"],)
    ).fetchall()
    conn.close()
    return jsonify({
        "ok": True,
        "order": dict(order),
        "finance_entry_count": len(finance_rows),
        "finance_entries": [dict(r) for r in finance_rows],
    })


@bp.route("/api/diagnose-order/<code>")
@owner_required
def api_diagnose_order(code):
    """Diagnose order customer linkage — check if wrong customer is linked."""
    conn = get_db()
    # Get order + customer details
    order = conn.execute("""
        SELECT o.id, o.order_code, o.customer_id,
               c.id as cid, c.name as cname, c.mobile, c.address,
               o.advance_paid, o.payable_amount
        FROM orders o
        LEFT JOIN customers c ON c.id = o.customer_id
        WHERE o.order_code = ?
    """, (code,)).fetchone()

    if not order:
        conn.close()
        return jsonify({"ok": False, "error": f"Order #{code} not found"})

    # Find all orders linked to same customer
    same_cust_orders = conn.execute("""
        SELECT order_code, order_date, status, payable_amount
        FROM orders WHERE customer_id = ?
        ORDER BY id DESC LIMIT 10
    """, (order["customer_id"],)).fetchall()

    # Find all customers with same mobile
    same_mobile_custs = []
    if order["mobile"]:
        same_mobile_custs = conn.execute("""
            SELECT id, name, mobile FROM customers WHERE mobile = ?
        """, (order["mobile"],)).fetchall()

    conn.close()
    return jsonify({
        "ok": True,
        "order": {
            "code": order["order_code"],
            "customer_id": order["customer_id"],
            "customer_name": order["cname"],
            "customer_mobile": order["mobile"],
        },
        "all_orders_of_this_customer": [dict(r) for r in same_cust_orders],
        "customers_with_same_mobile": [dict(r) for r in same_mobile_custs],
        "fix_url": f"/owner/api/fix-order-customer?code={code}&name=SAHI_NAAM_YAHAN"
    })


@bp.route("/api/fix-split-customer", methods=["GET","POST"])
@owner_required
def api_fix_split_customer():
    """Fix two orders wrongly sharing a customer.
    Usage: GET  ?code1=2408&code2=3279  → diagnose
           POST ?code1=2408&mobile1=9876543210&code2=3279&mobile2=9999999999
                → creates separate customers, fixes both orders
    """
    from datetime import datetime as _dt
    code1  = request.args.get("code1","").strip()
    code2  = request.args.get("code2","").strip()
    conn   = get_db()

    def get_order(code):
        return conn.execute("""
            SELECT o.id, o.order_code, o.customer_id,
                   c.name as cname, c.mobile, c.address
            FROM orders o
            LEFT JOIN customers c ON c.id=o.customer_id
            WHERE o.order_code=?
        """, (code,)).fetchone()

    o1 = get_order(code1)
    o2 = get_order(code2) if code2 else None

    # GET → diagnose only
    if request.method == "GET":
        conn.close()
        return jsonify({
            "order1": dict(o1) if o1 else None,
            "order2": dict(o2) if o2 else None,
            "same_customer": o1 and o2 and o1["customer_id"]==o2["customer_id"],
            "fix_instructions": {
                "method": "POST",
                "params": f"code1={code1}&mobile1=CORRECT_MOBILE_FOR_{code1}&code2={code2}&mobile2=CORRECT_MOBILE_FOR_{code2}"
            }
        })

    # POST → actually fix
    mobile1 = request.args.get("mobile1","").strip()
    mobile2 = request.args.get("mobile2","").strip()
    now_str = _dt.now().strftime("%Y-%m-%d %H:%M:%S")

    fixed = []

    if o1 and mobile1:
        # Check if customer with this name+mobile already exists
        existing = conn.execute(
            "SELECT id FROM customers WHERE mobile=? AND name=?",
            (mobile1, o1["cname"])
        ).fetchone()
        if existing:
            cust_id = existing["id"]
        else:
            cust_id = conn.execute(
                "INSERT INTO customers(name, mobile, address, created_at) VALUES(?,?,?,?) RETURNING id",
                (o1["cname"], mobile1, o1["address"] or "", now_str)
            ).fetchone()["id"]
        conn.execute("UPDATE orders SET customer_id=? WHERE order_code=?", (cust_id, code1))
        fixed.append(f"Order #{code1} → customer_id={cust_id} (mobile={mobile1})")

    if o2 and mobile2:
        existing = conn.execute(
            "SELECT id FROM customers WHERE mobile=? AND name=?",
            (mobile2, o2["cname"])
        ).fetchone()
        if existing:
            cust_id2 = existing["id"]
        else:
            cust_id2 = conn.execute(
                "INSERT INTO customers(name, mobile, address, created_at) VALUES(?,?,?,?) RETURNING id",
                (o2["cname"], mobile2, o2["address"] or "", now_str)
            ).fetchone()["id"]
        conn.execute("UPDATE orders SET customer_id=? WHERE order_code=?", (cust_id2, code2))
        fixed.append(f"Order #{code2} → customer_id={cust_id2} (mobile={mobile2})")

    conn.commit()
    conn.close()
    return jsonify({"ok": True, "fixed": fixed,
                    "message": "✅ Dono orders ke alag alag customers ban gaye!"})


@bp.route("/api/quick-fix-customers")
@owner_required
def api_quick_fix_customers():
    """Quick GET-based customer fix for specific orders."""
    from datetime import datetime as _dt
    code1, mob1 = request.args.get("c1",""), request.args.get("m1","")
    code2, mob2 = request.args.get("c2",""), request.args.get("m2","")
    if not all([code1, mob1, code2, mob2]):
        return jsonify({"ok": False, "error": "Need c1,m1,c2,m2 params"})

    conn = get_db()
    now_str = _dt.now().strftime("%Y-%m-%d %H:%M:%S")
    results = []

    for code, mob in [(code1, mob1), (code2, mob2)]:
        o = conn.execute("""
            SELECT o.id, o.order_code, o.customer_id,
                   c.name as cname, c.address
            FROM orders o
            LEFT JOIN customers c ON c.id=o.customer_id
            WHERE o.order_code=?
        """, (code,)).fetchone()
        if not o:
            results.append(f"❌ Order #{code} not found")
            continue

        # Check if customer with correct name+mobile already exists
        existing = conn.execute(
            "SELECT id FROM customers WHERE mobile=? AND name=?",
            (mob, o["cname"])
        ).fetchone()

        if existing:
            cust_id = existing["id"]
            results.append(f"✅ #{code} → existing customer id={cust_id}")
        else:
            row = conn.execute(
                "INSERT INTO customers(name, mobile, address, created_at) VALUES(?,?,?,?) RETURNING id",
                (o["cname"], mob, o["address"] or "", now_str)
            ).fetchone()
            cust_id = row["id"]
            results.append(f"✅ #{code} → new customer id={cust_id} created (mobile={mob})")

        conn.execute("UPDATE orders SET customer_id=? WHERE order_code=?", (cust_id, code))

    conn.commit()
    conn.close()
    return jsonify({"ok": True, "results": results,
                    "message": "Dono orders fix ho gaye! Refresh karo diary page."})

@bp.route("/api/fix-order-code")
@owner_required
def fix_order_code():
    from database import invalidate_settings_cache

    # Allow manual override: /owner/api/fix-order-code?set=3922
    manual = request.args.get("set", "").strip()
    if manual and manual.isdigit():
        max_code = int(manual)
        set_setting("last_order_code", str(max_code))
        set_setting("recycled_order_codes", "")
        invalidate_settings_cache()
        return f"<h2>✅ Manually set! last_order_code={max_code}. Next new order = #{max_code+1}</h2><a href='/new-order'>Go to New Order →</a>"

    # Auto-detect: only consider codes <= 9999 (new orders)
    # Past orders often have old diary codes like 34001, 50000 etc.
    # which should NOT affect the new order counter
    conn = get_db()
    rows = conn.execute("SELECT order_code FROM orders").fetchall()
    nums = [int(r["order_code"]) for r in rows
            if str(r["order_code"]).isdigit() and int(r["order_code"]) <= 9999]
    max_code = max(nums) if nums else 3599
    conn.close()

    set_setting("last_order_code", str(max_code))
    set_setting("recycled_order_codes", "")
    invalidate_settings_cache()
    return (f"<h2>✅ Fixed! Highest new order = #{max_code}. Next order = #{max_code+1}</h2>"
            f"<p style='color:#666'>Past order codes (>9999) were ignored.</p>"
            f"<p>To manually set: <a href='/owner/api/fix-order-code?set=3922'>/owner/api/fix-order-code?set=3922</a></p>"
            f"<a href='/new-order'>Go to New Order →</a>")


@bp.route("/api/force-order-code/<int:value>")
@owner_required
def force_order_code(value):
    """Directly set last_order_code to an exact value, bypassing the max-of-all-orders
    auto-detect used by /api/fix-order-code — that logic gets thrown off by a single
    one-off/custom order code sitting far outside the normal sequential range.
    Also lists any orders that exist ABOVE the given value, so you can see what caused
    the auto-detect to jump ahead in the first place."""
    conn = get_db()
    rows = conn.execute("""
        SELECT o.order_code, o.order_date, o.created_at, c.name, c.mobile
        FROM orders o LEFT JOIN customers c ON c.id = o.customer_id
    """).fetchall()
    conn.close()

    outliers = [r for r in rows if str(r["order_code"]).isdigit() and int(r["order_code"]) > value]
    outliers.sort(key=lambda r: int(r["order_code"]), reverse=True)

    set_setting("last_order_code", str(value))
    try:
        from database import invalidate_settings_cache
        invalidate_settings_cache()
    except Exception:
        pass

    rows_html = "".join(
        f"<tr><td>#{r['order_code']}</td><td>{r['name'] or '—'}</td><td>{r['mobile'] or '—'}</td>"
        f"<td>{r['order_date'] or '—'}</td><td>{r['created_at'] or '—'}</td></tr>"
        for r in outliers
    ) or "<tr><td colspan='5' style='text-align:center;color:#888;'>None — no orders above this value.</td></tr>"

    return f"""<h2>✅ last_order_code forcibly set to {value}. Next order = #{value+1}</h2>
    <p>Orders that exist <b>above</b> {value} (this is what was confusing the auto-detect endpoint — check if these are real orders or old test/junk data):</p>
    <table border="1" cellpadding="6" style="border-collapse:collapse;">
    <tr><th>Order Code</th><th>Customer</th><th>Mobile</th><th>Order Date</th><th>Created At</th></tr>
    {rows_html}
    </table>
    <br><a href='/owner/settings'>← Settings</a> | <a href='/new-order'>Go to New Order →</a>"""


@bp.route("/api/db-diagnostics")
@owner_required
def db_diagnostics():
    """Read-only: shows current DB connections/queries and what (if anything) is
    blocking what. Use this if a save/request ever seems to hang indefinitely
    (button stuck on 'Saving...' with no success or error) — it'll show which
    connection is stuck and which other connection is holding it up."""
    conn = get_db()
    try:
        rows = conn.execute("""
            SELECT pid, state,
                   COALESCE(wait_event_type,'-') AS wait_event_type,
                   COALESCE(wait_event,'-') AS wait_event,
                   EXTRACT(EPOCH FROM (NOW() - query_start))::int AS secs_running,
                   pg_blocking_pids(pid) AS blocked_by,
                   LEFT(query, 150) AS query
            FROM pg_stat_activity
            WHERE datname = current_database() AND pid <> pg_backend_pid()
            ORDER BY query_start ASC NULLS LAST
        """).fetchall()
        err = None
    except Exception as e:
        rows, err = [], str(e)
    conn.close()

    if err:
        return f"<h2>DB Diagnostics — query failed</h2><p>{err}</p>"

    def row_html(r):
        blocked = r["blocked_by"]
        is_blocked = blocked not in (None, "", "{}")
        style = "color:#dc2626;font-weight:800;" if is_blocked else ""
        return (f"<tr><td>{r['pid']}</td><td>{r['state']}</td><td>{r['wait_event_type']}</td>"
                f"<td>{r['wait_event']}</td><td>{r['secs_running']}</td>"
                f"<td style='{style}'>{blocked if is_blocked else '-'}</td>"
                f"<td style='font-family:monospace;font-size:11px;'>{r['query']}</td></tr>")

    rows_html = "".join(row_html(r) for r in rows) or \
        "<tr><td colspan='7' style='text-align:center;color:#888;'>No active connections</td></tr>"

    return f"""<h2>DB Diagnostics</h2>
    <p>Rows with a red "Blocked By" PID are stuck waiting on another connection — that PID is the one holding things up.</p>
    <table border="1" cellpadding="6" style="border-collapse:collapse;font-size:13px;">
    <tr><th>PID</th><th>State</th><th>Wait Type</th><th>Wait Event</th><th>Secs Running</th><th>Blocked By</th><th>Query</th></tr>
    {rows_html}
    </table>
    <br><a href='/owner/settings'>← Settings</a>"""


@bp.route("/api/deploy-info")
@owner_required
def deploy_info():
    """Read-only: shows exactly what commit/code is actually running on THIS
    server right now, plus file modification times for key static assets.
    Use this to verify a deploy actually landed instead of trusting GitHub
    Actions' reported status alone (the workflow currently swallows real
    failures with continue-on-error / || true, so 'success' there does not
    guarantee the server actually updated)."""
    import subprocess, os as _os, datetime as _dt
    app_root = _os.path.dirname(_os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))))

    def run(cmd):
        try:
            # Service environment has a restricted PATH without git — use full path
            if cmd[0] == "git":
                cmd = ["/usr/bin/git"] + cmd[1:]
            return subprocess.check_output(cmd, cwd=app_root, stderr=subprocess.STDOUT, timeout=10).decode().strip()
        except Exception as e:
            return f"ERROR: {e}"

    commit_hash = run(["git", "rev-parse", "HEAD"])
    commit_msg  = run(["git", "log", "-1", "--format=%s"])
    commit_date = run(["git", "log", "-1", "--format=%ci"])
    git_status  = run(["git", "status", "--short"])

    def mtime_of(rel_path):
        try:
            full = _os.path.join(app_root, rel_path)
            ts = _os.path.getmtime(full)
            return _dt.datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S")
        except Exception as e:
            return f"ERROR: {e}"

    files_html = "".join(
        f"<tr><td>{f}</td><td>{mtime_of(f)}</td></tr>"
        for f in ["static/css/main.css", "templates/base.html", "templates/employee/order_status.html", "run.py"]
    )

    # Last deploy log written by the GitHub Actions workflow (tee on the server)
    try:
        with open("/home/ubuntu/last_deploy.log") as lf:
            deploy_log = lf.read()[-4000:]
    except Exception as e:
        deploy_log = f"(no deploy log yet: {e})"

    return f"""<h2>Deploy Info — what's ACTUALLY running right now</h2>
    <p><b>Server time:</b> {_dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</p>
    <p><b>Git commit:</b> {commit_hash}</p>
    <p><b>Commit message:</b> {commit_msg}</p>
    <p><b>Commit date:</b> {commit_date}</p>
    <p><b>Uncommitted local changes (should be empty):</b> <pre>{git_status or '(none)'}</pre></p>
    <h3>File modification times on disk</h3>
    <table border="1" cellpadding="6" style="border-collapse:collapse;">
    <tr><th>File</th><th>Last Modified</th></tr>
    {files_html}
    </table>
    <h3>Last deploy log</h3>
    <pre style="background:#111;color:#0f0;padding:12px;border-radius:8px;font-size:12px;overflow-x:auto;">{deploy_log}</pre>
    <br><a href='/owner/settings'>← Settings</a>"""


@bp.route("/api/sync-order-images")
@owner_required
def sync_order_images():
    """Backfill order_images table from filesystem for all existing orders."""
    import os as _os
    img_base = _os.path.join(_os.path.dirname(_os.path.dirname(_os.path.dirname(__file__))), "static", "order_images")
    conn = get_db()
    synced = 0
    if _os.path.isdir(img_base):
        for code in _os.listdir(img_base):
            folder = _os.path.join(img_base, code)
            if not _os.path.isdir(folder): continue
            order_row = conn.execute("SELECT id FROM orders WHERE order_code=?", (code,)).fetchone()
            if not order_row: continue
            order_id = order_row["id"]
            imgs = sorted(f for f in _os.listdir(folder)
                          if f.lower().endswith((".jpg",".jpeg",".png",".webp")) and not f.startswith("."))
            for img in imgs:
                fp = f"/static/order_images/{code}/{img}"
                exists = conn.execute("SELECT id FROM order_images WHERE order_id=? AND file_path=?",
                                      (order_id, fp)).fetchone()
                if not exists:
                    conn.execute("INSERT INTO order_images(order_id, file_path) VALUES(?,?)", (order_id, fp))
                    synced += 1
        conn.commit()
    conn.close()
    return f"<h2>✅ {synced} images synced to DB!</h2><p>Ab Diary mein images dikhenge.</p><a href='/owner/measurement-book'>Open Diary →</a>"


@bp.route("/measurement-book")
def measurement_book():
    import os as _os, json as _json

    def fmtd(d):
        if not d: return "—"
        p = str(d).split("-")
        return f"{p[2]}-{p[1]}-{p[0]}" if len(p)==3 else d

    conn = get_db()
    try:
        uc = conn.execute("SELECT COUNT(*) as c FROM orders WHERE is_urgent=1 AND status!='delivered'").fetchone()["c"]
        total_orders = conn.execute("SELECT COUNT(*) as c FROM orders").fetchone()["c"]
        db_codes = set(r["order_code"] for r in conn.execute("SELECT order_code FROM orders").fetchall())

        # Recent already-filled-in orders (one card per customer, their latest
        # order) so the Diary isn't empty until you search — old/manually
        # entered orders show here too, not just fresh QR uploads.
        # Capped at 40 to avoid the old thousands-of-orders perf issue.
        recent_rows = conn.execute("""
            SELECT * FROM (
                SELECT DISTINCT ON (o.customer_id)
                       o.id, o.order_code, o.repeat_of, o.order_date, o.delivery_date, o.status,
                       o.payable_amount, o.advance_paid, o.remaining, o.note, o.is_urgent,
                       c.name as cname, c.mobile, c.address
                FROM (SELECT * FROM orders ORDER BY id DESC LIMIT 200) o
                JOIN customers c ON c.id = o.customer_id
                ORDER BY o.customer_id, o.id DESC
            ) sub
            ORDER BY sub.id DESC LIMIT 40
        """).fetchall()

        recent_ids = [r["id"] for r in recent_rows]
        garments_by_order, image_by_order = {}, {}
        if recent_ids:
            ph = ",".join("?" * len(recent_ids))
            for g in conn.execute(
                f"SELECT order_id, garment_type, quantity, rate, notes, measurements FROM order_items WHERE order_id IN ({ph}) ORDER BY order_id, id",
                recent_ids
            ).fetchall():
                try:
                    meas = _json.loads(g["measurements"] or "{}")
                except Exception:
                    meas = {}
                garments_by_order.setdefault(g["order_id"], []).append({
                    "type": g["garment_type"], "qty": g["quantity"],
                    "rate": int(g["rate"] or 0),
                    "notes": g["notes"] or "", "meas": meas
                })
            for r in conn.execute(
                f"SELECT order_id, file_path FROM order_images WHERE order_id IN ({ph}) AND file_path NOT LIKE ? ORDER BY order_id, id",
                recent_ids + ["temp:%"]
            ).fetchall():
                if r["order_id"] not in image_by_order and r["file_path"]:
                    image_by_order[r["order_id"]] = r["file_path"]
    finally:
        conn.close()

    # Initial page: pending QR image-only cards (usually a handful) FIRST —
    # they need action — followed by recent already-filled-in orders below.
    # Previously this route rendered a full card for EVERY order in the DB
    # (~thousands, hidden with display:none, searched client-side) — a 10-20MB
    # HTML response that effectively never finished loading on mobile.
    # Search is now server-side via /owner/api/diary-search for anything
    # beyond this recent set.
    orders_data = []
    img_base = _os.path.join(_os.path.dirname(_os.path.dirname(_os.path.dirname(__file__))), "static", "order_images")
    if _os.path.isdir(img_base):
        # Only scan last 100 folders (most recent) — avoids slow scan of 1000s of folders
        all_codes = sorted(
            (c for c in _os.listdir(img_base) if c.isdigit()),
            key=lambda x: int(x), reverse=True
        )[:100]
        for code in all_codes:
            if code in db_codes:
                continue
            folder = _os.path.join(img_base, code)
            if not _os.path.isfile(_os.path.join(folder, ".diary_upload")):
                continue
            imgs = [f for f in _os.listdir(folder)
                    if f.lower().endswith((".jpg",".jpeg",".png",".webp"))]
            if not imgs:
                continue
            orders_data.append({
                "code": code, "odate": "—", "ddate": "—",
                "status": "image_only", "urgent": False,
                "payable": 0, "paid": 0, "due": 0,
                "note": "", "cname": "— Tap to fill details —",
                "mobile": "—", "address": "—",
                "garments": [],
                "image": f"/static/order_images/{code}/{imgs[0]}",
                "image_only": True,
            })

    for o in recent_rows:
        orders_data.append({
            "code": o["order_code"],
            "display_code": o["repeat_of"] if o["repeat_of"] else o["order_code"],
            "entry_code": o["order_code"] if o["repeat_of"] else "",
            "odate": fmtd(o["order_date"]),
            "ddate": fmtd(o["delivery_date"]), "status": o["status"],
            "urgent": bool(o["is_urgent"]),
            "payable": int(o["payable_amount"] or 0),
            "paid":    int(o["advance_paid"]  or 0),
            "due":     int(o["remaining"]     or 0),
            "note":    o["note"] or "",
            "cname":   o["cname"]   or "—",
            "mobile":  o["mobile"]  or "—",
            "address": o["address"] or "—",
            "garments": garments_by_order.get(o["id"], []),
            "image": image_by_order.get(o["id"]),
            "image_only": False,
        })

    return render_template("owner/measurement_book.html",
        active_page="measurement_book", show_voice=False,
        urgent_count=uc, orders=orders_data, total_orders=total_orders)


@bp.route("/api/diary-search")
def api_diary_search():
    """Server-side Diary search. Returns rendered card HTML for matching orders
    (word-AND across name/mobile/address/order code/garment type), capped at 60."""
    import json as _json
    q = (request.args.get("q") or "").strip()
    if not q:
        return ""

    def fmtd(d):
        if not d: return "—"
        p = str(d).split("-")
        return f"{p[2]}-{p[1]}-{p[0]}" if len(p)==3 else d

    words = [w.lower() for w in q.split() if w][:6]
    if not words:
        return ""

    word_clauses = " AND ".join(
        """(LOWER(c.name) LIKE ?
           OR c.mobile LIKE ?
           OR LOWER(c.address) LIKE ?
           OR o.order_code LIKE ?
           OR EXISTS (
               SELECT 1 FROM order_items oi
               WHERE oi.order_id=o.id AND LOWER(oi.garment_type) LIKE ?
           ))"""
        for _ in words
    )
    word_params = []
    for w in words:
        lk = f"%{w}%"
        word_params.extend([lk, lk, lk, f"%{w.lstrip('#')}%", lk])

    conn = get_db()
    try:
        rows = None

        # Numeric query that exactly matches an order code → show ONLY that
        # order. (Typing "2890" was also substring-matching every mobile number
        # containing 2890, burying the actual order.) If no order has that
        # exact code, fall through to normal search so typing the beginning
        # of a mobile number still works.
        qnum = q.lstrip("#")
        if qnum.isdigit():
            exact = conn.execute("""
                SELECT o.id, o.order_code, o.order_date, o.delivery_date, o.status,
                       o.payable_amount, o.advance_paid, o.remaining, o.note, o.is_urgent,
                       o.customer_id, c.name as cname, c.mobile, c.address
                FROM orders o JOIN customers c ON c.id=o.customer_id
                WHERE o.order_code = ?
            """, (qnum,)).fetchall()
            if exact:
                # Also pull this customer's other orders (old + new) so the
                # whole history shows together, matched by customer/mobile —
                # not just the one order that was searched for.
                cust_id = exact[0]["customer_id"]
                rows = conn.execute("""
                    SELECT o.id, o.order_code, o.order_date, o.delivery_date, o.status,
                           o.payable_amount, o.advance_paid, o.remaining, o.note, o.is_urgent,
                           o.repeat_of, o.customer_id, c.name as cname, c.mobile, c.address
                    FROM orders o JOIN customers c ON c.id=o.customer_id
                    WHERE o.customer_id = ?
                    ORDER BY o.id DESC
                """, (cust_id,)).fetchall()
            else:
                rows = None

        if rows is None:
            # DISTINCT ON (customer_id): only each customer's LATEST order — the
            # diary is a measurements lookup, so one card per person (their current
            # naap) instead of flooding results with the person's whole history.
            # Customers sharing a mobile (family members) stay separate cards.
            rows = conn.execute(f"""
                SELECT * FROM (
                    SELECT DISTINCT ON (o.customer_id)
                           o.id, o.order_code, o.order_date, o.delivery_date, o.status,
                           o.payable_amount, o.advance_paid, o.remaining, o.note, o.is_urgent,
                           o.repeat_of, o.customer_id, c.name as cname, c.mobile, c.address
                    FROM orders o JOIN customers c ON c.id=o.customer_id
                    WHERE {word_clauses}
                    ORDER BY o.customer_id, o.id DESC
                ) sub
                ORDER BY sub.id DESC LIMIT 60
            """, word_params).fetchall()

        ids = [r["id"] for r in rows]
        garments_by_order, image_by_order = {}, {}
        if ids:
            ph = ",".join("?" * len(ids))
            for g in conn.execute(
                f"SELECT order_id, garment_type, quantity, rate, notes, measurements FROM order_items WHERE order_id IN ({ph}) ORDER BY order_id, id",
                ids
            ).fetchall():
                try:
                    meas = _json.loads(g["measurements"] or "{}")
                except Exception:
                    meas = {}
                garments_by_order.setdefault(g["order_id"], []).append({
                    "type": g["garment_type"], "qty": g["quantity"],
                    "rate": int(g["rate"] or 0),
                    "notes": g["notes"] or "", "meas": meas
                })
            for r in conn.execute(
                f"SELECT order_id, file_path FROM order_images WHERE order_id IN ({ph}) AND file_path NOT LIKE ? ORDER BY order_id, id",
                ids + ["temp:%"]
            ).fetchall():
                if r["order_id"] not in image_by_order and r["file_path"]:
                    image_by_order[r["order_id"]] = r["file_path"]
    finally:
        conn.close()

    # Mark the newest order per customer so the "Latest" badge and the
    # original-vs-entry code split work the same way as Order Status/Pickup.
    latest_id_per_cust = {}
    for r in rows:
        cid = r["customer_id"]
        if cid not in latest_id_per_cust or r["id"] > latest_id_per_cust[cid]:
            latest_id_per_cust[cid] = r["id"]

    orders_data = [{
        "code": o["order_code"], "odate": fmtd(o["order_date"]),
        "ddate": fmtd(o["delivery_date"]), "status": o["status"],
        "urgent": bool(o["is_urgent"]),
        "payable": int(o["payable_amount"] or 0),
        "paid":    int(o["advance_paid"]  or 0),
        "due":     int(o["remaining"]     or 0),
        "note":    o["note"] or "",
        "cname":   o["cname"]   or "—",
        "mobile":  o["mobile"]  or "—",
        "address": o["address"] or "—",
        "garments": garments_by_order.get(o["id"], []),
        "image": image_by_order.get(o["id"]),
        "image_only": False,
        "display_code": o["repeat_of"] if o["repeat_of"] else o["order_code"],
        "entry_code":   o["order_code"] if o["repeat_of"] else "",
        "is_latest":    latest_id_per_cust.get(o["customer_id"]) == o["id"],
    } for o in rows]

    return render_template("owner/_diary_cards.html", orders=orders_data)




@bp.route("/settings")
@owner_required
def settings():
    conn = get_db()
    today = date.today().isoformat()
    urgent_count = conn.execute("SELECT COUNT(*) as c FROM orders WHERE is_urgent=1 AND status != 'delivered' AND delivery_date >= ?",(today,)).fetchone()["c"]
    conn.close()
    current_settings = {
        "shop_name":        get_setting("shop_name","Uttam Tailors"),
        "shop_name_hi":     get_setting("shop_name_hi","उत्तम टेलर्स"),
        "whatsapp_number":  get_setting("whatsapp_number",""),
        "owner_pin":        get_setting("owner_pin","1234"),
        "default_language": get_setting("default_language","hinglish"),
        "work_rate_measurement": get_setting("work_rate_measurement","0"),
        "work_rate_cutting":     get_setting("work_rate_cutting","25"),
        "work_rate_alteration":  get_setting("work_rate_alteration","15"),
        "work_rate_stitching":   get_setting("work_rate_stitching",""),
        "salary_fresh_start_date": get_setting("salary_fresh_start_date",""),
        "utms_fresh_start":      get_setting("utms_fresh_start","0"),
        "utms_fresh_start_date": get_setting("utms_fresh_start_date",""),
        "finance_income_cats":   get_setting("finance_income_cats","advance,payment,alteration,other income"),
        "finance_expense_cats":  get_setting("finance_expense_cats","thread,buttons,fabric,electricity,rent,salary,transport,maintenance,other expense"),
        "shop_logo":             get_setting("shop_logo",""),
        "order_code_start":      str(int(get_setting("last_order_code","3898")) + 1),
     
    }
    # Garment chip styles
    garment_type_chips_raw = get_setting("garment_type_chips","")
    try:
        import json as _json
        garment_type_chips = _json.loads(garment_type_chips_raw) if garment_type_chips_raw else {}
    except Exception:
        garment_type_chips = {}

    # Garment rates
    garment_names = [
        "Shirt","Shirt Linen","Pant","Pant Double","Jeans","Suit 2pc","Suit 3pc",
        "Blazer","Kurta","Kurta Pajama","Pajama","Pathani","Sherwani","Safari","Waistcoat",
        "Alteration","Cutting Only"
    ]
    deleted_csv = get_setting("deleted_customer_rates","")
    deleted_set = set(x.strip() for x in deleted_csv.split(",") if x.strip())
    garment_rates = {}
    for n in garment_names:
        if n in deleted_set: continue
        garment_rates[n] = get_setting("customer_rate_"+n,"") or get_setting("rate_"+n,"0")

    # Work rates map
    work_rates_map = {
        "work_rate_measurement": get_setting("work_rate_measurement","0"),
        "work_rate_cutting":     get_setting("work_rate_cutting","25"),
        "work_rate_alteration":  get_setting("work_rate_alteration","15"),
    }
    # Stitch rates
    stitch_rates = {n: get_setting("stitch_rate_"+n,"0") for n in garment_names}

    return render_template("owner/settings.html",
        active_page="settings", show_voice=False,
        urgent_count=urgent_count,
        settings=current_settings,
        current_settings=current_settings,
        garment_type_chips=garment_type_chips,
        garment_rates=garment_rates,
        stitch_rates=stitch_rates,
        work_rates_map=work_rates_map,
        last_backup=get_setting("last_backup_at",""))


@bp.route("/settings/save", methods=["POST"])
@owner_required
def settings_save():
    section = request.form.get("section")
    if section == "shop":
        # Handle logo upload
        import base64
        logo_file = request.files.get("shop_logo")
        if logo_file and logo_file.filename:
            data = logo_file.read()
            ext = logo_file.filename.rsplit(".",1)[-1].lower()
            b64 = base64.b64encode(data).decode()
            set_setting("shop_logo", f"data:image/{ext};base64,{b64}")
            flash("Logo updated!", "success")
        set_setting("shop_name",        request.form.get("shop_name","").strip())
        set_setting("shop_name_hi",     request.form.get("shop_name_hi","").strip())
        set_setting("whatsapp_number",  request.form.get("whatsapp_number","").strip())
        set_setting("default_language", request.form.get("default_language","hl"))
        # Order code start - only update last_order_code if value given and valid
        new_code = request.form.get("order_code_start","").strip()
        if new_code.isdigit():
            set_setting("last_order_code", str(int(new_code) - 1))  # next call will increment to this
        from database import invalidate_settings_cache; invalidate_settings_cache()
        flash("Shop settings saved!", "success")
    elif section == "pin":
        current = request.form.get("current_pin","")
        new_pin = request.form.get("new_pin","")
        confirm = request.form.get("confirm_pin","")
        real_pin = get_setting("owner_pin","1234")
        if current != real_pin:
            flash("Current PIN is wrong.", "error")
        elif len(new_pin) != 4 or not new_pin.isdigit():
            flash("New PIN must be exactly 4 digits.", "error")
        elif new_pin != confirm:
            flash("PINs do not match.", "error")
        else:
            set_setting("owner_pin", new_pin)
            flash("PIN changed successfully!", "success")
    elif section == "customer_rates":
        # Load currently deleted garments list
        deleted_csv = get_setting("deleted_customer_rates", "")
        deleted_names = set(x.strip() for x in deleted_csv.split(",") if x.strip())
        # Process new deletes
        for name in request.form.getlist("delete_rate"):
            deleted_names.add(name)
            conn2 = get_db()
            conn2.execute("DELETE FROM settings WHERE key=?",("customer_rate_"+name,))
            conn2.execute("DELETE FROM settings WHERE key=?",("rate_"+name,))
            conn2.commit(); conn2.close()
        # Save deleted list
        set_setting("deleted_customer_rates", ",".join(sorted(deleted_names)))
        from database import invalidate_settings_cache; invalidate_settings_cache()
        # Then: save remaining rates (skip deleted ones)
        for key, val in request.form.items():
            if key.startswith("customer_rate_"):
                name = key[14:]
                if name in deleted_names:
                    continue  # skip re-saving deleted items
                if val.strip():
                    set_setting(key, val)
                    # Also sync to rate_X so new_order picks it up
                    set_setting("rate_"+name, val)
                    # If user re-adds a previously deleted garment, remove from deleted list
                    deleted_names.discard(name)
        set_setting("deleted_customer_rates", ",".join(sorted(deleted_names)))
        from database import invalidate_settings_cache; invalidate_settings_cache()
        flash("Customer rates saved!", "success")
    elif section == "stitch_rates":
        for key, val in request.form.items():
            if key.startswith("stitch_rate_") and val.strip():
                set_setting(key, val)
        from database import invalidate_settings_cache; invalidate_settings_cache()
        flash("Stitching rates saved!", "success")
    elif section == "rate_image":
        import base64
        img = request.files.get("rate_list_image")
        if img and img.filename:
            data = img.read()
            ext = img.filename.rsplit(".",1)[-1].lower()
            b64 = base64.b64encode(data).decode()
            set_setting("rate_list_image", f"data:image/{ext};base64,{b64}")
            flash("Rate list image uploaded!", "success")
    elif section == "add_employee":
        name  = request.form.get("emp_name","").strip()
        phone = request.form.get("emp_phone","").strip()
        if name:
            conn2 = get_db()
            try:
                conn2.execute("INSERT INTO employees(name,phone) VALUES(?,?)", (name, phone))
                conn2.commit()
                flash(f"Employee '{name}' added!", "success")
            except:
                flash("Employee name already exists", "warning")
            conn2.close()
    elif section == "remove_employee":
        emp_id = request.form.get("emp_id","")
        if emp_id:
            conn2 = get_db()
            conn2.execute("UPDATE employees SET active=0 WHERE id=?", (emp_id,))
            conn2.commit(); conn2.close()
            flash("Employee removed.", "success")
    elif section == "rates":
        for name in request.form.getlist("delete_rate"):
            conn2 = get_db()
            conn2.execute("DELETE FROM settings WHERE key=?",("rate_"+name,))
            conn2.commit(); conn2.close()
        for key, val in request.form.items():
            if key.startswith("rate_") and val.strip():
                set_setting(key, val)
        from database import invalidate_settings_cache; invalidate_settings_cache()
        flash("Rates saved!", "success")
    return redirect(url_for("owner.settings"))

@bp.route("/api/owner/earnings-7days")
@owner_required
def earnings_7days():
    conn = get_db()
    labels, income_data, expense_data = [], [], []
    for i in range(6,-1,-1):
        d = (date.today() - timedelta(days=i)).isoformat()
        labels.append(d[5:])
        rows = conn.execute("SELECT tx_type, SUM(amount) as total FROM finance WHERE tx_date=? GROUP BY tx_type",(d,)).fetchall()
        fin = {r["tx_type"]: r["total"] or 0 for r in rows}
        income_data.append(fin.get("income",0))
        expense_data.append(fin.get("expense",0))
    conn.close()
    return jsonify({"labels":labels,"income":income_data,"expense":expense_data})

@bp.route("/api/settings/logo")
def api_logo():
    return jsonify({"value": get_setting("shop_logo","")})

@bp.route("/measurement-fields")
@owner_required
def measurement_fields():
    conn = get_db()
    today = date.today().isoformat()
    urgent_count = conn.execute("SELECT COUNT(*) as c FROM orders WHERE is_urgent=1 AND status!='delivered' AND delivery_date>=?",(today,)).fetchone()["c"]
    # Get all garment types
    garment_types = [
        "Shirt","Shirt Linen","Pant","Pant Double","Jeans","Suit 2pc","Suit 3pc",
        "Blazer","Kurta","Kurta Pajama","Pajama","Pathani","Sherwani","Safari","Waistcoat",
        "Alteration","Cutting Only"
    ]
    # Also get any custom ones from DB
    extra = conn.execute("SELECT DISTINCT garment_type FROM measurement_fields WHERE garment_type NOT IN ({})".format(
        ",".join("?"*len(garment_types))), garment_types).fetchall()
    garment_types += [r["garment_type"] for r in extra]

    fields_by_garment = {}
    rows = conn.execute("SELECT garment_type,field_name,id FROM measurement_fields ORDER BY sort_order ASC,id ASC").fetchall()
    for r in rows:
        fields_by_garment.setdefault(r["garment_type"],[]).append({"id":r["id"],"name":r["field_name"]})
    conn.close()
    return render_template("owner/measurement_fields.html",
        active_page="settings", show_voice=False, urgent_count=urgent_count,
        garment_types=garment_types, fields_by_garment=fields_by_garment)

@bp.route("/measurement-fields/add", methods=["POST"])
@owner_required
def add_measurement_field():
    garment = request.form.get("garment_type","").strip()
    field   = request.form.get("field_name","").strip()
    if garment and field:
        conn = get_db()
        conn.execute("INSERT INTO measurement_fields (garment_type,field_name,sort_order) VALUES (?,?,99) ON CONFLICT DO NOTHING",(garment,field))
        conn.commit(); conn.close()
        flash(f"Field added to {garment}!", "success")
    return redirect(url_for("owner.measurement_fields"))

@bp.route("/measurement-fields/delete/<int:fid>")
@owner_required
def delete_measurement_field(fid):
    conn = get_db()
    conn.execute("DELETE FROM measurement_fields WHERE id=?",(fid,))
    conn.commit(); conn.close()
    flash("Field removed.", "success")
    return redirect(url_for("owner.measurement_fields"))

@bp.route("/measurement-fields/reorder", methods=["POST"])
@owner_required
def reorder_measurement_fields():
    """Save drag-and-drop order. Expects JSON: { ids: [1, 4, 2, ...] }"""
    import json as _json2
    data = request.get_json(silent=True) or {}
    ids  = data.get("ids", [])
    if not ids:
        return ("", 204)
    conn = get_db()
    for position, fid in enumerate(ids):
        conn.execute("UPDATE measurement_fields SET sort_order=? WHERE id=?", (position, fid))
    conn.commit()
    conn.close()
    return ("", 204)


# ══════════════════════════════════════════════
#  EXCEL EXPORT & IMPORT
# ══════════════════════════════════════════════

@bp.route("/export/orders")
@owner_required
def export_orders():
    """Export all orders with every field to Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    import json as _json

    conn = get_db()
    orders = conn.execute("""
        SELECT o.*, c.name as cname, c.mobile, c.address
        FROM orders o LEFT JOIN customers c ON c.id=o.customer_id
        ORDER BY o.id DESC
    """).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders"

    # Header style
    hdr_fill = PatternFill("solid", fgColor="6366F1")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)

    headers = [
        "Order Code", "Customer Name", "Mobile", "Address",
        "Order Date", "Delivery Date", "Status", "Is Urgent",
        "Garments", "Measurements",
        "Total Amount", "Extra Charges", "Payable Amount",
        "Advance Paid", "Remaining Due", "Payment Mode",
        "Repeat Of", "Note", "Delivered At", "Created At"
    ]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"

    def fmtd(d):
        if not d: return ""
        p = str(d).split("-")
        return f"{p[2]}-{p[1]}-{p[0]}" if len(p)==3 else d

    for row_idx, o in enumerate(orders, 2):
        items = conn.execute(
            "SELECT garment_type, quantity, rate, amount, measurements FROM order_items WHERE order_id=?",
            (o["id"],)
        ).fetchall()

        garments_str = "; ".join(f"{i['garment_type']} x{i['quantity']} @₹{int(i['rate'])}" for i in items)
        meas_parts = []
        for it in items:
            try:
                m = _json.loads(it["measurements"] or "{}")
                if m:
                    meas_parts.append(f"{it['garment_type']}: " + ", ".join(f"{k}={v}" for k,v in m.items()))
            except: pass
        meas_str = "; ".join(meas_parts)

        delivered_at = ""
        try: delivered_at = fmtd((o["delivered_at"] or "")[:10])
        except: pass

        row = [
            o["order_code"], o["cname"] or "", o["mobile"] or "", o["address"] or "",
            fmtd(o["order_date"]), fmtd(o["delivery_date"]),
            o["status"], "Yes" if o["is_urgent"] else "No",
            garments_str, meas_str,
            o["total_amount"] or 0, o["extra_charges"] or 0, o["payable_amount"] or 0,
            o["advance_paid"] or 0, o["remaining"] or 0, o["payment_mode"] or "",
            o["repeat_of"] or "", o["note"] or "", delivered_at,
            (o["created_at"] or "")[:16]
        ]
        for col, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col, value=val)
        # Highlight urgent in red
        if o["is_urgent"]:
            for col in range(1, len(headers)+1):
                ws.cell(row=row_idx, column=col).fill = PatternFill("solid", fgColor="FEE2E2")

    # Auto column widths
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

    conn.close()
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    from flask import send_file
    from datetime import datetime as _dt
    fname = f"uttam_tailors_orders_{_dt.now().strftime('%d-%m-%Y')}.xlsx"
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=fname)


@bp.route("/import/customers", methods=["POST"])
@owner_required
def import_customers():
    """Import customers from Excel sheet."""
    import openpyxl
    from io import BytesIO
    f = request.files.get("customer_file")
    if not f:
        flash("No file selected", "warning")
        return redirect(url_for("owner.settings"))

    try:
        wb = openpyxl.load_workbook(BytesIO(f.read()))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            flash("File is empty", "warning")
            return redirect(url_for("owner.settings"))

        # Try to detect header row
        header = [str(c or "").lower().strip() for c in rows[0]]
        name_col   = next((i for i,h in enumerate(header) if "name" in h), None)
        mobile_col = next((i for i,h in enumerate(header) if "mobile" in h or "phone" in h), None)
        addr_col   = next((i for i,h in enumerate(header) if "address" in h or "addr" in h), None)

        if name_col is None:
            flash("Could not find 'Name' column in Excel file", "error")
            return redirect(url_for("owner.settings"))

        conn = get_db()
        added = skipped = 0
        from datetime import datetime as _dt
        now = _dt.now().strftime("%Y-%m-%d %H:%M:%S")

        for row in rows[1:]:
            name   = str(row[name_col] or "").strip() if name_col is not None else ""
            mobile = str(row[mobile_col] or "").strip() if mobile_col is not None else ""
            addr   = str(row[addr_col] or "").strip() if addr_col is not None else ""
            if not name:
                continue
            existing = conn.execute("SELECT id FROM customers WHERE name=? AND mobile=?", (name, mobile)).fetchone()
            if existing:
                skipped += 1
            else:
                conn.execute("INSERT INTO customers(name,mobile,address,created_at) VALUES(?,?,?,?)",
                             (name, mobile, addr, now))
                added += 1

        conn.commit(); conn.close()
        flash(f"Import complete: {added} customers added, {skipped} already existed.", "success")
    except Exception as e:
        flash(f"Import error: {str(e)}", "error")

    return redirect(url_for("owner.settings"))


# ══════════════════════════════════════════════
#  INVENTORY MODULE
# ══════════════════════════════════════════════

@bp.route("/inventory")
@owner_required
def inventory():
    conn = get_db()
    today = date.today().isoformat()
    items = conn.execute("SELECT * FROM inventory ORDER BY item_name").fetchall()
    urgent_count = conn.execute(
        "SELECT COUNT(*) as c FROM orders WHERE is_urgent=1 AND status!='delivered'"
    ).fetchone()["c"]
    conn.close()
    return render_template("owner/inventory.html",
        active_page="inventory", show_voice=False,
        urgent_count=urgent_count, items=items)


@bp.route("/inventory/save", methods=["POST"])
@owner_required
def inventory_save():
    name      = request.form.get("item_name","").strip()
    quantity  = int(request.form.get("quantity",0) or 0)
    unit      = request.form.get("unit","").strip()
    threshold = int(request.form.get("low_threshold",0) or 0)
    if name:
        conn = get_db()
        # Upsert
        existing = conn.execute("SELECT id FROM inventory WHERE item_name=?", (name,)).fetchone()
        if existing:
            conn.execute("UPDATE inventory SET quantity=?, unit=?, low_threshold=? WHERE item_name=?",
                        (quantity, unit, threshold, name))
        else:
            conn.execute("INSERT INTO inventory(item_name,quantity,unit,low_threshold) VALUES(?,?,?,?)",
                        (name, quantity, unit, threshold))
        conn.commit(); conn.close()
        flash(f"'{name}' saved!", "success")
    return redirect(url_for("owner.inventory"))


@bp.route("/inventory/delete/<int:item_id>", methods=["POST"])
@owner_required
def inventory_delete(item_id):
    conn = get_db()
    conn.execute("DELETE FROM inventory WHERE id=?", (item_id,))
    conn.commit(); conn.close()
    flash("Item deleted", "success")
    return redirect(url_for("owner.inventory"))


# ══════════════════════════════════════════════
#  OWNER CUSTOMERS MODULE
# ══════════════════════════════════════════════

@bp.route("/customers")
@owner_required
def owner_customers():
    conn = get_db()
    try:
        rows = conn.execute("""
            SELECT c.id, c.name, COALESCE(c.mobile,'') as mobile,
                   COALESCE(c.address,'') as address,
                   COUNT(o.id) as order_count,
                   COALESCE(SUM(o.payable_amount),0) as total_billed,
                   COALESCE(SUM(o.remaining),0) as total_due,
                   MAX(o.order_date) as last_order_date
            FROM customers c LEFT JOIN orders o ON o.customer_id=c.id
            GROUP BY c.id, c.name, c.mobile, c.address ORDER BY c.id DESC
        """).fetchall()
    except Exception as e:
        conn.close()
        return f"<h2>DB Error in /owner/customers</h2><pre>{e}</pre>", 500

    def fmtd(d):
        if not d: return "—"
        p = str(d).split("-")
        return f"{p[2]}-{p[1]}-{p[0]}" if len(p)==3 else d

    # Get all order codes per customer in one query
    all_codes = conn.execute(
        "SELECT customer_id, order_code FROM orders ORDER BY id DESC"
    ).fetchall()
    codes_by_cust = {}
    for row in all_codes:
        codes_by_cust.setdefault(row["customer_id"], []).append(row["order_code"])

    customers = [{
        "id":             r["id"],
        "name":           r["name"],
        "mobile":         r["mobile"] or "",
        "address":        r["address"] or "",
        "order_count":    r["order_count"],
        "total_billed":   r["total_billed"] or 0,
        "total_due":      r["total_due"] or 0,
        "last_order_date":fmtd(r["last_order_date"]),
        "order_codes":    " ".join(codes_by_cust.get(r["id"], []))
    } for r in rows]
    urgent_count = conn.execute(
        "SELECT COUNT(*) as c FROM orders WHERE is_urgent=1 AND status!='delivered'"
    ).fetchone()["c"]
    conn.close()

    return render_template_string(CUSTOMERS_PAGE,
        active_page="customers", show_voice=False,
        urgent_count=urgent_count, customers=customers, total=len(customers))


# ══════════════════════════════════════════════
#  OWNER FINANCE MODULE
# ══════════════════════════════════════════════


@bp.route("/api/diagnose-ready/<code>")
@owner_required
def api_diagnose_ready(code):
    """Read-only: show exactly what check_and_auto_ready sees for this order —
    required qty per garment, how each work_log note got classified into
    naap/kataai/silai, and whether all_done evaluates true. Used to see
    exactly why an order isn't becoming 'ready' automatically."""
    conn = get_db()
    order = conn.execute("SELECT id, order_code, status FROM orders WHERE order_code=?", (code,)).fetchone()
    if not order:
        conn.close()
        return jsonify({"ok": False, "error": f"Order #{code} not found"})

    required = {}
    for r in conn.execute(
        "SELECT garment_type, SUM(quantity) as total FROM order_items WHERE order_id=? GROUP BY garment_type",
        (order["id"],)
    ).fetchall():
        required[r["garment_type"]] = r["total"]

    all_logs = conn.execute(
        "SELECT garment_type, notes, COALESCE(SUM(qty_done),0) as total FROM work_logs WHERE order_code=? GROUP BY garment_type, notes",
        (code,)
    ).fetchall()

    naap, kataai, silai = {}, {}, {}
    raw_logs = []
    for r in all_logs:
        gt, n, qty = r["garment_type"], (r["notes"] or "").strip(), r["total"] or 0
        if any(x in n for x in ["Measurement","Naap","नाप"]):
            bucket = "naap"; naap[gt] = naap.get(gt,0) + qty
        elif any(x in n for x in ["Kataai","Cutting","कटाई"]):
            bucket = "kataai"; kataai[gt] = kataai.get(gt,0) + qty
        else:
            bucket = "silai"; silai[gt] = silai.get(gt,0) + qty
        raw_logs.append({"garment_type": gt, "notes": n, "qty": qty, "classified_as": bucket})

    all_done = all(
        naap.get(gt,0) >= qty and kataai.get(gt,0) >= qty and silai.get(gt,0) >= qty
        for gt, qty in required.items()
    ) if required else False

    conn.close()
    return jsonify({
        "ok": True,
        "order_code": order["order_code"],
        "current_status": order["status"],
        "required_qty_per_garment": required,
        "raw_work_logs": raw_logs,
        "naap_totals": naap,
        "kataai_totals": kataai,
        "silai_totals": silai,
        "all_done": all_done,
    })


@bp.route("/api/recheck-ready-status")
@owner_required
def api_recheck_ready_status():
    """Re-run the (now-fixed) auto-ready check against every currently
    'pending' order. Needed because the auto-ready check only used to
    recognise English work-log labels (Naap/Cutting) and missed Hindi ones
    (नाप/कटाई) — so orders that were actually 100% done stayed stuck as
    'pending' until something new was logged against them. This catches
    those already-stuck orders retroactively. Only touches status; nothing
    else about the order is changed."""
    from app.routes.employee import check_and_auto_ready
    conn = get_db()
    pending = conn.execute("SELECT order_code FROM orders WHERE status='pending'").fetchall()
    fixed = []
    for r in pending:
        if check_and_auto_ready(conn, r["order_code"]):
            fixed.append(r["order_code"])
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "checked": len(pending), "fixed_count": len(fixed), "fixed_orders": fixed})


@bp.route("/auto-bypass-fix")
@owner_required
def auto_bypass_fix_page():
    """List every work-log entry currently credited to 'AUTO-BYPASS' (from
    Force Mark as Ready) so the owner can reassign each one to the real
    employee who actually did the work — for correct pay/stats tracking."""
    conn = get_db()
    rows = conn.execute("""
        SELECT wl.id, wl.order_code, wl.garment_type, wl.qty_done, wl.notes,
               wl.log_date, wl.created_at, c.name as cname
        FROM work_logs wl
        LEFT JOIN orders o ON o.id = wl.order_id
        LEFT JOIN customers c ON c.id = o.customer_id
        WHERE wl.employee_name = 'AUTO-BYPASS'
        ORDER BY wl.id DESC
    """).fetchall()
    employees = conn.execute(
        "SELECT id, name, COALESCE(hindi_name,'') as hindi_name FROM employees WHERE active=1 ORDER BY name"
    ).fetchall()
    conn.close()

    def fmtd(d):
        if not d: return "—"
        p = str(d).split("-")
        return f"{p[2]}-{p[1]}-{p[0]}" if len(p)==3 else d

    entries = [{
        "id": r["id"], "order_code": r["order_code"], "cname": r["cname"] or "—",
        "garment_type": r["garment_type"], "qty": r["qty_done"],
        "notes": r["notes"] or "", "date": fmtd(r["log_date"]),
    } for r in rows]

    return render_template("owner/auto_bypass_fix.html",
        active_page="auto_bypass_fix", entries=entries,
        employees=[dict(e) for e in employees])


@bp.route("/api/reassign-worklog", methods=["POST"])
@owner_required
def api_reassign_worklog():
    """Change a work_log entry's employee_name from AUTO-BYPASS to a real employee."""
    data = request.get_json(silent=True) or {}
    wl_id = data.get("id")
    emp_name = (data.get("employee_name") or "").strip()
    if not wl_id or not emp_name:
        return jsonify({"ok": False, "error": "Missing id or employee_name"})
    conn = get_db()
    row = conn.execute("SELECT id FROM work_logs WHERE id=? AND employee_name='AUTO-BYPASS'", (wl_id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({"ok": False, "error": "Entry not found or already reassigned"})
    conn.execute("UPDATE work_logs SET employee_name=? WHERE id=?", (emp_name, wl_id))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@bp.route("/full-search")
@owner_required
def full_search_page():
    """One search box → everything about an order/customer: garments &
    measurements, who logged नाप/कटाई/सिलाई and when, every payment/advance
    with mode and date, and every other order (old + new) for that customer."""
    q         = request.args.get("q", "").strip()
    from_date = request.args.get("from", "")
    to_date   = request.args.get("to", "")

    def fmtd(d):
        if not d: return "—"
        p = str(d).split("-")
        return f"{p[2]}-{p[1]}-{p[0]}" if len(p)==3 else d

    def fmt_dt(ts):
        if not ts or len(ts) < 16: return ts or "—"
        try:
            d, t = ts[:10], ts[11:16]
            h, m = int(t[:2]), t[3:5]
            ampm = "AM" if h < 12 else "PM"
            h12 = h % 12 or 12
            return f"{fmtd(d)} · {h12}:{m} {ampm}"
        except Exception:
            return ts

    orders_out = []
    if q:
        conn = get_db()
        qnum = q.lstrip("#").strip()
        base_orders = []
        if qnum.isdigit():
            exact = conn.execute(
                "SELECT customer_id FROM orders WHERE order_code=?", (qnum,)
            ).fetchone()
            if exact:
                base_orders = conn.execute("""
                    SELECT o.id, o.order_code, o.repeat_of, o.order_date, o.delivery_date,
                           o.delivered_at, o.status, o.payable_amount, o.advance_paid,
                           o.remaining, o.payment_mode, o.is_urgent, o.note, o.customer_id,
                           c.name as cname, c.mobile, c.address
                    FROM orders o JOIN customers c ON c.id=o.customer_id
                    WHERE o.customer_id=?
                    ORDER BY o.id DESC
                """, (exact["customer_id"],)).fetchall()
            else:
                # Fall back to mobile match
                base_orders = conn.execute("""
                    SELECT o.id, o.order_code, o.repeat_of, o.order_date, o.delivery_date,
                           o.delivered_at, o.status, o.payable_amount, o.advance_paid,
                           o.remaining, o.payment_mode, o.is_urgent, o.note, o.customer_id,
                           c.name as cname, c.mobile, c.address
                    FROM orders o JOIN customers c ON c.id=o.customer_id
                    WHERE c.mobile LIKE ?
                    ORDER BY o.id DESC LIMIT 100
                """, (f"%{qnum}%",)).fetchall()
        else:
            base_orders = conn.execute("""
                SELECT o.id, o.order_code, o.repeat_of, o.order_date, o.delivery_date,
                       o.delivered_at, o.status, o.payable_amount, o.advance_paid,
                       o.remaining, o.payment_mode, o.is_urgent, o.note, o.customer_id,
                       c.name as cname, c.mobile, c.address
                FROM orders o JOIN customers c ON c.id=o.customer_id
                WHERE LOWER(c.name) LIKE LOWER(?)
                ORDER BY o.id DESC LIMIT 100
            """, (f"%{q}%",)).fetchall()

        if from_date:
            base_orders = [o for o in base_orders if o["order_date"] and str(o["order_date"]) >= from_date]
        if to_date:
            base_orders = [o for o in base_orders if o["order_date"] and str(o["order_date"]) <= to_date]

        latest_id_per_cust = {}
        for o in base_orders:
            cid = o["customer_id"]
            if cid not in latest_id_per_cust or o["id"] > latest_id_per_cust[cid]:
                latest_id_per_cust[cid] = o["id"]

        order_ids = [o["id"] for o in base_orders]
        garments_by_order, wl_by_order, fin_by_order = {}, {}, {}
        if order_ids:
            ph = ",".join("?" * len(order_ids))
            for g in conn.execute(
                f"SELECT order_id, garment_type, quantity, rate, amount, notes, measurements FROM order_items WHERE order_id IN ({ph}) ORDER BY order_id, id",
                order_ids
            ).fetchall():
                try: meas = json.loads(g["measurements"] or "{}")
                except Exception: meas = {}
                garments_by_order.setdefault(g["order_id"], []).append({
                    "garment_type": g["garment_type"], "quantity": g["quantity"],
                    "rate": g["rate"], "amount": g["amount"],
                    "notes": (g["notes"] or "").split("[")[0].strip(),
                    "measurements": meas,
                })
            for w in conn.execute(
                f"SELECT order_id, employee_name, garment_type, qty_done, notes, log_date, created_at FROM work_logs WHERE order_id IN ({ph}) ORDER BY created_at DESC",
                order_ids
            ).fetchall():
                n = (w["notes"] or "")
                stage = "नाप" if any(x in n for x in ["Measurement","Naap","नाप"]) else \
                        "कटाई" if any(x in n for x in ["Kataai","Cutting","कटाई"]) else "सिलाई"
                wl_by_order.setdefault(w["order_id"], []).append({
                    "employee": w["employee_name"] or "—", "garment_type": w["garment_type"],
                    "stage": stage, "qty": w["qty_done"], "when": fmt_dt(w["created_at"] or w["log_date"]),
                })
            for f in conn.execute(
                f"SELECT order_id, tx_date, tx_type, category, amount, mode, note, created_by, created_at FROM finance WHERE order_id IN ({ph}) ORDER BY created_at DESC",
                order_ids
            ).fetchall():
                fin_by_order.setdefault(f["order_id"], []).append({
                    "type": f["tx_type"], "category": f["category"], "amount": f["amount"],
                    "mode": f["mode"], "note": f["note"] or "", "by": f["created_by"] or "—",
                    "when": fmt_dt(f["created_at"]),
                })

        conn.close()

        for o in base_orders:
            orders_out.append({
                "id": o["id"],
                "code": o["order_code"],
                "display_code": o["repeat_of"] if o["repeat_of"] else o["order_code"],
                "entry_code": o["order_code"] if o["repeat_of"] else "",
                "is_latest": latest_id_per_cust.get(o["customer_id"]) == o["id"],
                "status": o["status"],
                "urgent": bool(o["is_urgent"]),
                "order_date": fmtd(o["order_date"]),
                "delivery_date": fmtd(o["delivery_date"]),
                "delivered_at": fmt_dt(o["delivered_at"]) if o["delivered_at"] else "—",
                "payable": int(o["payable_amount"] or 0),
                "advance": int(o["advance_paid"] or 0),
                "due": int(o["remaining"] or 0),
                "mode": o["payment_mode"] or "—",
                "note": o["note"] or "",
                "cname": o["cname"] or "—",
                "mobile": o["mobile"] or "—",
                "address": o["address"] or "—",
                "garments": garments_by_order.get(o["id"], []),
                "work_logs": wl_by_order.get(o["id"], []),
                "finance": fin_by_order.get(o["id"], []),
            })
        # Newest first
        orders_out.sort(key=lambda x: -x["id"])

    # ── Monthly Business Overview — shown when a date range is set,
    # regardless of whether a specific order/customer was also searched.
    overview = None
    if from_date and to_date:
        conn2 = get_db()

        # Orders created in this period, by status
        created_rows = conn2.execute("""
            SELECT status, COUNT(*) as cnt, COALESCE(SUM(payable_amount),0) as total_value,
                   COALESCE(SUM(remaining),0) as total_due
            FROM orders
            WHERE order_date >= ? AND order_date <= ?
            GROUP BY status
        """, (from_date, to_date)).fetchall()

        status_map = {r["status"]: r for r in created_rows}
        created_total = sum(r["cnt"] for r in created_rows)

        # Orders actually delivered in this period (regardless of when created)
        delivered_row = conn2.execute("""
            SELECT COUNT(*) as cnt, COALESCE(SUM(payable_amount),0) as total_value
            FROM orders
            WHERE status='delivered' AND delivered_at >= ? AND delivered_at <= ?
        """, (from_date, to_date + " 23:59:59")).fetchone()

        # Finance: income & expense totals in this period
        fin_rows = conn2.execute("""
            SELECT tx_type, COALESCE(SUM(amount),0) as total
            FROM finance
            WHERE tx_date >= ? AND tx_date <= ?
            GROUP BY tx_type
        """, (from_date, to_date)).fetchall()
        income  = next((r["total"] for r in fin_rows if r["tx_type"]=="income"), 0)
        expense = next((r["total"] for r in fin_rows if r["tx_type"]=="expense"), 0)

        # Salary paid per employee in this period
        salary_rows = conn2.execute("""
            SELECT employee_name, COALESCE(SUM(amount),0) as total
            FROM salary_advances
            WHERE advance_date >= ? AND advance_date <= ?
            GROUP BY employee_name
            ORDER BY total DESC
        """, (from_date, to_date)).fetchall()
        total_salary = sum(r["total"] for r in salary_rows)

        conn2.close()

        def st(name):
            r = status_map.get(name)
            return {"count": r["cnt"] if r else 0, "value": int(r["total_value"] or 0) if r else 0,
                    "due": int(r["total_due"] or 0) if r else 0}

        overview = {
            "created_total": created_total,
            "pending": st("pending"),
            "ready": st("ready"),
            "delivered_created": st("delivered"),
            "cancelled": st("cancelled"),
            "delivered_period": {
                "count": delivered_row["cnt"] or 0,
                "value": int(delivered_row["total_value"] or 0),
            },
            "income": int(income),
            "expense": int(expense),
            "net": int(income) - int(expense),
            "total_salary": int(total_salary),
            "by_employee": [{"name": r["employee_name"], "amount": int(r["total"])} for r in salary_rows],
        }

    return render_template("owner/full_search.html",
        active_page="full_search", q=q, from_date=from_date, to_date=to_date,
        orders=orders_out, overview=overview)


@bp.route("/mixed-customers-3701")
@owner_required
def mixed_customers_3701():
    """Diagnostic (read-only): customers linked to MULTIPLE orders within the
    order_code >= 3701 range. A customer legitimately having several past
    orders is normal — but if two DIFFERENT people were entered with the same
    mobile number (e.g. typo, or family number reused), Past Orders' "same
    mobile = same person" matching silently overwrote the earlier person's
    name/address, and every order sharing that customer_id now shows the
    latest name. This page surfaces those groups so they can be reviewed and
    split via /owner/api/fix-order-customer where needed. Nothing is changed here."""
    conn = get_db()
    rows = conn.execute("""
        SELECT c.id as customer_id, c.name, c.mobile, c.address,
               COUNT(*) as order_count,
               STRING_AGG(
                   o.order_code || '~~' || COALESCE(o.delivery_date::text,'') || '~~' || COALESCE(o.payable_amount::text,'0') || '~~' || o.status,
                   '||' ORDER BY o.order_code::integer
               ) as details
        FROM orders o
        JOIN customers c ON c.id = o.customer_id
        WHERE o.order_code ~ '^[0-9]+$' AND o.order_code::integer >= 3701
        GROUP BY c.id, c.name, c.mobile, c.address
        HAVING COUNT(*) > 1
        ORDER BY COUNT(*) DESC
        LIMIT 200
    """).fetchall()
    conn.close()

    groups = []
    for r in rows:
        orders = []
        for part in (r["details"] or "").split("||"):
            bits = part.split("~~")
            if len(bits) == 4:
                orders.append({"code": bits[0], "delivery": bits[1] or "—", "amount": bits[2], "status": bits[3]})
        groups.append({
            "customer_id": r["customer_id"], "name": r["name"] or "—",
            "mobile": r["mobile"] or "—", "address": r["address"] or "—",
            "count": r["order_count"], "orders": orders
        })

    return render_template("owner/mixed_customers_3701.html",
        active_page="mixed_customers_3701", groups=groups)

@bp.route("/same-mobile")
@owner_required
def same_mobile_page():
    """Find mobile numbers shared by different-named customers."""
    conn = get_db()
    rows = conn.execute("""
        SELECT c.mobile,
               COUNT(o.id) as total_orders,
               COUNT(DISTINCT LOWER(TRIM(c.name))) as unique_names,
               STRING_AGG(DISTINCT TRIM(c.name), ' | ' ORDER BY TRIM(c.name)) as all_names,
               STRING_AGG(o.order_code || '~~' || TRIM(COALESCE(c.name,'')) || '~~' || CAST(c.id AS TEXT),
                          '||' ORDER BY o.id DESC) as details
        FROM orders o
        JOIN customers c ON c.id = o.customer_id
        WHERE c.mobile IS NOT NULL
          AND LENGTH(c.mobile) >= 8
          AND c.mobile NOT IN ('', '-', 'None', '0')
        GROUP BY c.mobile
        HAVING COUNT(DISTINCT LOWER(TRIM(c.name))) > 1
        ORDER BY COUNT(DISTINCT LOWER(TRIM(c.name))) DESC, COUNT(o.id) DESC
    """).fetchall()
    problems = []
    for row in rows:
        orders = []
        for part in (row["details"] or "").split("||"):
            bits = part.split("~~")
            if len(bits) == 3:
                orders.append({"code": bits[0].strip(), "name": bits[1].strip()})
        problems.append({
            "mobile": row["mobile"],
            "total": int(row["total_orders"] or 0),
            "names": row["all_names"] or "",
            "orders": orders[:15]
        })
    conn.close()
    return render_template("owner/same_mobile_orders.html",
        active_page="same_mobile",
        problems=problems,
        problem_count=len(problems))



@bp.route("/api/bulk-split-customers")
@owner_required
def api_bulk_split_customers():
    """BULK FIX: For every mobile that has 2+ different-named customers,
    ensure each (name, mobile) pair has its OWN customer record.
    Orders stay linked to correct customers. Fixes all merged records at once."""
    from datetime import datetime as _dt
    conn = get_db()
    now_str = _dt.now().strftime("%Y-%m-%d %H:%M:%S")

    # Find all orders where customer mobile is shared by different names
    # We get each order's code + customer name + mobile
    rows = conn.execute("""
        SELECT o.id as order_id, o.order_code,
               c.id as cust_id, TRIM(c.name) as cname,
               c.mobile, c.address
        FROM orders o
        JOIN customers c ON c.id = o.customer_id
        WHERE c.mobile IS NOT NULL
          AND LENGTH(c.mobile) >= 8
          AND c.mobile NOT IN ('', '-', 'None', '0')
        ORDER BY c.mobile, TRIM(c.name), o.id
    """).fetchall()

    # Group by mobile
    from collections import defaultdict
    mobile_map = defaultdict(lambda: defaultdict(list))
    for r in rows:
        mobile_map[r["mobile"]][r["cname"].lower()].append({
            "order_id":  r["order_id"],
            "order_code": r["order_code"],
            "cust_id":   r["cust_id"],
            "name":      r["cname"],
            "address":   r["address"] or ""
        })

    fixed_orders = 0
    fixed_customers = 0

    for mobile, name_groups in mobile_map.items():
        if len(name_groups) <= 1:
            continue  # Only one name for this mobile — no problem

        # Multiple names share same mobile
        # Sort names — first one keeps the original customer record
        sorted_names = sorted(name_groups.keys())
        first_name = sorted_names[0]

        # For each name group after the first → ensure they have their own customer
        for name_key in sorted_names[1:]:
            orders_in_group = name_groups[name_key]
            actual_name = orders_in_group[0]["name"]

            # Check if a separate customer exists for this name+mobile
            existing = conn.execute(
                "SELECT id FROM customers WHERE LOWER(TRIM(name))=LOWER(?) AND mobile=?",
                (actual_name, mobile)
            ).fetchone()

            if existing:
                correct_cust_id = existing["id"]
            else:
                # Create new customer for this name
                row = conn.execute(
                    "INSERT INTO customers(name, mobile, address, created_at) VALUES(?,?,?,?) RETURNING id",
                    (actual_name, mobile, orders_in_group[0]["address"], now_str)
                ).fetchone()
                correct_cust_id = row["id"]
                fixed_customers += 1

            # Update all orders in this name group to point to correct customer
            for o in orders_in_group:
                if o["cust_id"] != correct_cust_id:
                    conn.execute(
                        "UPDATE orders SET customer_id=? WHERE id=?",
                        (correct_cust_id, o["order_id"])
                    )
                    fixed_orders += 1

    conn.commit()
    conn.close()

    return jsonify({
        "ok": True,
        "fixed_orders": fixed_orders,
        "new_customers_created": fixed_customers,
        "message": f"✅ {fixed_orders} orders alag ho gaye! {fixed_customers} naye customer records bane."
    })


@bp.route("/api/fix-order-mobile", methods=["POST"])
@owner_required
def api_fix_order_mobile():
    """Fix a single order's customer mobile number. POST JSON: {order_code, name, mobile}"""
    from datetime import datetime as _dt
    d = request.get_json(silent=True) or {}
    code   = d.get("order_code","").strip()
    name   = d.get("name","").strip()
    mobile = d.get("mobile","").strip()
    if not code or not mobile or not name:
        return jsonify({"ok": False, "error": "order_code, name, mobile required"})
    conn = get_db()
    order = conn.execute("SELECT id, customer_id FROM orders WHERE order_code=?", (code,)).fetchone()
    if not order:
        conn.close()
        return jsonify({"ok": False, "error": f"Order #{code} not found"})
    now_str = _dt.now().strftime("%Y-%m-%d %H:%M:%S")
    existing = conn.execute(
        "SELECT id FROM customers WHERE mobile=? AND LOWER(TRIM(name))=LOWER(TRIM(?))",
        (mobile, name)
    ).fetchone()
    if existing:
        cust_id = existing["id"]
    else:
        old = conn.execute("SELECT address FROM customers WHERE id=?", (order["customer_id"],)).fetchone()
        row = conn.execute(
            "INSERT INTO customers(name, mobile, address, created_at) VALUES(?,?,?,?) RETURNING id",
            (name, mobile, (old["address"] if old else "") or "", now_str)
        ).fetchone()
        cust_id = row["id"]
    conn.execute("UPDATE orders SET customer_id=? WHERE order_code=?", (cust_id, code))
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "message": f"✅ Order #{code} → {name} ({mobile})"})


@bp.route("/api/diagnose-orders")
@owner_required
def api_diagnose_orders():
    """Check if two orders share same customer_id."""
    c1 = request.args.get("c1","").strip()
    c2 = request.args.get("c2","").strip()
    conn = get_db()

    def get_order_info(code):
        r = conn.execute("""
            SELECT o.id, o.order_code, o.customer_id,
                   c.id as cid, c.name as cname, c.mobile, c.address
            FROM orders o
            LEFT JOIN customers c ON c.id=o.customer_id
            WHERE o.order_code=?
        """, (code,)).fetchone()
        if not r: return None
        # Also get all orders linked to this customer
        others = conn.execute(
            "SELECT order_code FROM orders WHERE customer_id=? ORDER BY id",
            (r["customer_id"],)).fetchall()
        return {
            "order_code": r["order_code"],
            "customer_id": r["customer_id"],
            "name": r["cname"],
            "mobile": r["mobile"],
            "all_orders_of_this_customer": [x["order_code"] for x in others]
        }

    o1 = get_order_info(c1)
    o2 = get_order_info(c2)
    conn.close()

    same_cust = o1 and o2 and o1["customer_id"] == o2["customer_id"]
    return jsonify({
        "order_1": o1,
        "order_2": o2,
        "PROBLEM": same_cust,
        "reason": "Dono same customer_id share kar rahe hain!" if same_cust else "Alag customer_id hain"
    })


@bp.route("/api/mobile-audit-count")
@owner_required
def api_mobile_audit_count():
    """Count exactly how many orders are actually affected."""
    conn = get_db()

    # Total orders
    total = conn.execute("SELECT COUNT(*) as c FROM orders").fetchone()["c"]

    # Orders with unique mobile (safe - no conflict possible)
    unique_mobile = conn.execute("""
        SELECT COUNT(o.id) as c FROM orders o
        JOIN customers c ON c.id=o.customer_id
        WHERE c.mobile IS NOT NULL
        GROUP BY c.mobile
        HAVING COUNT(DISTINCT LOWER(TRIM(c.name))) = 1
    """).fetchall()

    # Mobile conflicts (same mobile, different names)
    conflicts = conn.execute("""
        SELECT c.mobile,
               COUNT(DISTINCT LOWER(TRIM(c.name))) as names,
               COUNT(o.id) as order_count,
               STRING_AGG(DISTINCT TRIM(c.name), ' | ' ORDER BY TRIM(c.name)) as all_names
        FROM orders o
        JOIN customers c ON c.id=o.customer_id
        WHERE c.mobile IS NOT NULL AND LENGTH(c.mobile) >= 8
        GROUP BY c.mobile
        HAVING COUNT(DISTINCT LOWER(TRIM(c.name))) > 1
        ORDER BY COUNT(o.id) DESC
    """).fetchall()

    conflict_orders = sum(int(r["order_count"]) for r in conflicts)

    # Orders with no mobile
    no_mobile = conn.execute("""
        SELECT COUNT(o.id) as c FROM orders o
        JOIN customers c ON c.id=o.customer_id
        WHERE c.mobile IS NULL OR c.mobile=''
    """).fetchone()["c"]

    conn.close()
    return jsonify({
        "total_orders": total,
        "no_mobile_orders": no_mobile,
        "conflict_mobiles": len(conflicts),
        "orders_in_conflict": conflict_orders,
        "safe_orders": total - conflict_orders - no_mobile,
        "top_conflicts": [{"mobile": r["mobile"], "names": r["all_names"],
                           "orders": r["order_count"]} for r in conflicts[:10]]
    })

@bp.route("/api/smart-fix-mobile-conflicts")
@owner_required
def api_smart_fix_mobile_conflicts():
    """Smart fix: auto-merge typo names (same person), flag truly different people."""
    from datetime import datetime as _dt
    conn = get_db()

    def name_similarity(a, b):
        """Check if two names are likely the same person (typo/abbreviation)."""
        a = a.lower().strip()
        b = b.lower().strip()
        if a == b: return True
        # Remove common suffixes
        for suf in [' ji', ' kumar', ' lal', ' ram', ' singh']:
            a = a.replace(suf, '')
            b = b.replace(suf, '')
        a = a.strip(); b = b.strip()
        if a == b: return True
        # One contains the other
        if a in b or b in a: return True
        # First word matches
        if a.split()[0] == b.split()[0] and len(a.split()[0]) > 3: return True
        # Very similar (edit distance <= 2)
        if abs(len(a)-len(b)) <= 2:
            diff = sum(1 for x,y in zip(a,b) if x!=y) + abs(len(a)-len(b))
            if diff <= 2: return True
        return False

    # Get all conflicts
    rows = conn.execute("""
        SELECT c.mobile, c.id as cust_id, TRIM(c.name) as cname,
               COUNT(o.id) as ord_count
        FROM customers c
        JOIN orders o ON o.customer_id=c.id
        WHERE c.mobile IS NOT NULL AND LENGTH(c.mobile) >= 8
        GROUP BY c.mobile, c.id, c.name
        HAVING COUNT(o.id) >= 1
        ORDER BY c.mobile, c.id
    """).fetchall()

    # Group by mobile
    from collections import defaultdict
    mobile_groups = defaultdict(list)
    for r in rows:
        mobile_groups[r["mobile"]].append({
            "cust_id": r["cust_id"],
            "name": r["cname"],
            "orders": r["ord_count"]
        })

    merged = 0
    needs_review = []
    now_str = _dt.now().strftime("%Y-%m-%d %H:%M:%S")

    for mobile, customers in mobile_groups.items():
        if len(customers) <= 1:
            continue

        # Group customers by similarity
        groups = []
        used = set()
        for i, c in enumerate(customers):
            if i in used: continue
            group = [c]
            used.add(i)
            for j, c2 in enumerate(customers):
                if j in used: continue
                if name_similarity(c["name"], c2["name"]):
                    group.append(c2)
                    used.add(j)
            groups.append(group)

        for group in groups:
            if len(group) <= 1:
                continue
            # All in this group are likely same person — merge into oldest record
            group.sort(key=lambda x: x["cust_id"])
            keep_id = group[0]["cust_id"]
            # Use longest name as canonical
            best_name = max([g["name"] for g in group], key=len)
            conn.execute("UPDATE customers SET name=? WHERE id=?", (best_name, keep_id))
            for dup in group[1:]:
                # Move all orders to canonical customer
                conn.execute("UPDATE orders SET customer_id=? WHERE customer_id=?",
                             (keep_id, dup["cust_id"]))
                merged += 1

        # Check for truly different people (groups that didn't merge)
        if len(groups) > 1:
            group_names = [max([g["name"] for g in grp], key=len) for grp in groups]
            total_orders = sum(c["orders"] for c in customers)
            needs_review.append({
                "mobile": mobile,
                "different_people": group_names,
                "orders": total_orders
            })

    conn.commit()
    conn.close()

    return jsonify({
        "ok": True,
        "typo_merges_done": merged,
        "truly_different_people": len(needs_review),
        "review_list": needs_review[:50],
        "message": f"✅ {merged} typo duplicates merged. {len(needs_review)} mobiles need manual review (genuinely different people)."
    })


@bp.route("/api/server-stats")
@owner_required
def api_server_stats():
    """Return server disk, RAM, CPU usage."""
    import shutil, subprocess, os
    result = {}

    # Disk usage
    try:
        total, used, free = shutil.disk_usage("/")
        result["disk"] = {
            "total_gb": round(total / 1e9, 1),
            "used_gb":  round(used  / 1e9, 1),
            "free_gb":  round(free  / 1e9, 1),
            "used_pct": round(used / total * 100, 1)
        }
    except Exception as e:
        result["disk"] = {"error": str(e)}

    # RAM usage
    try:
        with open("/proc/meminfo") as f:
            mem = {}
            for line in f:
                k, v = line.split(":")
                mem[k.strip()] = int(v.strip().split()[0])
        total_mb = round(mem["MemTotal"] / 1024, 0)
        free_mb  = round((mem.get("MemAvailable", mem.get("MemFree", 0))) / 1024, 0)
        used_mb  = total_mb - free_mb
        result["ram"] = {
            "total_mb": total_mb,
            "used_mb":  used_mb,
            "free_mb":  free_mb,
            "used_pct": round(used_mb / total_mb * 100, 1)
        }
    except Exception as e:
        result["ram"] = {"error": str(e)}

    # CPU usage (via /proc/loadavg — no external binary needed, works even without procps installed)
    try:
        with open("/proc/loadavg") as f:
            load1, load5, load15 = f.read().split()[:3]
        cores = os.cpu_count() or 1
        result["cpu_line"] = f"load avg: {load1}, {load5}, {load15} (over {cores} core{'s' if cores != 1 else ''})"
        result["cpu_load_pct_1min"] = round(float(load1) / cores * 100, 1)
    except Exception as e:
        result["cpu_line"] = str(e)

    # UTMS app process memory (scan /proc/[pid]/ directly — no ps needed)
    try:
        total_kb = 0
        for pid in os.listdir("/proc"):
            if not pid.isdigit():
                continue
            try:
                with open(f"/proc/{pid}/cmdline", "rb") as f:
                    cmdline = f.read().replace(b"\x00", b" ").decode(errors="ignore")
                if "gunicorn" in cmdline:
                    with open(f"/proc/{pid}/status") as f:
                        for line in f:
                            if line.startswith("VmRSS:"):
                                total_kb += int(line.split()[1])
                                break
            except Exception:
                continue
        result["gunicorn_ram"] = f"{round(total_kb / 1024, 1)} MB"
    except Exception as e:
        result["gunicorn_ram"] = str(e)

    # Uptime (via /proc/uptime — no external binary needed)
    try:
        with open("/proc/uptime") as f:
            secs = float(f.read().split()[0])
        d, h, m = int(secs // 86400), int((secs % 86400) // 3600), int((secs % 3600) // 60)
        parts = ([f"{d}d"] if d else []) + ([f"{h}h"] if h else []) + [f"{m}m"]
        result["uptime"] = "up " + " ".join(parts)
    except Exception:
        result["uptime"] = "unknown"

    # Image folder size
    try:
        out = subprocess.check_output(
            "du -sh /home/ubuntu/UTMS/order_images/ 2>/dev/null || echo '0'",
            shell=True, text=True).strip()
        result["images_size"] = out.split()[0] if out else "0"
    except:
        result["images_size"] = "unknown"

    return jsonify({"ok": True, "stats": result})


@bp.route("/api/upload-order-image", methods=["POST"])
@owner_required
def api_upload_order_image():
    """Replace/add image for an order from diary page."""
    import os, uuid
    from datetime import datetime as _dt
    code  = (request.form.get("order_code") or "").strip().lstrip("#")
    file  = request.files.get("image")
    if not code or not file:
        return jsonify({"ok": False, "error": "code and image required"})
    conn = get_db()
    order = conn.execute("SELECT id FROM orders WHERE order_code=?", (code,)).fetchone()
    if not order:
        conn.close()
        return jsonify({"ok": False, "error": f"Order #{code} not found"})
    # Save image to static/order_images directory (Flask static folder)
    img_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
                           "static", "order_images", code)
    os.makedirs(img_dir, exist_ok=True)
    ext     = os.path.splitext(file.filename or "img.jpg")[1] or ".jpg"
    fname   = f"{uuid.uuid4().hex[:8]}{ext}"
    fpath   = os.path.join(img_dir, fname)
    file.save(fpath)
    # Insert into order_images table
    now_str = _dt.now().strftime("%Y-%m-%d %H:%M:%S")
    file_path = f"/static/order_images/{code}/{fname}"
    conn.execute(
        "INSERT INTO order_images(order_id, file_path) VALUES(?,?)",
        (order["id"], file_path)
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "filename": fname, "message": "✅ Image uploaded!"})

@bp.route("/finance")
@owner_required
def owner_finance():
    today = date.today().isoformat()
    # Support ?filter=today or ?filter=month from dashboard cards
    filt = request.args.get("filter", "")
    if filt == "today":
        from_date = request.args.get("date", today)
        to_date   = from_date
    elif filt == "month":
        from_date = today[:7] + "-01"
        to_date   = today
    else:
        from_date = request.args.get("from", today[:7]+"-01")
        to_date   = request.args.get("to",   today)

    conn = get_db()
    rows = conn.execute("""
        SELECT f.*, o.order_code, o.repeat_of
        FROM finance f LEFT JOIN orders o ON o.id=f.order_id
        WHERE f.tx_date >= ? AND f.tx_date <= ?
        ORDER BY f.tx_date DESC, f.id DESC
    """, (from_date, to_date)).fetchall()

    stats_r = conn.execute("""
        SELECT
            COALESCE(SUM(CASE WHEN tx_type='income' THEN amount ELSE 0 END),0) as income,
            COALESCE(SUM(CASE WHEN tx_type='income' AND mode='cash' THEN amount ELSE 0 END),0) as cash_i,
            COALESCE(SUM(CASE WHEN tx_type='income' AND mode='upi' THEN amount ELSE 0 END),0) as upi_i,
            COALESCE(SUM(CASE WHEN tx_type='expense' THEN amount ELSE 0 END),0) as expense,
            COUNT(CASE WHEN tx_type='expense' THEN 1 END) as exp_count
        FROM finance WHERE tx_date >= ? AND tx_date <= ?
    """, (from_date, to_date)).fetchone()

    pending_due = conn.execute(
        "SELECT COALESCE(SUM(remaining),0) as d FROM orders WHERE status!='delivered'"
    ).fetchone()["d"]
    urgent_count = conn.execute(
        "SELECT COUNT(*) as c FROM orders WHERE is_urgent=1 AND status!='delivered'"
    ).fetchone()["c"]
    pending_orders_rows = conn.execute("""
        SELECT o.order_code, o.repeat_of, c.name as cname, o.remaining, o.payable_amount,
               o.advance_paid, o.delivery_date, o.status
        FROM orders o JOIN customers c ON c.id=o.customer_id
        WHERE o.status!='delivered' AND o.remaining>0
        ORDER BY o.remaining DESC
    """).fetchall()
    conn.close()

    def fmtd(d):
        if not d: return "—"
        p = str(d).split("-")
        return f"{p[2]}-{p[1]}-{p[0]}" if len(p)==3 else d

    def fmt_time12(ts):
        """Convert 'YYYY-MM-DD HH:MM:SS' (24h) to 12-hour 'h:MM AM/PM'."""
        if not ts or len(ts) < 16:
            return ""
        try:
            hh = int(ts[11:13])
            mm = ts[14:16]
            suffix = "AM" if hh < 12 else "PM"
            hh12 = hh % 12
            if hh12 == 0:
                hh12 = 12
            return f"{hh12}:{mm} {suffix}"
        except (ValueError, IndexError):
            return ts[11:16]

    pending_orders = [{
        "order_code":   (r["repeat_of"] if r["repeat_of"] else r["order_code"]),
        "entry_code":   r["order_code"] if r["repeat_of"] else "",
        "cname":        r["cname"],
        "remaining":    int(r["remaining"] or 0),
        "payable":      int(r["payable_amount"] or 0),
        "advance_paid": int(r["advance_paid"] or 0),
        "delivery_date":fmtd(r["delivery_date"]),
        "status":       r["status"]
    } for r in pending_orders_rows]

    transactions = [{
        "id":          r["id"],
        "tx_date":     r["tx_date"],
        "tx_date_fmt": fmtd(r["tx_date"]),
        "tx_time":     fmt_time12(r["created_at"]),
        "tx_type":     r["tx_type"],
        "category":    r["category"] or "",
        "note":        r["note"] or "",
        "mode":        r["mode"] or "",
        "amount":      r["amount"] or 0,
        "order_code":  (r["repeat_of"] if r["repeat_of"] else r["order_code"]) or "",
        "entry_code":  r["order_code"] if r["repeat_of"] else "",
        "created_by":  r["created_by"] or ""
    } for r in rows]

    net = int((stats_r["income"] or 0) - (stats_r["expense"] or 0))
    return render_template("owner/finance.html",
        active_page="finance", show_voice=False,
        urgent_count=urgent_count,
        from_date=from_date, to_date=to_date,
        pending_orders=pending_orders,
        stats={
            "total_income":  int(stats_r["income"] or 0),
            "cash_income":   int(stats_r["cash_i"] or 0),
            "upi_income":    int(stats_r["upi_i"] or 0),
            "total_expense": int(stats_r["expense"] or 0),
            "expense_count": int(stats_r["exp_count"] or 0),
            "net":           net,
            "pending_due":   int(pending_due or 0)
        },
        transactions=transactions
    )


# ══════════════════════════════════════════════
#  FINANCE ENTRY DELETE
# ══════════════════════════════════════════════

@bp.route("/api/finance/update-mode/<int:tx_id>", methods=["POST"])
@owner_required
def api_finance_update_mode(tx_id):
    """Fix a transaction's payment mode (Cash/UPI) — for when the wrong one
    was accidentally selected while adding the entry. Also updates the
    linked order's payment_mode if this entry is tied to one, so both stay
    consistent."""
    data = request.get_json(silent=True) or {}
    new_mode = (data.get("mode") or "").strip().lower()
    if new_mode not in ("cash", "upi"):
        return jsonify({"ok": False, "error": "Mode must be cash or upi"})

    conn = get_db()
    entry = conn.execute("SELECT * FROM finance WHERE id=?", (tx_id,)).fetchone()
    if not entry:
        conn.close()
        return jsonify({"ok": False, "error": "Entry not found"})

    conn.execute("UPDATE finance SET mode=? WHERE id=?", (new_mode, tx_id))

    if entry["order_id"]:
        conn.execute("UPDATE orders SET payment_mode=? WHERE id=?", (new_mode, entry["order_id"]))

    conn.commit()
    conn.close()
    return jsonify({"ok": True, "mode": new_mode})


@bp.route("/api/finance/delete/<int:tx_id>", methods=["POST"])
@owner_required
def delete_finance_entry(tx_id):
    """Delete a finance entry. Reverses advance/payment impact on order if applicable."""
    conn = get_db()
    entry = conn.execute("SELECT * FROM finance WHERE id=?", (tx_id,)).fetchone()
    if not entry:
        conn.close()
        return jsonify({"ok": False, "error": "Entry not found"})

    # If it was an income entry linked to an order (advance/payment/
    # remaining amount), reverse it
    if entry["tx_type"] == "income" and entry["order_id"] and \
       entry["category"] and entry["category"].lower() in ("advance", "payment", "remaining amount"):
        order = conn.execute(
            "SELECT advance_paid, remaining, payable_amount FROM orders WHERE id=?",
            (entry["order_id"],)
        ).fetchone()
        if order:
            amount       = float(entry["amount"] or 0)
            new_advance  = max(0.0, float(order["advance_paid"] or 0) - amount)
            payable      = float(order["payable_amount"] or 0)
            new_remaining = max(0.0, payable - new_advance)
            conn.execute(
                "UPDATE orders SET advance_paid=?, remaining=? WHERE id=?",
                (new_advance, new_remaining, entry["order_id"])
            )

    conn.execute("DELETE FROM finance WHERE id=?", (tx_id,))
    # If it was a salary expense, also remove from salary_advances
    if entry["tx_type"] == "expense" and entry["category"] and "salary" in entry["category"].lower():
        emp_name = entry["employee_name"] or entry["note"] or ""
        if emp_name:
            conn.execute("""DELETE FROM salary_advances WHERE employee_name=?
                AND amount=? AND advance_date=?""",
                (emp_name, entry["amount"], entry["tx_date"]))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


# ══════════════════════════════════════════════
#  FRESH START (DATE FILTER) TOGGLE
# ══════════════════════════════════════════════

@bp.route("/api/fresh-start/toggle", methods=["POST"])
@owner_required
def toggle_fresh_start():
    """Toggle the 'show only this month' fresh-start filter across all of UTMS."""
    from database import set_setting, invalidate_settings_cache
    current = get_setting("utms_fresh_start", "0")
    new_val = "0" if current == "1" else "1"
    set_setting("utms_fresh_start", new_val)
    # Set the cutoff date to the start of the current month if not already set
    if new_val == "1":
        existing_date = get_setting("utms_fresh_start_date", "")
        if not existing_date:
            from datetime import date as _date
            d = _date.today()
            set_setting("utms_fresh_start_date", f"{d.year}-{d.month:02d}-01")
    invalidate_settings_cache()
    enabled_date = get_setting("utms_fresh_start_date", "2026-06-01")
    return jsonify({"ok": True, "enabled": new_val == "1", "date": enabled_date})


@bp.route("/api/fresh-start/set-date", methods=["POST"])
@owner_required
def set_fresh_start_date():
    """Update the fresh-start cutoff date."""
    from database import set_setting, invalidate_settings_cache
    data = request.get_json(silent=True) or {}
    new_date = (data.get("date") or "").strip()
    if not new_date:
        return jsonify({"ok": False, "error": "Date required"})
    set_setting("utms_fresh_start_date", new_date)
    set_setting("utms_fresh_start", "1")  # Enable when date is explicitly set
    invalidate_settings_cache()
    return jsonify({"ok": True, "date": new_date})


# ══════════════════════════════════════════════
#  WHATSAPP BROADCAST
# ══════════════════════════════════════════════

@bp.route("/whatsapp")
@owner_required
def whatsapp():
    conn = get_db()
    # Load all customers with due amounts and ready order flags
    rows = conn.execute("""
        SELECT c.id, c.name, c.mobile,
               COALESCE(SUM(o.remaining),0) as due,
               MAX(CASE WHEN o.status='ready' THEN 1 ELSE 0 END) as has_ready
        FROM customers c
        LEFT JOIN orders o ON o.customer_id=c.id AND o.status!='delivered'
        WHERE c.mobile IS NOT NULL AND c.mobile != ''
        GROUP BY c.id, c.name, c.mobile, c.address ORDER BY c.name
    """).fetchall()

    customers = [{"id":r["id"],"name":r["name"],"mobile":r["mobile"] or "",
                  "due":r["due"] or 0,"has_ready":r["has_ready"] or 0}
                 for r in rows]

    # Recent broadcast log
    logs_raw = conn.execute("""
        SELECT message_type, COUNT(*) as count, MAX(sent_at) as sent_at
        FROM whatsapp_log GROUP BY message_type ORDER BY MAX(sent_at) DESC LIMIT 10
    """).fetchall()

    def fmtd(d):
        if not d: return "—"
        p = str(d).split("-")
        return f"{p[2]}-{p[1]}-{p[0]}" if len(p)==3 else d

    broadcast_log = [{"message_type":r["message_type"],"count":r["count"],
                      "sent_at_fmt":fmtd((r["sent_at"] or "")[:10])} for r in logs_raw]

    urgent_count = conn.execute(
        "SELECT COUNT(*) as c FROM orders WHERE is_urgent=1 AND status!='delivered'"
    ).fetchone()["c"]
    shop_name = get_setting("shop_name","Uttam Tailors")

    # Default templates (user's exact format)
    _D_CONFIRM_HI = "*उत्तम टेलर्स*\n\nनमस्ते *{name}*,\n\nआपका ऑर्डर सफलतापूर्वक दर्ज हो गया है।\n------------------------------------------\n\n- ऑर्डर नंबर: *#{code}*\n- ऑर्डर तारीख: {odate}\n- डिलीवरी तारीख: *{ddate}*\n\n- कपड़े: {items}\n\n- कुल राशि: *₹{total}*\n- अग्रिम भुगतान: ₹{advance}\n- बकाया राशि: *₹{due}*\n- भुगतान माध्यम: {mode}\n\n-----------------------------------------------\nआपके विश्वास के लिए धन्यवाद। 🙏🏻\nउत्तम टेलर्स"
    _D_CONFIRM_EN = "*Uttam Tailors*\n\nHello *{name}*,\nYour order has been confirmed successfully.\n------------------------------------------\n\n- Order No: *#{code}*\n- Order Date: {odate}\n- Delivery Date: *{ddate}*\n\n- Garments: {items}\n\n- Total Amount: *₹{total}*\n- Advance Paid: ₹{advance}\n- Balance Due: *₹{due}*\n- Payment Mode: {mode}\n\n-----------------------------------------------\nThank you for your trust. 🙏🏻\nUttam Tailors"
    _D_READY_HI   = "*उत्तम टेलर्स*\n\nनमस्ते *{name}*,\nआपका ऑर्डर तैयार है। कृपया अपनी सुविधानुसार आकर अपने कपड़े प्राप्त कर लें।\n------------------------------------------\n\n- ऑर्डर नंबर: *#{code}*\n- कपड़े: {items}\n- कुल राशि: *₹{total}*\n- बकाया राशि: *₹{due}*\n\n-----------------------------------------------\nआपके विश्वास के लिए धन्यवाद। 🙏🏻\nउत्तम टेलर्स"
    _D_READY_EN   = "*Uttam Tailors*\n\nHello *{name}*,\nYour order is ready. Please collect your clothes at your convenience. Thank you!\n------------------------------------------\n\n- Order No: *#{code}*\n- Garments: {items}\n- Total: *₹{total}*\n- Balance Due: *₹{due}*\n\n-----------------------------------------------\nThank you for your trust. 🙏🏻\nUttam Tailors"
    _D_DELIV_HI   = "*उत्तम टेलर्स*\n\nनमस्ते *{name}*,\nआपके कपड़े सफलतापूर्वक डिलीवर कर दिए गए हैं।\n------------------------------------------\n\n- ऑर्डर नंबर: *#{code}*\n- ऑर्डर तारीख: {odate}\n- डिलीवरी तारीख: {ddate}\n\n- कपड़े: {items}\n\n- कुल भुगतान: *₹{total}*\n- भुगतान माध्यम: {mode}\n\n-----------------------------------------------\nआपके विश्वास के लिए धन्यवाद। 🙏🏻\nआगे भी सेवा का अवसर दें, यही हमारी शुभकामना है।\n\nधन्यवाद!\nउत्तम टेलर्स"
    _D_DELIV_EN   = "*Uttam Tailors*\n\nHello *{name}*,\nYour garments have been delivered successfully.\n------------------------------------------\n\n- Order No: *#{code}*\n- Order Date: {odate}\n- Delivery Date: {ddate}\n\n- Garments: {items}\n\n- Total Paid: *₹{total}*\n- Payment Mode: {mode}\n\n-----------------------------------------------\nThank you for your trust. 🙏🏻\nWe look forward to serving you again.\n\nThank you!\nUttam Tailors"
    _D_REMIND_HI  = "*उत्तम टेलर्स*\n\nनमस्ते *{name}*,\nआपका ऑर्डर तैयार है और अभी तक pickup नहीं हुआ है। कृपया जल्द आकर अपने कपड़े ले जाएं।\n------------------------------------------\n\n- ऑर्डर नंबर: *#{code}*\n- कपड़े: {items}\n- बकाया राशि: *₹{due}*\n\n-----------------------------------------------\nधन्यवाद। 🙏🏻\nउत्तम टेलर्स"
    _D_REMIND_EN  = "*Uttam Tailors*\n\nHello *{name}*,\nYour order is ready and still awaiting pickup. Please collect your garments soon.\n------------------------------------------\n\n- Order No: *#{code}*\n- Garments: {items}\n- Balance Due: *₹{due}*\n\n-----------------------------------------------\nThank you. 🙏🏻\nUttam Tailors"

    wa_templates = {
        "confirm":   {"hi": get_setting("wa_order_confirm_hi",   _D_CONFIRM_HI),
                      "en": get_setting("wa_order_confirm_en",   _D_CONFIRM_EN)},
        "ready":     {"hi": get_setting("wa_order_ready_hi",     _D_READY_HI),
                      "en": get_setting("wa_order_ready_en",     _D_READY_EN)},
        "delivered": {"hi": get_setting("wa_order_delivered_hi", _D_DELIV_HI),
                      "en": get_setting("wa_order_delivered_en", _D_DELIV_EN)},
        "reminder":  {"hi": get_setting("wa_order_reminder_hi",  _D_REMIND_HI),
                      "en": get_setting("wa_order_reminder_en",  _D_REMIND_EN)},
    }

    conn.close()

    templates = [
        {"name":"Order Confirmed", "icon":"✅", "key":"confirm"},
        {"name":"Order Ready",     "icon":"🔔", "key":"ready"},
        {"name":"Delivered",       "icon":"📦", "key":"delivered"},
        {"name":"Pickup Reminder", "icon":"⏰", "key":"reminder"},
        {"name":"Payment Due",     "icon":"💰", "key":""},
        {"name":"Festival Wishes", "icon":"🎉", "key":""},
        {"name":"Eid Mubarak",     "icon":"🌙", "key":""},
        {"name":"New Collection",  "icon":"✨", "key":""},
        {"name":"Shop Closed",     "icon":"🚪", "key":""},
        {"name":"General Reminder","icon":"📢", "key":""},
        {"name":"Diwali Wishes",   "icon":"🪔", "key":""},
    ]

    return render_template("owner/whatsapp.html",
        active_page="whatsapp", show_voice=False,
        urgent_count=urgent_count,
        customers=customers,
        customers_json=json.dumps(customers),
        shop_name=shop_name,
        templates=templates,
        broadcast_log=broadcast_log,
        wa_templates=wa_templates,
        wa_templates_json=json.dumps({k: v for k, v in wa_templates.items()}),
    )


@bp.route("/api/whatsapp-log", methods=["POST"])
@owner_required
def whatsapp_log():
    data  = request.get_json(silent=True) or {}
    count = data.get("count", 1)
    btype = data.get("type", "broadcast")
    now   = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn  = get_db()
    # Log one entry per send session
    conn.execute(
        "INSERT INTO whatsapp_log(order_id,mobile,message_type,sent_at) VALUES(?,?,?,?)",
        (None, f"bulk:{count}", btype, now)
    )
    conn.commit(); conn.close()
    return jsonify({"ok":True})


# ══════════════════════════════════════════════
#  FINANCE CATEGORIES MANAGEMENT
# ══════════════════════════════════════════════

@bp.route("/api/finance-categories", methods=["GET"])
@owner_required
def get_finance_categories():
    income_cats  = get_setting("finance_income_cats",  "advance,payment,alteration,other income").split(",")
    expense_cats = get_setting("finance_expense_cats", "thread,buttons,fabric,electricity,rent,salary,transport,maintenance,other expense").split(",")
    return jsonify({
        "income":  [c.strip() for c in income_cats  if c.strip()],
        "expense": [c.strip() for c in expense_cats if c.strip()]
    })

@bp.route("/api/finance-categories/save", methods=["POST"])
@owner_required
def save_finance_categories():
    data    = request.get_json(silent=True) or {}
    cat_type = data.get("type","")   # "income" or "expense"
    cats    = data.get("categories", [])
    if cat_type not in ("income","expense"):
        return jsonify({"ok":False,"error":"Invalid type"})
    key = f"finance_{cat_type}_cats"
    set_setting(key, ",".join([c.strip() for c in cats if c.strip()]))
    return jsonify({"ok":True})


# ══════════════════════════════════════════════
#  WORK PROGRESS OVERVIEW (Admin)
# ══════════════════════════════════════════════

@bp.route("/work-progress")
@owner_required
def work_progress():
    conn = get_db()
    orders = conn.execute("""
        SELECT o.order_code, o.repeat_of, o.status, o.delivery_date, o.is_urgent,
               COALESCE(c.name,'—') as cname, COALESCE(c.mobile,'') as mobile,
               STRING_AGG(CAST(oi.garment_type||' x'||oi.quantity AS TEXT), ', ') as garments_str,
               SUM(oi.quantity) as total_qty
        FROM orders o
        LEFT JOIN customers c ON c.id=o.customer_id
        LEFT JOIN order_items oi ON oi.order_id=o.id
        WHERE o.status NOT IN ('delivered','cancelled')
        GROUP BY o.id, o.order_code, o.repeat_of, o.status, o.delivery_date, o.is_urgent, c.name, c.mobile
        ORDER BY o.delivery_date ASC, o.is_urgent DESC
    """).fetchall()

    def fmtd(d):
        if not d: return "—"
        p = str(d).split("-")
        return f"{p[2]}-{p[1]}-{p[0]}" if len(p)==3 else d

    result = []
    for o in orders:
        wl_rows = conn.execute(
            "SELECT garment_type, qty_done, notes FROM work_logs WHERE order_code=?",
            (o["order_code"],)
        ).fetchall()
        naap_done = kataai_done = silai_done = 0
        for wl in wl_rows:
            n = (wl["notes"] or "").strip()
            q = wl["qty_done"] or 0
            if any(x in n for x in ["Measurement","Naap","नाप"]):
                naap_done += q
            elif any(x in n for x in ["Kataai","Cutting","कटाई"]):
                kataai_done += q
            else:
                silai_done += q
        tq = o["total_qty"] or 1
        result.append({
            "order_code":    o["order_code"],
            "display_code":  o["repeat_of"] if o["repeat_of"] else o["order_code"],
            "entry_code":    o["order_code"] if o["repeat_of"] else "",
            "status":        o["status"],
            "cname":         o["cname"],
            "mobile":        o["mobile"],
            "delivery_date": fmtd(o["delivery_date"]),
            "is_urgent":     o["is_urgent"],
            "garments":      o["garments_str"] or "—",
            "total_qty":     tq,
            "naap_done":     min(naap_done, tq),
            "cut_done":      min(kataai_done, tq),
            "stitch_done":   min(silai_done, tq),
            "naap_pct":      min(100, int((min(naap_done,tq)/tq)*100)),
            "cut_pct":       min(100, int((min(kataai_done,tq)/tq)*100)),
            "stitch_pct":    min(100, int((min(silai_done,tq)/tq)*100)),
        })

    urgent_count = conn.execute(
        "SELECT COUNT(*) as c FROM orders WHERE is_urgent=1 AND status!='delivered'"
    ).fetchone()["c"]
    conn.close()
    return render_template_string(WORK_PROGRESS_PAGE,
        active_page="work_progress", show_voice=False,
        urgent_count=urgent_count, orders=result, total=len(result))



# ══════════════════════════════════════════════
#  DESIGN GALLERY (Admin)
# ══════════════════════════════════════════════

@bp.route("/gallery")
@owner_required
def gallery_admin():
    conn = get_db()
    # Tables already created by init_db
    try:
        types = conn.execute("SELECT * FROM gallery_types ORDER BY parent_id NULLS FIRST, sort_order, id").fetchall()
        images = conn.execute("SELECT * FROM gallery_images ORDER BY type_id, sort_order, id").fetchall()
    except Exception:
        types = []
        images = []
    urgent_count = conn.execute("SELECT COUNT(*) as c FROM orders WHERE is_urgent=1 AND status!='delivered'").fetchone()["c"]
    conn.close()
    return render_template("owner/gallery_admin.html",
        active_page="gallery_admin", show_voice=False,
        urgent_count=urgent_count, types=types, images=images)

@bp.route("/gallery/add-type", methods=["POST"])
@owner_required
def gallery_add_type():
    name      = request.form.get("name","").strip()
    parent_id = request.form.get("parent_id","").strip() or None
    if name:
        conn = get_db()
        conn.execute("INSERT INTO gallery_types(name,parent_id) VALUES(?,?)", (name, parent_id))
        conn.commit(); conn.close()
        flash(f"Type '{name}' added.", "success")
    return redirect(url_for("owner.gallery_admin"))

@bp.route("/gallery/delete-type/<int:tid>", methods=["POST"])
@owner_required
def gallery_delete_type(tid):
    conn = get_db()
    conn.execute("DELETE FROM gallery_images WHERE type_id=?", (tid,))
    conn.execute("DELETE FROM gallery_types WHERE id=?", (tid,))
    conn.execute("DELETE FROM gallery_types WHERE parent_id=?", (tid,))
    conn.commit(); conn.close()
    flash("Type deleted.", "success")
    return redirect(url_for("owner.gallery_admin"))

@bp.route("/gallery/upload-image", methods=["POST"])
@owner_required
def gallery_upload_image():
    import os as _os, time as _time
    try:
        type_id = request.form.get("type_id","").strip()
        caption = request.form.get("caption","").strip()
        file    = request.files.get("image")
        if not type_id or not file or not file.filename:
            flash("Please select a category and image.", "error")
            return redirect(url_for("owner.gallery_admin"))

        use_cloudinary = bool(_os.environ.get("CLOUDINARY_CLOUD_NAME"))
        if use_cloudinary:
            import cloudinary, cloudinary.uploader
            cloudinary.config(
                cloud_name=_os.environ.get("CLOUDINARY_CLOUD_NAME"),
                api_key=_os.environ.get("CLOUDINARY_API_KEY"),
                api_secret=_os.environ.get("CLOUDINARY_API_SECRET")
            )
            result = cloudinary.uploader.upload(
                file,
                folder="uttam_tailors/gallery",
                public_id=f"gal_{type_id}_{int(_time.time())}",
                overwrite=True
            )
            fname = result.get("secure_url","")
        else:
            ext = _os.path.splitext(file.filename)[1].lower()
            if ext not in [".jpg",".jpeg",".png",".gif",".webp"]: ext = ".jpg"
            folder = _os.path.join(_os.path.dirname(_os.path.dirname(_os.path.dirname(__file__))), "static", "order_images", "gallery")
            _os.makedirs(folder, exist_ok=True)
            fname = f"gal_{type_id}_{int(_time.time())}{ext}"
            file.save(_os.path.join(folder, fname))

        conn = get_db()
        conn.execute("INSERT INTO gallery_images(type_id,filename,caption) VALUES(?,?,?)", (type_id, fname, caption))
        conn.commit(); conn.close()
        flash("Image uploaded successfully.", "success")
    except Exception as e:
        flash(f"Upload failed: {str(e)}", "error")
    return redirect(url_for("owner.gallery_admin"))

@bp.route("/gallery/delete-image/<int:iid>", methods=["POST"])
@owner_required
def gallery_delete_image(iid):
    import os as _os
    conn = get_db()
    img = conn.execute("SELECT filename FROM gallery_images WHERE id=?", (iid,)).fetchone()
    if img:
        fpath = _os.path.join(Config.UPLOAD_FOLDER, "gallery", img["filename"])
        if _os.path.exists(fpath):
            _os.remove(fpath)
        conn.execute("DELETE FROM gallery_images WHERE id=?", (iid,))
        conn.commit()
    conn.close()
    flash("Image deleted.", "success")
    return redirect(url_for("owner.gallery_admin"))

@bp.route("/api/gallery")
def api_gallery():
    """Public API for employee gallery page"""
    conn = get_db()
    try:
        types  = conn.execute("SELECT * FROM gallery_types ORDER BY parent_id NULLS FIRST, sort_order, id").fetchall()
        images = conn.execute("SELECT * FROM gallery_images ORDER BY type_id, sort_order, id").fetchall()
    except:
        conn.close()
        return jsonify({"types":[], "images":[]})
    conn.close()
    return jsonify({
        "types":  [{"id":t["id"],"name":t["name"],"parent_id":t["parent_id"]} for t in types],
        "images": [{"id":i["id"],"type_id":i["type_id"],"filename":i["filename"],"caption":i["caption"]} for i in images]
    })



# ══════════════════════════════════════════════
#  CLEANUP OLD BASE64 DATA (one-time fix)
# ══════════════════════════════════════════════

@bp.route("/cleanup-images", methods=["POST"])
@owner_required
def cleanup_images():
    """Remove old base64 image data from database - run once to fix slowness"""
    conn = get_db()
    try:
        # Delete rows where file_path starts with data: (base64) or temp:
        conn.execute("DELETE FROM order_images WHERE file_path LIKE 'data:%'")
        conn.execute("DELETE FROM order_images WHERE file_path LIKE 'temp:%'")
        conn.commit()
        deleted = conn.execute("SELECT changes() as c").fetchone()
        conn.close()
        flash("✅ Old image data cleaned up. System should be faster now.", "success")
    except Exception as e:
        try: conn.close()
        except: pass
        flash(f"Cleanup done (some errors ignored).", "success")
    return redirect(url_for("owner.settings"))

# ══════════════════════════════════════════════
#  DATABASE BACKUP / RESTORE
# ══════════════════════════════════════════════

@bp.route("/backup/download")
@owner_required
def backup_download():
    """Download database backup - exports PostgreSQL data as JSON"""
    import json as _json, os as _os, tempfile
    from flask import send_file
    try:
        conn = get_db()
        backup_data = {}
        tables = ["customers","orders","order_items","order_images","work_logs",
                  "finance","employees","settings","salary_advances","gallery_types",
                  "gallery_images","measurement_fields","notify_log"]
        for table in tables:
            try:
                rows = conn.execute(f"SELECT * FROM {table}").fetchall() if table in ["orders","customers","order_items","finance","work_logs","settings","salary_advances","order_images","notify_log"] else []
                backup_data[table] = [dict(zip(r.keys(), r.values())) for r in rows]
            except Exception:
                backup_data[table] = []
        conn.close()
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".json", mode="w")
        _json.dump(backup_data, tmp, default=str, ensure_ascii=False)
        tmp.close()
        from datetime import datetime
        fname = f"uttam_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        set_setting("last_backup_at", datetime.now().strftime("%Y-%m-%d %H:%M"))
        return send_file(tmp.name, as_attachment=True, download_name=fname, mimetype="application/json")
    except Exception as e:
        flash(f"Backup failed: {str(e)}", "error")
        return redirect(url_for("owner.settings"))

@bp.route("/backup/restore", methods=["POST"])
@owner_required
def backup_restore():
    """Restore database from uploaded JSON backup file."""
    import json as _json, os as _os
    try:
        file = request.files.get("db_file")
        if not file or not file.filename:
            flash("Please select a backup file.", "error")
            return redirect(url_for("owner.settings"))
        if not file.filename.endswith(".json"):
            flash("Invalid file. Must be a .json backup file.", "error")
            return redirect(url_for("owner.settings"))

        backup_data = _json.load(file)
        if not isinstance(backup_data, dict):
            flash("Invalid backup format.", "error")
            return redirect(url_for("owner.settings"))

        conn = get_db()
        tables = ["customers","orders","order_items","order_images","work_logs",
                  "finance","employees","settings","salary_advances","gallery_types",
                  "gallery_images","measurement_fields","notify_log"]
        for table in tables:
            rows = backup_data.get(table, [])
            if not rows:
                continue
            # Clear existing data
            try:
                conn.execute(f"DELETE FROM {table}") if table in ["orders","customers","order_items","finance","work_logs","settings","salary_advances","order_images","notify_log"] else None
            except Exception:
                continue
            # Insert rows
            for row in rows:
                cols = list(row.keys())
                placeholders = ", ".join(["?"] * len(cols))
                col_names = ", ".join(cols)
                vals = [row[c] for c in cols]
                try:
                    conn.execute(f"INSERT INTO {table} ({col_names}) VALUES ({placeholders})", vals) if table in ["orders","customers","order_items","finance","work_logs","settings","salary_advances","order_images","notify_log"] else None
                except Exception:
                    continue
        conn.commit()
        conn.close()
        flash("✅ Database restored successfully! All your orders and data are back.", "success")
    except Exception as e:
        flash(f"Restore failed: {str(e)}", "error")
    return redirect(url_for("owner.settings"))

# ══════════════════════════════════════════════
#  FACTORY RESET
# ══════════════════════════════════════════════

@bp.route("/reset", methods=["POST"])
@owner_required
def factory_reset():
    password = request.form.get("reset_password","").strip()
    if password != "8899":
        flash("Incorrect reset password.", "error")
        return redirect(url_for("owner.settings"))
    try:
        conn = get_db()
        conn.execute("DELETE FROM work_logs")
        conn.execute("DELETE FROM order_items")
        conn.execute("DELETE FROM finance")
        conn.execute("DELETE FROM orders")
        conn.execute("DELETE FROM customers")
        conn.execute("DELETE FROM salary_advances")
        try:
            conn.execute("DELETE FROM notify_log")
        except: pass
        # Reset order code counter using same connection
        conn.execute("UPDATE settings SET value='3599' WHERE key='last_order_code'")
        conn.execute("INSERT INTO settings (key,value) VALUES ('last_order_code','3599') ON CONFLICT DO NOTHING")
        conn.commit()
        conn.close()
        flash("✅ System reset! All orders and customers cleared. Rates and settings kept.", "success")
    except Exception as e:
        flash(f"Reset failed: {str(e)}", "error")
    return redirect(url_for("owner.settings"))




# ══════════════════════════════════════════════
#  API: ORDER DETAIL (for expandable row)
# ══════════════════════════════════════════════

@bp.route("/api/order-detail/<order_code>")
@owner_required
def api_order_detail(order_code):
    import json as _json, os as _os
    from config import Config
    conn = get_db()
    o = conn.execute("""
        SELECT o.*, c.name as cname, c.mobile, c.address
        FROM orders o LEFT JOIN customers c ON c.id=o.customer_id
        WHERE o.order_code=?
    """, (order_code,)).fetchone()
    if not o:
        conn.close()
        return jsonify({"error": "Not found"})

    items = conn.execute("SELECT * FROM order_items WHERE order_id=?", (o["id"],)).fetchall()
    wl_rows = conn.execute("SELECT qty_done, notes FROM work_logs WHERE order_code=?", (order_code,)).fetchall()

    # Progress
    naap=kataai=silai=0
    for wl in wl_rows:
        n=(wl["notes"] or "").strip(); q=wl["qty_done"] or 0
        if any(x in n for x in ["Measurement","Naap","नाप"]): naap+=q
        elif any(x in n for x in ["Kataai","Cutting","कटाई"]): kataai+=q
        else: silai+=q
    tq=sum(it["quantity"] for it in items) or 1

    garments=[]
    for it in items:
        try: meas=_json.loads(it["measurements"] or "{}")
        except: meas={}
        garments.append({"type":it["garment_type"],"qty":it["quantity"],"rate":it["rate"],"amount":it["amount"],"measurements":meas,"notes":it["notes"] or ""})

    # Images - DB first (Cloudinary), then filesystem fallback
    images=[]
    img_rows=conn.execute("SELECT file_path FROM order_images WHERE order_id=? ORDER BY id",(o["id"],)).fetchall()
    images=[r["file_path"] for r in img_rows if r["file_path"] and not r["file_path"].startswith("temp:")]
    conn.close()

    if not images:
        folder=_os.path.join(Config.UPLOAD_FOLDER, order_code)
        if _os.path.isdir(folder):
            images=[f"/static/order_images/{order_code}/{f}" for f in sorted(_os.listdir(folder)) if f.lower().endswith((".jpg",".jpeg",".png",".gif",".webp"))]

    return jsonify({
        "order": {
            "order_code": o["order_code"], "status": o["status"],
            "cname": o["cname"] or "—", "mobile": o["mobile"] or "—", "address": o["address"] or "—",
            "payable": o["payable_amount"] or 0, "advance": o["advance_paid"] or 0, "remaining": o["remaining"] or 0,
            "naap_pct": min(100,int(naap/tq*100)), "cut_pct": min(100,int(kataai/tq*100)), "stitch_pct": min(100,int(silai/tq*100)),
        },
        "garments": garments,
        "images": images
    })

# ══════════════════════════════════════════════
#  ADMIN ORDER DETAIL PAGE
# ══════════════════════════════════════════════

@bp.route("/orders/detail/<order_code>")
@owner_required
def order_detail(order_code):
    conn = get_db()
    import json as _json, os as _os
    from config import Config

    o = conn.execute("""
        SELECT o.*, c.name as cname, c.mobile, c.address,
               COALESCE(o.note,'') as note
        FROM orders o
        LEFT JOIN customers c ON c.id=o.customer_id
        WHERE o.order_code=?
    """, (order_code,)).fetchone()

    if not o:
        conn.close()
        return "<h2>Order not found</h2>", 404

    items = conn.execute(
        "SELECT * FROM order_items WHERE order_id=?", (o["id"],)
    ).fetchall()

    wl_rows = conn.execute("""
        SELECT garment_type, qty_done, notes, employee_name, log_date, making_rate
        FROM work_logs WHERE order_code=? ORDER BY log_date, id
    """, (order_code,)).fetchall()

    finance = conn.execute("""
        SELECT tx_date, tx_type, category, amount, mode, note
        FROM finance WHERE order_id=? ORDER BY id
    """, (o["id"],)).fetchall()

    urgent_count = conn.execute(
        "SELECT COUNT(*) as c FROM orders WHERE is_urgent=1 AND status!='delivered'"
    ).fetchone()["c"]
    conn.close()

    # Work progress
    naap_done = kataai_done = silai_done = 0
    for wl in wl_rows:
        n = (wl["notes"] or "").strip()
        q = wl["qty_done"] or 0
        if any(x in n for x in ["Measurement","Naap","नाप"]):
            naap_done += q
        elif any(x in n for x in ["Kataai","Cutting","कटाई"]):
            kataai_done += q
        else:
            silai_done += q

    total_qty = sum(it["quantity"] for it in items) or 1

    def fmtd(d):
        if not d: return "—"
        p = str(d).split("-")
        return f"{p[2]}-{p[1]}-{p[0]}" if len(p)==3 else d

    # Images from filesystem
    images = []
    folder = _os.path.join(Config.UPLOAD_FOLDER, order_code)
    if _os.path.isdir(folder):
        images = [f"/static/order_images/{order_code}/{f}"
                  for f in sorted(_os.listdir(folder))
                  if f.lower().endswith((".jpg",".jpeg",".png",".gif",".webp"))]

    garments = []
    for it in items:
        try:
            meas = _json.loads(it["measurements"] or "{}")
        except:
            meas = {}
        garments.append({
            "type": it["garment_type"],
            "qty":  it["quantity"],
            "rate": it["rate"],
            "amount": it["amount"],
            "measurements": meas,
            "notes": it["notes"] or ""
        })

    order_data = {
        "order_code":    o["order_code"],
        "display_code":  o["repeat_of"] if o["repeat_of"] else o["order_code"],
        "entry_code":    o["order_code"] if o["repeat_of"] else "",
        "status":        o["status"],
        "is_urgent":     o["is_urgent"],
        "note":          o["note"],
        "order_date":    fmtd(o["order_date"]),
        "delivery_date": fmtd(o["delivery_date"]),
        "delivered_at":  fmtd((o["delivered_at"] or "")[:10]),
        "cname":         o["cname"] or "—",
        "mobile":        o["mobile"] or "—",
        "address":       o["address"] or "—",
        "total_amount":  o["total_amount"] or 0,
        "extra_charges": o["extra_charges"] or 0,
        "payable":       o["payable_amount"] or 0,
        "advance":       o["advance_paid"] or 0,
        "remaining":     o["remaining"] or 0,
        "payment_mode":  o["payment_mode"] or "cash",
        "naap_pct":      min(100, int(naap_done/total_qty*100)),
        "cut_pct":       min(100, int(kataai_done/total_qty*100)),
        "stitch_pct":    min(100, int(silai_done/total_qty*100)),
        "total_qty":     total_qty,
    }

    return render_template("owner/order_detail.html",
        active_page="owner_orders", show_voice=False,
        urgent_count=urgent_count,
        order=order_data, garments=garments,
        work_logs=wl_rows, finance=finance, images=images,
        fmtd=fmtd)

# ══════════════════════════════════════════════
#  ORDER CANCEL
# ══════════════════════════════════════════════

@bp.route("/orders/cancel/<order_code>", methods=["POST"])
@owner_required
def cancel_order(order_code):
    conn = get_db()
    conn.execute("UPDATE orders SET status='cancelled' WHERE order_code=?", (order_code,))
    conn.commit()
    conn.close()
    flash(f"Order #{order_code} cancelled.", "success")
    return redirect(request.referrer or url_for("owner.owner_dashboard"))


# ══════════════════════════════════════════════
#  WHATSAPP ORDER TEMPLATE SAVE
# ══════════════════════════════════════════════

@bp.route("/api/wa-templates")
def api_wa_templates():
    """Return saved WA templates (public — no auth needed for employee pages)."""
    _D_CONFIRM_HI = "*उत्तम टेलर्स*\n\nनमस्ते *{name}*,\n\nआपका ऑर्डर सफलतापूर्वक दर्ज हो गया है।\n------------------------------------------\n\n- ऑर्डर नंबर: *#{code}*\n- ऑर्डर तारीख: {odate}\n- डिलीवरी तारीख: *{ddate}*\n\n- कपड़े: {items}\n\n- कुल राशि: *₹{total}*\n- अग्रिम भुगतान: ₹{advance}\n- बकाया राशि: *₹{due}*\n- भुगतान माध्यम: {mode}\n\n-----------------------------------------------\nआपके विश्वास के लिए धन्यवाद। 🙏🏻\nउत्तम टेलर्स"
    _D_CONFIRM_EN = "*Uttam Tailors*\n\nHello *{name}*,\nYour order has been confirmed successfully.\n------------------------------------------\n\n- Order No: *#{code}*\n- Order Date: {odate}\n- Delivery Date: *{ddate}*\n\n- Garments: {items}\n\n- Total Amount: *₹{total}*\n- Advance Paid: ₹{advance}\n- Balance Due: *₹{due}*\n- Payment Mode: {mode}\n\n-----------------------------------------------\nThank you for your trust. 🙏🏻\nUttam Tailors"
    _D_READY_HI   = "*उत्तम टेलर्स*\n\nनमस्ते *{name}*,\nआपका ऑर्डर तैयार है। कृपया अपनी सुविधानुसार आकर अपने कपड़े प्राप्त कर लें।\n------------------------------------------\n\n- ऑर्डर नंबर: *#{code}*\n- कपड़े: {items}\n- कुल राशि: *₹{total}*\n- बकाया राशि: *₹{due}*\n\n-----------------------------------------------\nआपके विश्वास के लिए धन्यवाद। 🙏🏻\nउत्तम टेलर्स"
    _D_READY_EN   = "*Uttam Tailors*\n\nHello *{name}*,\nYour order is ready. Please collect your clothes at your convenience. Thank you!\n------------------------------------------\n\n- Order No: *#{code}*\n- Garments: {items}\n- Total: *₹{total}*\n- Balance Due: *₹{due}*\n\n-----------------------------------------------\nThank you for your trust. 🙏🏻\nUttam Tailors"
    _D_DELIV_HI   = "*उत्तम टेलर्स*\n\nनमस्ते *{name}*,\nआपके कपड़े सफलतापूर्वक डिलीवर कर दिए गए हैं।\n------------------------------------------\n\n- ऑर्डर नंबर: *#{code}*\n- ऑर्डर तारीख: {odate}\n- डिलीवरी तारीख: {ddate}\n\n- कपड़े: {items}\n\n- कुल भुगतान: *₹{total}*\n- भुगतान माध्यम: {mode}\n\n-----------------------------------------------\nआपके विश्वास के लिए धन्यवाद। 🙏🏻\nआगे भी सेवा का अवसर दें, यही हमारी शुभकामना है।\n\nधन्यवाद!\nउत्तम टेलर्स"
    _D_DELIV_EN   = "*Uttam Tailors*\n\nHello *{name}*,\nYour garments have been delivered successfully.\n------------------------------------------\n\n- Order No: *#{code}*\n- Order Date: {odate}\n- Delivery Date: {ddate}\n\n- Garments: {items}\n\n- Total Paid: *₹{total}*\n- Payment Mode: {mode}\n\n-----------------------------------------------\nThank you for your trust. 🙏🏻\nWe look forward to serving you again.\n\nThank you!\nUttam Tailors"
    return jsonify({
        "confirm":   {"hi": get_setting("wa_order_confirm_hi",   _D_CONFIRM_HI),
                      "en": get_setting("wa_order_confirm_en",   _D_CONFIRM_EN)},
        "ready":     {"hi": get_setting("wa_order_ready_hi",     _D_READY_HI),
                      "en": get_setting("wa_order_ready_en",     _D_READY_EN)},
        "delivered": {"hi": get_setting("wa_order_delivered_hi", _D_DELIV_HI),
                      "en": get_setting("wa_order_delivered_en", _D_DELIV_EN)},
    })


@bp.route("/api/save-wa-template", methods=["POST"])
@owner_required
def save_wa_template():
    data = request.get_json(silent=True) or {}
    tmpl = data.get("template", "confirm")
    key_map = {
        "confirm":   ("wa_order_confirm_en",   "wa_order_confirm_hi"),
        "ready":     ("wa_order_ready_en",     "wa_order_ready_hi"),
        "delivered": ("wa_order_delivered_en", "wa_order_delivered_hi"),
        "reminder":  ("wa_order_reminder_en",  "wa_order_reminder_hi"),
    }
    keys = key_map.get(tmpl, key_map["confirm"])
    if "en" in data: set_setting(keys[0], data["en"])
    if "hi" in data: set_setting(keys[1], data["hi"])
    return jsonify({"ok": True})


# ══════════════════════════════════════════════
#  ORDER MANAGEMENT (Owner)
# ══════════════════════════════════════════════

@bp.route("/orders")
@owner_required
def owner_orders():
    conn = get_db()
    try:
        rows = conn.execute("""
            SELECT o.order_code, o.status, o.order_date, o.delivery_date,
                   COALESCE(o.payable_amount,0) as payable_amount,
                   COALESCE(o.remaining,0) as remaining,
                   COALESCE(o.is_urgent,0) as is_urgent,
                   COALESCE(o.note,'') as note,
                   COALESCE(o.repeat_of,'') as repeat_of,
                   COALESCE(c.name,'—') as cname, COALESCE(c.mobile,'') as mobile,
                   STRING_AGG(CAST(oi.garment_type||' x'||oi.quantity AS TEXT), ', ') as garments_str
            FROM orders o
            LEFT JOIN customers c ON c.id=o.customer_id
            LEFT JOIN order_items oi ON oi.order_id=o.id
            GROUP BY o.id, o.order_code, o.status, o.order_date, o.delivery_date,
                     o.total_amount, o.extra_charges, o.payable_amount, o.advance_paid,
                     o.remaining, o.payment_mode, o.is_urgent, o.note, o.repeat_of,
                     o.delivered_at, o.created_at, c.name, c.mobile, c.address
            ORDER BY o.id DESC
        """).fetchall()
    except Exception as e:
        conn.close()
        return f"<h2>DB Error in /owner/orders</h2><pre>{e}</pre>", 500

    def fmtd(d):
        if not d: return "—"
        p = str(d).split("-")
        return f"{p[2]}-{p[1]}-{p[0]}" if len(p)==3 else d

    orders = [{
        "order_code":    r["order_code"],
        "display_code":  r["repeat_of"] if r["repeat_of"] else r["order_code"],
        "entry_code":    r["order_code"] if r["repeat_of"] else "",
        "repeat_of":     r["repeat_of"] or "",
        "status":        r["status"],
        "order_date":    fmtd(r["order_date"]),
        "delivery_date": fmtd(r["delivery_date"]),
        "delivery_date_iso": (r["delivery_date"] or "").replace("-",""),
        "payable":       r["payable_amount"] or 0,
        "remaining":     r["remaining"] or 0,
        "is_urgent":     r["is_urgent"],
        "note":          r["note"] or "",
        "cname":         r["cname"] or "—",
        "mobile":        r["mobile"] or "",
        "garments":      r["garments_str"] or "—"
    } for r in rows]

    # Work progress data for combined page
    try:
        wp_rows = conn.execute("""
            SELECT o.order_code, o.repeat_of, o.status, o.delivery_date, o.is_urgent,
                   COALESCE(c.name,'—') as cname, COALESCE(c.mobile,'') as mobile,
                   STRING_AGG(CAST(oi.garment_type||' x'||oi.quantity AS TEXT), ', ') as garments_str,
                   SUM(oi.quantity) as total_qty
            FROM orders o
            LEFT JOIN customers c ON c.id=o.customer_id
            LEFT JOIN order_items oi ON oi.order_id=o.id
            WHERE o.status NOT IN ('delivered','cancelled')
            GROUP BY o.id, o.order_code, o.repeat_of, o.status, o.delivery_date, o.is_urgent, c.name, c.mobile
            ORDER BY o.delivery_date ASC, o.is_urgent DESC
        """).fetchall()
    except:
        wp_rows = []

    wp_result = []
    for o in wp_rows:
        wl = conn.execute("SELECT qty_done, notes FROM work_logs WHERE order_code=?", (o["order_code"],)).fetchall()
        naap=kataai=silai=0
        for w in wl:
            n=(w["notes"] or "").strip(); q=w["qty_done"] or 0
            if any(x in n for x in ["Measurement","Naap","नाप"]): naap+=q
            elif any(x in n for x in ["Kataai","Cutting","कटाई"]): kataai+=q
            else: silai+=q
        tq=o["total_qty"] or 1
        wp_result.append({
            "order_code": o["order_code"],
            "display_code": o["repeat_of"] if o["repeat_of"] else o["order_code"],
            "entry_code": o["order_code"] if o["repeat_of"] else "",
            "status": o["status"],
            "cname": o["cname"], "mobile": o["mobile"],
            "delivery_date": fmtd(o["delivery_date"]),
            "is_urgent": o["is_urgent"], "garments": o["garments_str"] or "—",
            "total_qty": tq,
            "naap_pct": min(100,int(min(naap,tq)/tq*100)),
            "cut_pct":  min(100,int(min(kataai,tq)/tq*100)),
            "stitch_pct": min(100,int(min(silai,tq)/tq*100)),
        })

    urgent_count = conn.execute(
        "SELECT COUNT(*) as c FROM orders WHERE is_urgent=1 AND status!='delivered'"
    ).fetchone()["c"]
    conn.close()

    filter_mode = request.args.get("filter", "")
    return render_template_string(ORDERS_PAGE,
        active_page="owner_orders", show_voice=False,
        urgent_count=urgent_count, orders=orders, total=len(orders),
        filter_mode=filter_mode, wp_orders=wp_result)

# ══════════════════════════════════════════════
#  ORDER DELETE (Owner)
# ══════════════════════════════════════════════

@bp.route("/orders/delete/<order_code>", methods=["POST"])
@owner_required
def delete_order(order_code):
    conn = get_db()
    order = conn.execute("SELECT id FROM orders WHERE order_code=?", (order_code,)).fetchone()
    if order:
        conn.execute("DELETE FROM work_logs WHERE order_id=?", (order["id"],))
        conn.execute("DELETE FROM work_logs WHERE order_code=?", (order_code,))
        conn.execute("DELETE FROM order_items WHERE order_id=?", (order["id"],))
        conn.execute("DELETE FROM order_images WHERE order_id=?", (order["id"],))
        conn.execute("DELETE FROM finance WHERE order_id=?", (order["id"],))
        conn.execute("DELETE FROM notify_log WHERE order_code=?", (order_code,))
        conn.execute("DELETE FROM orders WHERE id=?", (order["id"],))
        conn.commit()
        conn.close()
        # Release the number ONLY if it's the most recently issued code (undo-style)
        from database import release_order_code_if_latest
        release_order_code_if_latest(order_code)
        flash(f"Order #{order_code} deleted.", "success")
    else:
        conn.close()
        flash(f"Order #{order_code} not found.", "error")
    return redirect(request.referrer or url_for("owner.owner_dashboard"))


@bp.route("/customers/<int:customer_id>/delete", methods=["POST"])
@owner_required
def delete_customer(customer_id):
    """Delete a customer and ALL their orders, work logs, images, finance entries."""
    conn = get_db()
    cust = conn.execute("SELECT name FROM customers WHERE id=?", (customer_id,)).fetchone()
    if not cust:
        conn.close()
        flash("Customer not found.", "error")
        return redirect(url_for("owner.owner_dashboard"))

    cust_name = cust["name"]
    # Get all orders for this customer
    orders = conn.execute("SELECT id, order_code FROM orders WHERE customer_id=?", (customer_id,)).fetchall()
    for o in orders:
        conn.execute("DELETE FROM work_logs WHERE order_id=?", (o["id"],))
        conn.execute("DELETE FROM work_logs WHERE order_code=?", (o["order_code"],))
        conn.execute("DELETE FROM order_items WHERE order_id=?", (o["id"],))
        conn.execute("DELETE FROM order_images WHERE order_id=?", (o["id"],))
        conn.execute("DELETE FROM finance WHERE order_id=?", (o["id"],))
        conn.execute("DELETE FROM notify_log WHERE order_code=?", (o["order_code"],))
    conn.execute("DELETE FROM orders WHERE customer_id=?", (customer_id,))
    conn.execute("DELETE FROM customers WHERE id=?", (customer_id,))
    conn.commit()
    conn.close()
    # Release order codes that were the latest issued (undo-style) — never recycle old ones
    from database import release_order_code_if_latest
    for o in orders:
        release_order_code_if_latest(o["order_code"])
    flash(f"Customer '{cust_name}' and {len(orders)} order(s) deleted permanently.", "success")
    return redirect(url_for("owner.owner_customers"))


# ══════════════════════════════════════════════
#  CUSTOMER DETAIL (Owner)
# ══════════════════════════════════════════════

@bp.route("/customers/<int:customer_id>")
@owner_required
def owner_customer_detail(customer_id):
    conn = get_db()
    cust = conn.execute("SELECT * FROM customers WHERE id=?", (customer_id,)).fetchone()
    if not cust:
        conn.close()
        return "Customer not found", 404

    orders = conn.execute("""
        SELECT o.*, STRING_AGG(CAST(oi.garment_type||' x'||oi.quantity AS TEXT), ', ') as garments_str
        FROM orders o LEFT JOIN order_items oi ON oi.order_id=o.id
        WHERE o.customer_id=? GROUP BY o.id, o.order_code, o.status, o.order_date,
                 o.delivery_date, o.payable_amount, o.advance_paid, o.remaining,
                 o.payment_mode, o.is_urgent, o.note, o.delivered_at
        ORDER BY o.id DESC
    """, (customer_id,)).fetchall()

    def fmtd(d):
        if not d: return "—"
        p = str(d).split("-")
        return f"{p[2]}-{p[1]}-{p[0]}" if len(p)==3 else d

    urgent_count = conn.execute(
        "SELECT COUNT(*) as c FROM orders WHERE is_urgent=1 AND status!='delivered'"
    ).fetchone()["c"]
    conn.close()

    latest_id = max((o["id"] for o in orders), default=None)

    orders_list = [{
        "order_code":       o["order_code"],
        "display_code":     o["repeat_of"] if o["repeat_of"] else o["order_code"],
        "entry_code":       o["order_code"] if o["repeat_of"] else "",
        "is_latest":        o["id"] == latest_id,
        "status":           o["status"],
        "order_date_fmt":   fmtd(o["order_date"]),
        "delivery_date_fmt":fmtd(o["delivery_date"]),
        "payable_amount":   o["payable_amount"] or 0,
        "advance_paid":     o["advance_paid"] or 0,
        "remaining":        o["remaining"] or 0,
        "is_urgent":        o["is_urgent"],
        "note":             o["note"] or "",
        "garments_str":     o["garments_str"] or "",
    } for o in orders]

    total_billed = sum(o["payable_amount"] for o in orders_list)
    total_due    = sum(o["remaining"] for o in orders_list)

    return render_template("owner/customer_detail.html",
        active_page="customers", show_voice=False,
        urgent_count=urgent_count,
        cust={"id":cust["id"],"name":cust["name"],"mobile":cust["mobile"] or "",
              "address":cust["address"] or ""},
        orders=orders_list,
        total_billed=int(total_billed), total_due=int(total_due)
    )


# ══════════════════════════════════════════════
#  CUSTOMER EDIT (Owner)
# ══════════════════════════════════════════════

@bp.route("/customers/<int:customer_id>/edit", methods=["POST"])
@owner_required
def owner_customer_edit(customer_id):
    name    = request.form.get("name","").strip()
    mobile  = request.form.get("mobile","").strip()
    address = request.form.get("address","").strip()
    if not name:
        flash("Name is required", "error")
        return redirect(url_for("owner.owner_customer_detail", customer_id=customer_id))
    conn = get_db()
    conn.execute("UPDATE customers SET name=?,mobile=?,address=? WHERE id=?",
                 (name, mobile, address, customer_id))
    conn.commit()
    conn.close()
    flash("Customer updated!", "success")
    return redirect(url_for("owner.owner_customer_detail", customer_id=customer_id))


@bp.route("/notify-log")
@owner_required
def notify_log_view():
    """Owner sees all WhatsApp notifies sent."""
    conn = get_db()
    logs = conn.execute("""
        SELECT n.*, o.repeat_of
        FROM notify_log n LEFT JOIN orders o ON o.order_code = n.order_code
        ORDER BY n.sent_at DESC
    """).fetchall()
    urgent_count = conn.execute(
        "SELECT COUNT(*) as c FROM orders WHERE is_urgent=1 AND status!='delivered'"
    ).fetchone()["c"]
    conn.close()

    def fmtdt(d):
        if not d: return "—"
        parts = str(d).split(" ")
        dp = parts[0].split("-") if parts else []
        date_fmt = f"{dp[2]}-{dp[1]}-{dp[0]}" if len(dp)==3 else parts[0]
        time_fmt = parts[1][:5] if len(parts)>1 else ""
        return f"{date_fmt} {time_fmt}"

    log_list = [{"order_code":r["order_code"],
                 "display_code": r["repeat_of"] if r["repeat_of"] else r["order_code"],
                 "entry_code": r["order_code"] if r["repeat_of"] else "",
                 "customer":r["customer"],
                 "mobile":r["mobile"],"lang":r["lang"].upper(),"sent_at":fmtdt(r["sent_at"])}
                for r in logs]
    return render_template("owner/notify_log.html",
        active_page="whatsapp", show_voice=False,
        urgent_count=urgent_count, logs=log_list)



@bp.route("/api/salary/delete-advance", methods=["GET","POST"])
@owner_required
def api_delete_salary_advance():
    """Delete a specific salary advance entry — GET to preview, POST to delete."""
    name   = request.args.get("name", "").strip()
    amount = request.args.get("amount", "").strip()
    date_q = request.args.get("date", "").strip()

    conn = get_db()

    # Build query based on provided params
    conditions, params = [], []
    if name:
        conditions.append("LOWER(employee_name) LIKE ?")
        params.append(f"%{name.lower()}%")
    if amount:
        conditions.append("amount = ?")
        params.append(float(amount))
    if date_q:
        conditions.append("advance_date LIKE ?")
        params.append(f"{date_q}%")

    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
    rows = conn.execute(f"SELECT id, employee_name, amount, note, advance_date FROM salary_advances {where} ORDER BY id DESC LIMIT 20", params).fetchall()

    if request.method == "GET":
        # Preview mode
        table = "".join(f"<tr><td>{r['id']}</td><td>{r['employee_name']}</td><td>₹{r['amount']}</td><td>{r['note'] or ''}</td><td>{r['advance_date'] or ''}</td><td><form method='POST' action='/owner/api/salary/delete-advance?id={r['id']}' style='display:inline'><button style='background:#dc2626;color:#fff;border:none;padding:4px 10px;border-radius:6px;cursor:pointer;'>Delete</button></form></td></tr>" for r in rows)
        conn.close()
        return f"""<style>body{{font-family:sans-serif;padding:20px}} table{{border-collapse:collapse;width:100%}} td,th{{border:1px solid #e2e8f0;padding:8px;font-size:13px}}</style>
        <h2>🔍 Salary Advances — Preview</h2>
        <p>Found {len(rows)} entries</p>
        <table><tr><th>ID</th><th>Name</th><th>Amount</th><th>Note</th><th>Date</th><th>Action</th></tr>{table}</table>
        <br><a href='/owner/salary'>← Salary Page</a>"""

    # POST = delete by id
    del_id = request.args.get("id", "").strip()
    if del_id:
        conn.execute("DELETE FROM salary_advances WHERE id = ?", (del_id,))
        conn.commit()
        conn.close()
        return f"<h2>✅ Entry #{del_id} deleted from salary_advances!</h2><a href='/owner/salary'>← Salary Page</a><br><a href='/owner/api/salary/delete-advance?name={name}&amount={amount}'>← Back to list</a>"
    conn.close()
    return "<h2>❌ No id provided</h2>"

# ══════════════════════════════════════════════
#  SALARY MODULE
# ══════════════════════════════════════════════

@bp.route("/salary")
@owner_required
def salary():
    conn = get_db()
    period = request.args.get("period","month")
    today  = date.today().isoformat()

    fresh_start = get_setting("salary_fresh_start_date", "")

    if period == "week":
        from datetime import timedelta as td
        start = (date.today() - timedelta(days=date.today().weekday())).isoformat()
        period_label = f"This week ({start[8:]}-{start[5:7]} to {today[8:]}-{today[5:7]})"
    elif period == "all":
        start = "2000-01-01"
        period_label = "All time"
    elif period == "fresh":
        start = fresh_start if fresh_start else "2000-01-01"
        if fresh_start:
            p = fresh_start.split("-")
            period_label = f"Fresh start from {p[2]}-{p[1]}-{p[0]}"
        else:
            period_label = "Fresh start (not set)"
    else:
        period = "month"
        start = today[:7] + "-01"
        period_label = date.today().strftime("%B %Y")

    emps = conn.execute("SELECT name FROM employees WHERE active=1 ORDER BY name").fetchall()
    advances_all = conn.execute(
        "SELECT * FROM salary_advances ORDER BY advance_date DESC"
    ).fetchall()

    employees = []
    for emp in emps:
        name = emp["name"]
        logs = conn.execute("""
            SELECT wl.order_code, wl.garment_type, wl.qty_done, wl.making_rate, wl.log_date,
                   CAST(wl.qty_done AS REAL) * CAST(wl.making_rate AS REAL) as earning,
                   o.repeat_of
            FROM work_logs wl
            LEFT JOIN orders o ON o.order_code = wl.order_code
            WHERE wl.employee_name=? AND wl.log_date >= ?
            ORDER BY wl.log_date DESC, wl.id DESC
        """, (name, start)).fetchall()

        total_earned  = sum(r["earning"] or 0 for r in logs)
        total_pieces  = sum(r["qty_done"] or 0 for r in logs)
        total_orders  = len(set(r["order_code"] for r in logs if r["order_code"]))

        # Advances from salary_advances table (canonical source).
        # Both the Salary page "Record Advance" button AND the Finance page
        # salary expense sync route write here, so this captures everything.
        adv_period = conn.execute(
            "SELECT COALESCE(SUM(amount),0) as total FROM salary_advances WHERE employee_name=? AND advance_date >= ?",
            (name, start)
        ).fetchone()["total"] or 0

        employees.append({
            "name":         name,
            "total_pieces": total_pieces,
            "total_orders": total_orders,
            "total_earned": total_earned,
            "total_advance":adv_period,
            "net_payable":  total_earned - adv_period,
            "logs": [{"order_code":r["order_code"],
                      "display_code": r["repeat_of"] if r["repeat_of"] else r["order_code"],
                      "entry_code": r["order_code"] if r["repeat_of"] else "",
                      "garment_type":r["garment_type"],
                      "qty_done":r["qty_done"],"earning":r["earning"] or 0,
                      "log_date":r["log_date"]} for r in logs]
        })

    def fmtd(d):
        if not d: return "—"
        p = str(d).split("-")
        return f"{p[2]}-{p[1]}-{p[0]}" if len(p)==3 else d

    advances = [{"employee_name":a["employee_name"],"amount":a["amount"],
                 "note":a["note"] or "","advance_date":fmtd(a["advance_date"])} for a in advances_all]

    urgent_count = conn.execute(
        "SELECT COUNT(*) as c FROM orders WHERE is_urgent=1 AND status!='delivered'"
    ).fetchone()["c"]
    conn.close()

    return render_template("owner/salary.html",
        active_page="salary", show_voice=False,
        urgent_count=urgent_count,
        employees=employees, advances=advances,
        period=period, period_label=period_label,
        fresh_start=fresh_start)


@bp.route("/api/salary/advance", methods=["POST"])
@owner_required
def api_salary_advance():
    data   = request.get_json(silent=True) or {}
    name   = data.get("employee_name","").strip()
    amount = float(data.get("amount",0) or 0)
    note   = data.get("note","").strip() or f"Salary — {name}"
    if not name or amount <= 0:
        return jsonify({"ok":False,"error":"Name and amount required"})
    today = date.today().isoformat()
    now   = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn  = get_db()
    # Write to salary_advances (legacy/canonical table)
    conn.execute(
        "INSERT INTO salary_advances(employee_name,amount,note,advance_date,created_at) VALUES(?,?,?,?,?)",
        (name, amount, note, today, now)
    )
    # Also write to finance table so both pages stay in sync
    conn.execute("""
        INSERT INTO finance(tx_date,tx_type,category,amount,mode,note,employee_name,created_by,created_at)
        VALUES(?,'expense','salary',?,'cash',?,?,'owner',?)
    """, (today, amount, note, name, now))
    conn.commit(); conn.close()
    return jsonify({"ok":True})


@bp.route("/api/orders/change-code", methods=["POST"])
@owner_required
def api_change_order_code():
    data     = request.get_json(silent=True) or {}
    old_code = (data.get("old_code") or "").strip()
    new_code = (data.get("new_code") or "").strip()

    if not old_code or not new_code:
        return jsonify({"ok": False, "error": "Both old and new code required"})
    if old_code == new_code:
        return jsonify({"ok": False, "error": "Same as current code"})

    conn = get_db()

    # Check old code exists
    order = conn.execute("SELECT id FROM orders WHERE order_code=?", (old_code,)).fetchone()
    if not order:
        conn.close()
        return jsonify({"ok": False, "error": f"Order #{old_code} not found"})

    # Check new code not already taken
    existing = conn.execute("SELECT id FROM orders WHERE order_code=?", (new_code,)).fetchone()
    if existing:
        conn.close()
        return jsonify({"ok": False, "error": f"Order #{new_code} already exists — choose a different code"})

    # Update orders table
    conn.execute("UPDATE orders SET order_code=? WHERE order_code=?", (new_code, old_code))

    # Update finance notes that reference old code
    conn.execute("""
        UPDATE finance SET note = REPLACE(note, ?, ?)
        WHERE note LIKE ?
    """, (f"#{old_code}", f"#{new_code}", f"%{old_code}%"))

    conn.commit()
    conn.close()
    return jsonify({"ok": True})
@owner_required
def api_salary_sync_finance():
    """Sync Finance-page salary expenses into salary_advances table so the
    Salary page picks them up. Safe to run multiple times (deduplicates)."""
    conn = get_db()

    # Ensure employee_name column exists in finance table
    try:
        conn.execute("SAVEPOINT col_sp")
        conn.execute("ALTER TABLE finance ADD COLUMN IF NOT EXISTS employee_name TEXT DEFAULT NULL")
        conn.execute("RELEASE SAVEPOINT col_sp")
        conn.commit()
    except Exception:
        try: conn.execute("ROLLBACK TO SAVEPOINT col_sp")
        except: pass

    # Step 1: Backfill employee_name in finance table from note pattern
    try:
        rows = conn.execute("""
            SELECT id, note FROM finance
            WHERE tx_type='expense' AND LOWER(category)='salary'
            AND note LIKE 'Salary — %'
        """).fetchall()
        for r in rows:
            emp = r["note"].replace("Salary — ", "").strip()
            if emp:
                conn.execute(
                    "UPDATE finance SET employee_name=? WHERE id=? AND (employee_name IS NULL OR employee_name='')",
                    (emp, r["id"])
                )
        conn.commit()
    except Exception:
        pass

    # Step 2: For each finance salary expense, ensure it's also in salary_advances
    # (so the Salary page — which reads salary_advances — picks it up correctly)
    synced = 0
    try:
        fin_rows = conn.execute("""
            SELECT id, employee_name, amount, note, tx_date FROM finance
            WHERE tx_type='expense' AND LOWER(category)='salary'
            AND employee_name IS NOT NULL AND employee_name != ''
        """).fetchall()

        for r in fin_rows:
            emp    = r["employee_name"]
            amount = r["amount"]
            note   = r["note"] or f"Salary — {emp}"
            dt     = r["tx_date"]
            # Check if already in salary_advances (same employee, amount, date)
            exists = conn.execute("""
                SELECT id FROM salary_advances
                WHERE employee_name=? AND amount=? AND advance_date=?
            """, (emp, amount, dt)).fetchone()
            if not exists:
                now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                conn.execute(
                    "INSERT INTO salary_advances(employee_name,amount,note,advance_date,created_at) VALUES(?,?,?,?,?)",
                    (emp, amount, note, dt, now)
                )
                synced += 1
        conn.commit()
    except Exception as e:
        return f"<h2>❌ Sync error: {e}</h2>"

    conn.close()
    return (
        f"<h2>✅ Sync complete!</h2>"
        f"<p>{synced} new entries added to salary advances.</p>"
        f"<p><a href='/owner/salary'>← Go to Salary page</a></p>"
    )


@bp.route("/api/salary/fresh-start", methods=["POST"])
@owner_required
def api_salary_fresh_start():
    today = date.today().isoformat()
    set_setting("salary_fresh_start_date", today)
    return jsonify({"ok": True, "date": today})


@bp.route("/api/save-setting", methods=["POST"])
@owner_required
def api_save_setting():
    data  = request.get_json(silent=True) or {}
    key   = data.get("key","").strip()
    value = data.get("value","").strip()
    if not key:
        return jsonify({"ok":False,"error":"No key"})
    set_setting(key, value)
    return jsonify({"ok":True})


@bp.route("/api/employee-skill", methods=["POST"])
@owner_required
def api_employee_skill():
    data   = request.get_json(silent=True) or {}
    emp_id = data.get("employee_id")
    skills = data.get("skills","stitch")
    if not emp_id:
        return jsonify({"ok":False,"error":"No employee id"})
    conn = get_db()
    conn.execute("UPDATE employees SET skills=? WHERE id=?", (skills, emp_id))
    conn.commit(); conn.close()
    return jsonify({"ok":True})



# ══════════════════════════════════════════════
#  GARMENT MANAGER — Unified Garment Admin
# ══════════════════════════════════════════════

STANDARD_GARMENTS = [
    "Shirt","Shirt Linen","Pant","Pant Double","Jeans","Suit 2pc","Suit 3pc",
    "Blazer","Kurta","Kurta Pajama","Pajama","Pathani","Sherwani","Safari","Waistcoat",
    "Alteration","Cutting Only"
]

@bp.route("/garments")
@owner_required
def garment_manager():
    conn = get_db()
    today = date.today().isoformat()
    urgent_count = conn.execute(
        "SELECT COUNT(*) as c FROM orders WHERE is_urgent=1 AND status!='delivered' AND delivery_date>=?",
        (today,)
    ).fetchone()["c"]

    deleted_csv = get_setting("deleted_customer_rates", "")
    deleted_set = set(x.strip() for x in deleted_csv.split(",") if x.strip())

    all_settings_rows = conn.execute("SELECT key, value FROM settings").fetchall()
    sd = {r["key"]: r["value"] for r in all_settings_rows}

    # Collect ALL garment names (standard + custom via any key)
    garment_names = [g for g in STANDARD_GARMENTS if g not in deleted_set]
    for key in sd:
        for prefix in ("customer_rate_", "stitch_rate_", "types_"):
            if key.startswith(prefix):
                name = key[len(prefix):]
                if name not in garment_names and name not in deleted_set and name not in STANDARD_GARMENTS:
                    garment_names.append(name)

    garments = []
    for name in garment_names:
        cr = sd.get("customer_rate_"+name, "") or sd.get("rate_"+name, "0") or "0"
        sr = sd.get("stitch_rate_"+name, "0") or "0"
        chips_raw = sd.get("types_"+name, "")
        chips = []
        for t in chips_raw.split("|"):
            if ":" in t:
                k, v = t.split(":", 1)
                if k.strip() and v.strip():
                    chips.append({"k": k.strip(), "v": v.strip()})
        garments.append({
            "name": name,
            "is_standard": name in STANDARD_GARMENTS,
            "customer_rate": cr,
            "stitch_rate": sr,
            "chips": chips,
            "chips_raw": chips_raw,
        })

    # Also fetch measurement fields per garment so they can be managed inline
    mf_rows = conn.execute(
        "SELECT id, garment_type, field_name, sort_order FROM measurement_fields ORDER BY sort_order ASC, id ASC"
    ).fetchall()
    meas_by_garment = {}
    for r in mf_rows:
        meas_by_garment.setdefault(r["garment_type"], []).append({
            "id": r["id"], "field_name": r["field_name"], "sort_order": r["sort_order"]
        })

    conn.close()
    return render_template("owner/garment_manager.html",
        active_page="garments", show_voice=False, urgent_count=urgent_count,
        garments=garments,
        meas_by_garment=meas_by_garment,
    )


@bp.route("/api/garment/save", methods=["POST"])
@owner_required
def api_garment_save():
    data = request.get_json(silent=True) or {}
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"ok": False, "error": "No garment name"})
    customer_rate = str(data.get("customer_rate", "0") or "0").strip()
    stitch_rate   = str(data.get("stitch_rate", "0") or "0").strip()
    chips_raw     = (data.get("chips_raw") or "").strip()
    set_setting("customer_rate_"+name, customer_rate)
    set_setting("rate_"+name, customer_rate)
    set_setting("stitch_rate_"+name, stitch_rate)
    set_setting("types_"+name, chips_raw)
    # Un-delete if previously deleted
    deleted_csv = get_setting("deleted_customer_rates", "")
    deleted_set = set(x.strip() for x in deleted_csv.split(",") if x.strip())
    deleted_set.discard(name)
    set_setting("deleted_customer_rates", ",".join(sorted(deleted_set)))
    from database import invalidate_settings_cache; invalidate_settings_cache()
    return jsonify({"ok": True})


@bp.route("/api/garment/add", methods=["POST"])
@owner_required
def api_garment_add():
    data = request.get_json(silent=True) or {}
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"ok": False, "error": "No garment name"})
    customer_rate = str(data.get("customer_rate", "0") or "0").strip()
    stitch_rate   = str(data.get("stitch_rate", "0") or "0").strip()
    chips_raw     = (data.get("chips_raw") or "").strip()
    set_setting("customer_rate_"+name, customer_rate)
    set_setting("rate_"+name, customer_rate)
    set_setting("stitch_rate_"+name, stitch_rate)
    if chips_raw:
        set_setting("types_"+name, chips_raw)
    deleted_csv = get_setting("deleted_customer_rates", "")
    deleted_set = set(x.strip() for x in deleted_csv.split(",") if x.strip())
    deleted_set.discard(name)
    set_setting("deleted_customer_rates", ",".join(sorted(deleted_set)))
    from database import invalidate_settings_cache; invalidate_settings_cache()
    return jsonify({"ok": True, "name": name})


@bp.route("/api/garment/delete", methods=["POST"])
@owner_required
def api_garment_delete():
    data = request.get_json(silent=True) or {}
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"ok": False, "error": "No garment name"})
    conn = get_db()
    conn.execute("DELETE FROM settings WHERE key=?", ("customer_rate_"+name,))
    conn.execute("DELETE FROM settings WHERE key=?", ("rate_"+name,))
    conn.commit(); conn.close()
    deleted_csv = get_setting("deleted_customer_rates", "")
    deleted_set = set(x.strip() for x in deleted_csv.split(",") if x.strip())
    deleted_set.add(name)
    set_setting("deleted_customer_rates", ",".join(sorted(deleted_set)))
    from database import invalidate_settings_cache; invalidate_settings_cache()
    return jsonify({"ok": True})


# ── Measurement fields — inline management from Garment Manager ──

@bp.route("/api/garment/meas-fields/add", methods=["POST"])
@owner_required
def api_meas_field_add():
    data = request.get_json(silent=True) or {}
    garment = (data.get("garment_type") or "").strip()
    field   = (data.get("field_name") or "").strip()
    if not garment or not field:
        return jsonify({"ok": False, "error": "Missing garment or field"})
    conn = get_db()
    # Check not duplicate
    exists = conn.execute(
        "SELECT id FROM measurement_fields WHERE garment_type=? AND field_name=?", (garment, field)
    ).fetchone()
    if exists:
        conn.close()
        return jsonify({"ok": False, "error": "Field already exists"})
    # Get next sort_order for this garment type (so new fields go to end in order)
    max_order = conn.execute(
        "SELECT COALESCE(MAX(sort_order), 0) as m FROM measurement_fields WHERE garment_type=?", (garment,)
    ).fetchone()["m"]
    conn.execute(
        "INSERT INTO measurement_fields(garment_type, field_name, sort_order) VALUES(?,?,?)",
        (garment, field, max_order + 1)
    )
    conn.commit()
    new_id = conn.execute(
        "SELECT id FROM measurement_fields WHERE garment_type=? AND field_name=? ORDER BY id DESC LIMIT 1",
        (garment, field)
    ).fetchone()["id"]
    conn.close()
    return jsonify({"ok": True, "id": new_id})


@bp.route("/api/garment/meas-fields/delete/<int:fid>", methods=["POST"])
@owner_required
def api_meas_field_delete(fid):
    conn = get_db()
    conn.execute("DELETE FROM measurement_fields WHERE id=?", (fid,))
    conn.commit(); conn.close()
    return jsonify({"ok": True})


@bp.route("/api/garment/meas-fields/reorder", methods=["POST"])
@owner_required
def api_meas_field_reorder():
    data = request.get_json(silent=True) or {}
    ordered_ids = data.get("ids", [])
    conn = get_db()
    for pos, fid in enumerate(ordered_ids):
        conn.execute("UPDATE measurement_fields SET sort_order=? WHERE id=?", (pos, fid))
    conn.commit(); conn.close()
    return jsonify({"ok": True})


# ══════════════════════════════════════════════
#  PAST ORDERS (OLD DATA ENTRY)
# ══════════════════════════════════════════════

@bp.route("/past-upload")
def past_upload():
    """Mobile-optimized page for taking diary photos linked to past order codes."""
    return render_template("owner/past_upload.html")


@bp.route("/past-orders")
def past_orders():
    """Page to enter old/historical delivered orders — mirrors new_order flow."""
    import json as _json
    conn = get_db()

    # Garment rates — same logic as new_order
    RATE_DEFAULTS = {
        "Shirt":"350","Shirt Linen":"450","Pant":"450","Pant Double":"550",
        "Jeans":"550","Suit 2pc":"2800","Suit 3pc":"3500","Blazer":"2300",
        "Kurta":"800","Kurta Pajama":"1000","Pajama":"300","Pathani":"800",
        "Sherwani":"3500","Safari":"1500","Waistcoat":"800",
        "Alteration":"100","Cutting Only":"100"
    }
    deleted_csv = get_setting("deleted_customer_rates", "")
    deleted_set = set(x.strip() for x in deleted_csv.split(",") if x.strip())
    garment_rates = {}
    for n, default_rate in RATE_DEFAULTS.items():
        if n in deleted_set:
            continue
        cr = get_setting("customer_rate_"+n, "")
        if cr and cr.strip():
            garment_rates[n] = cr
        else:
            garment_rates[n] = get_setting("rate_"+n, default_rate)
    custom_rates = conn.execute("SELECT key,value FROM settings WHERE key LIKE 'customer_rate_%'").fetchall()
    for row in custom_rates:
        name = row["key"][14:]
        if name not in garment_rates and name not in deleted_set and row["value"] and row["value"].strip() and row["value"] != "0":
            garment_rates[name] = row["value"]

    # Measurement fields
    meas_rows = conn.execute("SELECT garment_type, field_name FROM measurement_fields ORDER BY garment_type, sort_order").fetchall()
    meas_fields = {}
    for r in meas_rows:
        meas_fields.setdefault(r["garment_type"], []).append(r["field_name"])

    # Garment type chips
    garment_types = {}
    for row in conn.execute("SELECT key, value FROM settings WHERE key LIKE 'types_%'").fetchall():
        gname = row["key"][6:]
        pairs = []
        for t in (row["value"] or "").split("|"):
            if ":" in t:
                k, v = t.split(":", 1)
                pairs.append({"k": k.strip(), "v": v.strip()})
        if pairs:
            garment_types[gname] = pairs

    # Peek next order code for display
    from database import peek_order_code
    # Past orders: user enters their OWN code — don't prefill with next new-order code
    next_code = ""

    conn.close()
    return render_template("owner/past_orders.html",
        active_page="past_orders",
        garment_rates=garment_rates,
        garment_rates_json=_json.dumps({k: float(v) for k, v in garment_rates.items()}),
        meas_fields_json=_json.dumps(meas_fields),
        garment_types_json=_json.dumps(garment_types),
        next_code=next_code,
    )

@bp.route("/past-orders/save", methods=["POST"])
def past_orders_save():
    """Save a past/historical order — bulletproof version."""
    import json as _json
    data = request.get_json(silent=True) or {}

    customer_name  = (data.get("customer_name") or "").strip()
    customer_mobile= (data.get("mobile") or "").strip()
    customer_address = (data.get("address") or "").strip()
    order_date     = (data.get("order_date") or "").strip()
    delivery_date  = (data.get("delivery_date") or "").strip()
    total_amount   = float(data.get("total_amount") or 0)
    advance_paid   = float(data.get("advance_paid") or 0)
    payment_mode   = (data.get("payment_mode") or "cash").strip()
    note           = (data.get("note") or "").strip()
    garments       = data.get("garments") or []
    order_code_override = (data.get("order_code") or "").strip()
    is_delivered   = data.get("is_delivered", True)

    if not customer_name:
        return jsonify({"ok": False, "error": "Customer name required"})

    payable   = total_amount
    remaining = max(0, payable - advance_paid)
    now_str   = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    status    = "delivered" if is_delivered else "pending"

    try:
        conn = get_db()

        # 1. Customer — no automatic matching by mobile or name. Every past
        # order gets its own separate, brand-new customer record; nothing is
        # ever silently merged or reused based on a guess.
        row = conn.execute(
            "INSERT INTO customers(name,mobile,address,created_at) VALUES(?,?,?,?) RETURNING id",
            (customer_name, customer_mobile or None, customer_address, now_str)).fetchone()
        customer_id = row["id"] if row else None

        if not customer_id:
            conn.close()
            return jsonify({"ok": False, "error": "Customer create failed"})


        # 2. Order code — ONLY use the user-provided code, NEVER auto-generate
        # Past orders must NEVER touch the last_order_code counter — that is
        # exclusively for new orders created through the new-order flow.
        order_code = order_code_override or ""
        if not order_code:
            conn.close()
            return jsonify({"ok": False, "error": "Order code required for past orders"})

        # Check for clash — if already exists, return error (do NOT generate a new code)
        clash = conn.execute("SELECT id FROM orders WHERE order_code=?", (order_code,)).fetchone()
        if clash:
            conn.close()
            return jsonify({"ok": False, "error": f"Order #{order_code} already exists. Please use a different code."})

        # 3. Insert order
        delivered_at = delivery_date if is_delivered else None
        conn.execute("""
            INSERT INTO orders(order_code,customer_id,order_date,delivery_date,
                total_amount,extra_charges,payable_amount,advance_paid,remaining,
                payment_mode,status,is_urgent,note,delivered_at,created_at)
            VALUES(?,?,?,?,?,0,?,?,?,?,?,0,?,?,?)
        """, (order_code, customer_id, order_date, delivery_date,
              total_amount, payable, advance_paid, remaining,
              payment_mode, status, note, delivered_at, now_str))
        conn.commit()

        r = conn.execute("SELECT id FROM orders WHERE order_code=?", (order_code,)).fetchone()
        order_id = r["id"] if r else None
        if not order_id:
            conn.close()
            return jsonify({"ok": False, "error": "Order insert failed"})

        # 4. Items
        for g in garments:
            gtype = (g.get("type") or "").strip()
            qty   = int(g.get("qty") or 1)
            rate  = float(g.get("rate") or 0)
            if gtype:
                meas  = g.get("measurements") or {}
                notes = (g.get("notes") or "").strip()
                sel   = g.get("selectedTypes") or []
                if sel:
                    notes = (notes.split("[")[0].strip() + " [" + ",".join(sel) + "]").strip()
                conn.execute("""
                    INSERT INTO order_items(order_id,garment_type,quantity,rate,amount,measurements,notes)
                    VALUES(?,?,?,?,?,?,?)
                """, (order_id, gtype, qty, rate, qty*rate, _json.dumps(meas), notes))

        # NOTE: Past orders are historical records only — the money for
        # these orders was already received outside/before UTMS, so we do
        # NOT log a finance entry here. Doing so would double-count income
        # in Finance reports (showing old money as if received today).
        # The order's own advance_paid/remaining fields still get set
        # correctly above, for accurate order-level payment reference.

        conn.commit()
        conn.close()
        from database import invalidate_settings_cache
        invalidate_settings_cache()
        return jsonify({"ok": True, "order_code": order_code})

    except Exception as e:
        try: conn.close()
        except: pass
        import traceback
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(e)})


# ══════════════════════════════════════════════
#  AUTO-BACKUP: FULL EXCEL EXPORT (ALL TABLES)
# ══════════════════════════════════════════════

@bp.route("/backup/excel")
@owner_required
def backup_excel():
    """Full Excel backup with all tables as separate sheets."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    import json as _json

    conn = get_db()
    wb   = openpyxl.Workbook()

    hdr_fill = PatternFill("solid", fgColor="6366F1")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)

    def write_sheet(ws, rows):
        if not rows:
            ws.append(["No data"])
            return
        keys = list(rows[0].keys())
        # Header row
        for col, k in enumerate(keys, 1):
            cell = ws.cell(row=1, column=col, value=k.upper())
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"
        for r_idx, row in enumerate(rows, 2):
            for col, k in enumerate(keys, 1):
                ws.cell(row=r_idx, column=col, value=str(row[k] or "") if row[k] is not None else "")
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 45)

    # Sheet 1: Orders (with customer name & garments)
    ws1 = wb.active
    ws1.title = "Orders"
    orders = conn.execute("""
        SELECT o.order_code, c.name as customer, c.mobile, o.order_date, o.delivery_date,
               o.status, o.total_amount, o.advance_paid, o.remaining, o.payment_mode,
               o.note, o.delivered_at, o.created_at
        FROM orders o LEFT JOIN customers c ON c.id=o.customer_id
        ORDER BY o.id DESC
    """).fetchall()
    write_sheet(ws1, orders)

    # Sheet 2: Order Items
    ws2 = wb.create_sheet("Order Items")
    items = conn.execute("""
        SELECT oi.id, o.order_code, c.name as customer, oi.garment_type,
               oi.quantity, oi.rate, oi.amount, oi.measurements, oi.notes
        FROM order_items oi
        JOIN orders o ON o.id=oi.order_id
        LEFT JOIN customers c ON c.id=o.customer_id
        ORDER BY oi.id DESC
    """).fetchall()
    write_sheet(ws2, items)

    # Sheet 3: Customers
    ws3 = wb.create_sheet("Customers")
    customers = conn.execute("SELECT * FROM customers ORDER BY id DESC").fetchall()
    write_sheet(ws3, customers)

    # Sheet 4: Finance
    ws4 = wb.create_sheet("Finance")
    finance = conn.execute("""
        SELECT f.id, f.tx_date, f.tx_type, f.category, f.amount, f.mode,
               o.order_code, f.note, f.created_by, f.created_at
        FROM finance f LEFT JOIN orders o ON o.id=f.order_id
        ORDER BY f.id DESC
    """).fetchall()
    write_sheet(ws4, finance)

    # Sheet 5: Work Logs
    ws5 = wb.create_sheet("Work Logs")
    wlogs = conn.execute("SELECT * FROM work_logs ORDER BY id DESC").fetchall()
    write_sheet(ws5, wlogs)

    # Sheet 6: Employees
    ws6 = wb.create_sheet("Employees")
    emps = conn.execute("SELECT * FROM employees ORDER BY id").fetchall()
    write_sheet(ws6, emps)

    conn.close()

    # Update last_backup_at setting
    set_setting("last_backup_at", datetime.now().strftime("%Y-%m-%d %H:%M"))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    from flask import send_file
    fname = f"uttam_tailors_backup_{datetime.now().strftime('%d-%m-%Y_%H%M')}.xlsx"
    return send_file(buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name=fname)


# ══════════════════════════════════════════════
#  EDIT ORDER (full edit — all fields + images)
# ══════════════════════════════════════════════

@bp.route("/orders/edit/<order_code>")
def order_edit(order_code):
    import json as _json, os as _os
    from config import Config
    conn = get_db()

    o = conn.execute("""
        SELECT o.*, c.name as cname, c.mobile as cmobile, c.address as caddress, c.id as cid
        FROM orders o LEFT JOIN customers c ON c.id=o.customer_id
        WHERE o.order_code=?
    """, (order_code,)).fetchone()
    if not o:
        conn.close()
        return "<h2>Order not found</h2>", 404

    items = conn.execute("SELECT * FROM order_items WHERE order_id=?", (o["id"],)).fetchall()
    garments = []
    for it in items:
        try: meas = _json.loads(it["measurements"] or "{}")
        except: meas = {}
        garments.append({"type":it["garment_type"],"qty":it["quantity"],
                         "rate":it["rate"],"meas":meas,"notes":it["notes"] or ""})

    # Images
    images = []
    rows = conn.execute("SELECT file_path FROM order_images WHERE order_id=? ORDER BY id", (o["id"],)).fetchall()
    images = [r["file_path"] for r in rows if r["file_path"] and not r["file_path"].startswith("temp:")]
    if not images:
        folder = _os.path.join(Config.UPLOAD_FOLDER, order_code)
        if _os.path.isdir(folder):
            images = [f"/static/order_images/{order_code}/{f}"
                      for f in sorted(_os.listdir(folder))
                      if f.lower().endswith((".jpg",".jpeg",".png",".gif",".webp"))]

    # Garment rates + measurement fields
    garment_names = ["Shirt","Shirt Linen","Pant","Pant Double","Jeans","Suit 2pc","Suit 3pc",
        "Blazer","Kurta","Kurta Pajama","Pajama","Pathani","Sherwani","Safari","Waistcoat",
        "Alteration","Cutting Only"]
    garment_rates = {n: get_setting("customer_rate_"+n,"") or get_setting("rate_"+n,"0") for n in garment_names}
    custom = conn.execute("SELECT key,value FROM settings WHERE key LIKE 'customer_rate_%'").fetchall()
    for row in custom:
        name = row["key"][14:]
        if name not in garment_rates: garment_rates[name] = row["value"]

    meas_fields = {}
    for r in conn.execute("SELECT garment_type, field_name FROM measurement_fields ORDER BY garment_type, sort_order").fetchall():
        meas_fields.setdefault(r["garment_type"], []).append(r["field_name"])

    garment_types = {}
    for row in conn.execute("SELECT key, value FROM settings WHERE key LIKE 'types_%'").fetchall():
        gname = row["key"][6:]
        pairs = []
        for t in (row["value"] or "").split("|"):
            if ":" in t:
                k, v = t.split(":", 1)
                pairs.append({"k": k.strip(), "v": v.strip()})
        if pairs: garment_types[gname] = pairs

    urgent_count = conn.execute("SELECT COUNT(*) as c FROM orders WHERE is_urgent=1 AND status!='delivered'").fetchone()["c"]
    conn.close()

    # Existing images for this order
    import os as _os
    img_dir = _os.path.join(_os.path.dirname(_os.path.dirname(_os.path.dirname(__file__))),
                            "static", "order_images", order_code)
    existing_images = []
    if _os.path.exists(img_dir):
        for f in sorted(_os.listdir(img_dir)):
            if f.lower().endswith(('.jpg','.jpeg','.png','.webp','.gif')):
                existing_images.append(f"/static/order_images/{order_code}/{f}")

    return render_template("owner/order_edit.html",
        active_page="owner_orders", show_voice=False,
        urgent_count=urgent_count,
        order=dict(o),
        existing_images=existing_images,
        garments_json=_json.dumps(garments),
        images=images,
        garment_rates=garment_rates,
        garment_rates_json=_json.dumps({k: float(v) for k, v in garment_rates.items()}),
        meas_fields_json=_json.dumps(meas_fields),
        garment_types_json=_json.dumps(garment_types),
    )


@bp.route("/orders/edit/<order_code>/save", methods=["POST"])
def order_edit_save(order_code):
    import json as _json
    data = request.get_json(silent=True) or {}
    conn = get_db()
    try:
        o = conn.execute("SELECT id, customer_id, advance_paid FROM orders WHERE order_code=?", (order_code,)).fetchone()
        if not o:
            return jsonify({"ok": False, "error": "Order not found"})
        order_id    = o["id"]
        customer_id = o["customer_id"]
        old_advance = float(o["advance_paid"] or 0)

        # Update customer — NEVER overwrite mobile if it changed (different person!)
        name    = (data.get("customer_name") or "").strip()
        mobile  = (data.get("mobile") or "").strip()
        address = (data.get("address") or "").strip()
        if name:
            # Get existing customer's mobile
            ex_cust = conn.execute("SELECT mobile, name FROM customers WHERE id=?",
                                   (customer_id,)).fetchone()
            ex_mobile = ((ex_cust["mobile"] if ex_cust else "") or "").strip()

            if mobile and ex_mobile and mobile != ex_mobile:
                # Mobile changed on this order — no automatic matching against
                # other customers; always create a fresh, separate customer
                # record rather than guessing this is someone else's existing entry.
                from datetime import datetime as _dt
                now_str = _dt.now().strftime("%Y-%m-%d %H:%M:%S")
                row = conn.execute(
                    "INSERT INTO customers(name,mobile,address,created_at) VALUES(?,?,?,?) RETURNING id",
                    (name, mobile, address, now_str)).fetchone()
                new_cust_id = row["id"]
                conn.execute("UPDATE orders SET customer_id=? WHERE id=?",
                             (new_cust_id, order_id))
                customer_id = new_cust_id
            elif mobile and not ex_mobile:
                # Mobile was NULL/empty, now ADDING one → just update this customer
                conn.execute("UPDATE customers SET name=?,mobile=?,address=? WHERE id=?",
                             (name, mobile, address, customer_id))
            elif not mobile and ex_mobile:
                # Mobile explicitly CLEARED → set to NULL
                conn.execute("UPDATE customers SET name=?,mobile=NULL,address=? WHERE id=?",
                             (name, address, customer_id))
            else:
                # Same mobile or both blank → update name/address only
                conn.execute("UPDATE customers SET name=?,address=? WHERE id=?",
                             (name, address, customer_id))

        # Update order
        new_payable = float(data.get("payable_amount",0))
        new_advance = float(data.get("advance_paid",0))
        # Safety net: never allow advance to exceed the order's payable
        # amount, even if frontend validation was somehow bypassed.
        if new_payable > 0 and new_advance > new_payable:
            new_advance = new_payable
        new_remaining = max(0, round(new_payable - new_advance, 2))
        conn.execute("""
            UPDATE orders SET
                order_date=?, delivery_date=?, note=?, is_urgent=?,
                total_amount=?, extra_charges=?, payable_amount=?,
                advance_paid=?, remaining=?, payment_mode=?, status=?
            WHERE id=?
        """, (
            data.get("order_date",""),
            data.get("delivery_date",""),
            data.get("note",""),
            1 if data.get("is_urgent") else 0,
            float(data.get("total_amount",0)),
            float(data.get("extra_charges",0)),
            new_payable,
            new_advance,
            new_remaining,
            data.get("payment_mode","cash"),
            data.get("status","pending"),
            order_id
        ))

        # ── Keep Finance ledger in sync ──────────────────────────────
        # If the owner increased the advance/paid amount directly on
        # this edit screen (instead of going through the Finance page),
        # that money was never logged anywhere — meaning Finance reports
        # would be missing this income entirely. Auto-log the difference
        # so the two stay connected. (Decreases are NOT auto-logged here
        # since those are typically data-entry corrections, not refunds.)
        # If payment_mode changed → update ALL existing finance entries for this order
        new_mode = data.get("payment_mode", "cash")
        if new_mode != (data.get("old_payment_mode") or ""):
            conn.execute(
                "UPDATE finance SET mode=? WHERE order_id=? AND tx_type='income'",
                (new_mode, order_id)
            )

        payment_diff = round(new_advance - old_advance, 2)
        if payment_diff > 0:
            now0 = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            today0 = date.today().isoformat()
            conn.execute("""
                INSERT INTO finance(tx_date,tx_type,category,amount,mode,order_id,note,created_by,created_at)
                VALUES(?,'income','payment',?,?,?,?,'owner',?)
            """, (today0, payment_diff, data.get("payment_mode","cash"), order_id,
                  f"Payment updated via Order Edit for #{order_code}", now0))

        # Replace garment items
        conn.execute("DELETE FROM order_items WHERE order_id=?", (order_id,))
        for g in (data.get("garments") or []):
            gtype = (g.get("type") or "").strip()
            qty   = int(g.get("qty") or 1)
            rate  = float(g.get("rate") or 0)
            meas  = g.get("measurements") or {}
            notes = (g.get("notes") or "").strip()
            if gtype:
                conn.execute("""
                    INSERT INTO order_items(order_id,garment_type,quantity,rate,amount,measurements,notes)
                    VALUES(?,?,?,?,?,?,?)
                """, (order_id, gtype, qty, rate, qty*rate, _json.dumps(meas), notes))

        conn.commit()
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        try: conn.close()
        except: pass
        return jsonify({"ok": False, "error": str(e)})
# deploy-2026-07-23 08:11:46
